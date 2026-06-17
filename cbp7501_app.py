#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CBP Form 7501 - Drag & Drop Desktop-App (v4)
=============================================
Ziehen Sie eine oder mehrere PDF-Dateien per Drag & Drop in das Fenster.
Jede PDF wird sofort gelesen, die CBP Form 7501 (Entry Summary) darin
gesucht und alle Felder in die SQLite-Datenbank geschrieben.

Das Fenster hat zwei Reiter:

  1) "Import & Status"
     - Ablage-Zone fuer Drag & Drop
     - Liste: hochgeladene PDF-Dateien OHNE erkannte CBP-Seite (Form 7501)
       direkt unter der Ablage-Zone, damit Problemfaelle sofort sichtbar sind
     - Liste aller bisher verarbeiteten Dateien:
       Datei-Name, Gelesen am (Datum/Uhrzeit), Groesse, Status, Details
     - Knopf "Excel-Export (XLSX) erzeugen und oeffnen": exportiert den
       kompletten Datenbank-Inhalt als Excel-Arbeitsmappe und oeffnet sie

  2) "Datenbank-Inhalt"
     - Tabellarische Ansicht ALLER Felder aus der Datenbank:
       - Tabelle "entries" (Kopfdaten, Felder 1-26 + 35-43)
       - Tabelle "entry_lines" (Warenpositionen, Felder 27-34)
     - Aktualisiert sich automatisch nach jedem Import
       (oder per Knopfdruck "Aktualisieren")

Voraussetzungen (einmalig in der Eingabeaufforderung installieren):
    pip install pdfplumber tkinterdnd2

Start:
    python cbp7501_app.py

Wichtig: Diese Datei muss im selben Ordner liegen wie
'cbp7501_extractor.py' (wird als Bibliothek wiederverwendet).
"""

import os
import sys
import sqlite3
import threading
import html
import webbrowser
from datetime import datetime


import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    print("Das Paket 'tkinterdnd2' fehlt. Bitte installieren mit:")
    print("    pip install tkinterdnd2")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── PyInstaller-kompatibler App-Pfad ─────────────────────────────────────────
# Im gefrorenen EXE zeigt __file__ in den _internal/-Ordner.
# sys.executable zeigt dagegen immer auf die EXE selbst → Ordner daneben = APP_DIR.
if getattr(sys, "frozen", False):
    # Läuft als PyInstaller-EXE
    APP_DIR = os.path.dirname(sys.executable)
    # _internal/ dem Import-Pfad hinzufügen (cbp7501_extractor.py liegt dort)
    _internal = os.path.join(APP_DIR, "_internal")
    if os.path.isdir(_internal):
        sys.path.insert(0, _internal)
    sys.path.insert(0, APP_DIR)
else:
    # Normaler Python-Start
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, APP_DIR)
    # Sicherstellen dass immer der aktuelle Quellcode geladen wird
    import shutil as _shutil
    _pycache = os.path.join(APP_DIR, "__pycache__")
    if os.path.isdir(_pycache):
        _shutil.rmtree(_pycache, ignore_errors=True)
    del _shutil, _pycache

try:
    import cbp7501_extractor as extractor
except ImportError:
    print("Datei 'cbp7501_extractor.py' wurde nicht im selben Ordner gefunden.")
    sys.exit(1)
DB_PATH = os.path.join(APP_DIR, "cbp7501.db")
REPORT_PATH = os.path.join(APP_DIR, "CBP7501_Gesamtreport.html")
STARTER_BAT = os.path.join(APP_DIR, "CBP7501-Import-starten.bat")
EXPORT_XLSX_PATH = os.path.join(APP_DIR, "CBP7501_Datenbank_Export.xlsx")
CLAIM_TEMPLATE_PATH = os.path.join(APP_DIR, "template_Claim-Report_Kunde.xlsx")
PDF_ARCHIVE_DIR = os.path.join(APP_DIR, "pdf_archive")
os.makedirs(PDF_ARCHIVE_DIR, exist_ok=True)

LOG_SCHEMA = """
CREATE TABLE IF NOT EXISTS processed_files (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    file_name       TEXT NOT NULL,
    file_path       TEXT,
    file_size_bytes INTEGER,
    processed_at    TEXT NOT NULL,
    status          TEXT,
    message         TEXT,
    archived_path   TEXT
);
"""

def _migrate_ck_row_keys(conn):
    """Normalisiert alte ck_row_selection-Schluessel: 'None' → '' (einmalige Migration)."""
    try:
        rows = conn.execute(
            "SELECT row_key FROM ck_row_selection WHERE row_key LIKE '%None%'").fetchall()
        for (old_key,) in rows:
            new_key = "|".join(
                "" if p == "None" else p for p in old_key.split("|"))
            if new_key != old_key:
                conn.execute(
                    "INSERT OR REPLACE INTO ck_row_selection (row_key, selected) "
                    "SELECT ?, selected FROM ck_row_selection WHERE row_key=?",
                    (new_key, old_key))
                conn.execute(
                    "DELETE FROM ck_row_selection WHERE row_key=?", (old_key,))
        if rows:
            conn.commit()
    except Exception:
        pass


def _migrate_processed_files(conn):
    """Fügt archived_path-Spalte zu bestehenden DBs hinzu (einmalige Migration)."""
    cols = [r[1] for r in conn.execute("PRAGMA table_info(processed_files)").fetchall()]
    if "archived_path" not in cols:
        conn.execute("ALTER TABLE processed_files ADD COLUMN archived_path TEXT")
        conn.commit()

ENTRIES_LABELS = {
    "id": "ID", "source_file": "Quelldatei", "source_page": "Seite",
    "imported_at": "Importiert am",
    "filer_code_entry_no": "1. Filer Code/Entry No.", "entry_type": "2. Entry Type",
    "summary_date": "3. Summary Date", "surety_no": "4. Surety No.",
    "bond_type": "5. Bond Type", "port_code": "6. Port Code",
    "entry_date": "7. Entry Date", "importing_carrier": "8. Importing Carrier",
    "mode_of_transport": "9. Mode of Transport", "country_of_origin": "10. Country of Origin",
    "import_date": "11. Import Date", "bl_or_awb_no": "12. B/L or AWB No.",
    "manufacturer_id": "13. Manufacturer ID", "exporting_country": "14. Exporting Country",
    "export_date": "15. Export Date", "it_no": "16. I.T. No.", "it_date": "17. I.T. Date",
    "missing_docs": "18. Missing Docs", "foreign_port_of_lading": "19. Foreign Port of Lading",
    "us_port_of_unlading": "20. U.S. Port of Unlading",
    "location_of_goods_go_no": "21. Location of Goods/G.O.No.",
    "consignee_no": "22. Consignee No.", "importer_no": "23. Importer No.",
    "reference_no": "24. Reference No.",
    "ultimate_consignee_name_address": "25. Ultimate Consignee",
    "importer_of_record_name_address": "26. Importer of Record",
    "total_entered_value": "35. Total Entered Value", "mpf_total": "MPF Summe",
    "duty_total": "37. Duty", "tax_total": "38. Tax", "other_total": "39. Other",
    "grand_total": "40. Grand Total",
    "declarant_name_title_signature_date": "41. Declarant",
    "broker_filer_information": "42. Broker/Filer Info",
    "broker_importer_file_number": "43. Broker File No.",
}

LINES_LABELS = {
    "id": "ID", "entry_id": "Entry-ID", "source_block": "Quelle", "line_no": "Pos.-Nr.",
    "country_of_origin_line": "Ursprungsland", "program_code": "Programmcode",
    "description": "Warenbeschreibung", "htsus_no": "HTSUS-Nr.",
    "gross_weight": "Bruttogewicht", "net_quantity": "Nettomenge",
    "htsus_rate": "Zollsatz", "duty_amount": "Zollbetrag", "entered_value": "Eingetr. Wert",
    "manifest_qty": "Manifest-Menge", "relationship": "Relationship", "visa_no": "Visa-Nr.",
    "mpf_rate": "MPF-Satz", "mpf_amount": "MPF-Betrag", "invoice_no": "Rechnungs-Nr.",
    "invoice_reference": "Rechnungsreferenz", "invoice_qty": "Rechnungsmenge",
    "invoice_value_qty": "Rechnungswert-Menge", "invoice_value_amount": "Rechnungswert-Betrag",
    "invoice_value_rate": "Rechnungswert-Kurs", "invoice_value_currency": "Waehrung",
    "filer_code_entry_no": "Entry No.",
}

# Auswertung-Tab – JOIN entry_lines + entries + stammdaten_db_tariff + zmk (59 Spalten)
AUS_COLUMNS = [
    # ── CBP Entry-Kopfdaten ──────────────────────────────────────────────────
    "filer_code_entry_no", "entry_type", "import_date", "entry_date",
    "country_of_origin", "exporting_country", "importing_carrier",
    "mode_of_transport", "us_port_of_unlading", "source_file",
    "total_entered_value", "duty_total", "mpf_total", "grand_total",
    # ── CBP entry_lines ──────────────────────────────────────────────────────
    "line_no", "htsus_no", "description", "country_of_origin_line",
    "program_code", "gross_weight", "net_quantity", "manifest_qty",
    "entered_value", "htsus_rate", "duty_amount",
    "mpf_rate", "mpf_amount", "hmf_rate", "hmf_amount",
    "invoice_no", "invoice_qty", "invoice_value_amount",
    "invoice_value_rate", "invoice_value_currency",
    "relationship", "visa_no", "source_block",
    # ── Lieferanten-Daten (db_Tariff JOIN) ──────────────────────────────────
    "dbt_dachser_dokument", "dbt_entry_document",
    "dbt_vendor", "dbt_vendor_name",
    "dbt_vendor_origin_o", "dbt_vendor_origin_e",
    "dbt_invoice_no_delivery_note", "dbt_po_order_no",
    "dbt_mat_no_sap", "dbt_material_cat_sap",
    "dbt_quantity", "dbt_rate",
    "dbt_invoice_doc_value", "dbt_doc_currency",
    "dbt_local_value_sap", "dbt_local_currency",
    # ── ZuOrd-Mat-KD (Kundenzuordnung über Mat-Nr.) ──────────────────────────
    "kd_sap_customer_no", "kd_sap_customer_name", "kd_customer_part_no",
    # ── CBP-Reports (Claim-Felder) ───────────────────────────────────────────
    "dbt_claim_status", "dbt_claim_value_1", "dbt_claim_value_2",
]
AUS_HEADERS = [
    # ── CBP Entry-Kopfdaten ──────────────────────────────────────────────────
    "Entry-No.", "Typ", "Import Date", "Entry Date",
    "Land Urspr.", "Exportland", "Carrier",
    "Transport", "US-Port", "Quelldatei",
    "Total Entered Value", "Duty Total", "MPF Total", "Grand Total",
    # ── CBP entry_lines ──────────────────────────────────────────────────────
    "Pos.", "HTSUS-Nr.", "Warenbeschreibung", "Ursprungsland (Pos.)",
    "Programm", "Bruttogewicht", "Nettomenge", "Manifest Menge",
    "Eingetr. Wert", "HTSUS-Satz", "Zollbetrag",
    "MPF-Satz", "MPF-Betrag", "HMF-Satz", "HMF-Betrag",
    "Rechnungs-Nr.", "Rg-Menge", "Rg-Wert-Betrag",
    "Rg-Wert-Kurs", "Währung",
    "Relationship", "Visa-Nr.", "Quellblock",
    # ── Lieferanten-Daten ────────────────────────────────────────────────────
    "Dachser Dokument", "Entry Document",
    "Vendor", "Vendor Name",
    "Vendor Origin (O)", "Vendor Origin (E)",
    "Invoice-No/Del.Note", "PO/Order-No.",
    "Mat-No. SAP", "Material CAT-SAP",
    "Quantity (DACHSER)", "Rate",
    "Invoice Doc Value", "Doc. Currency",
    "Local Value SAP", "Local Currency",
    # ── ZuOrd-Mat-KD ─────────────────────────────────────────────────────────
    "SAP Kd-Nr.", "SAP Kd-Name", "Kd. Teilenr.",
    # ── CBP-Reports (Claim) ───────────────────────────────────────────────────
    "Claim Status", "Claim Value 1", "Claim Value 2",
]
_AUS_NUM_COLS = {
    "total_entered_value", "duty_total", "mpf_total", "grand_total",
    "entered_value", "duty_amount", "mpf_amount", "hmf_amount",
    "invoice_value_amount",
    "dbt_quantity", "dbt_rate", "dbt_invoice_doc_value", "dbt_local_value_sap",
    "dbt_claim_value_1", "dbt_claim_value_2",
}
_AUS_COL_WIDTHS = {
    "filer_code_entry_no": 150, "description": 260, "source_file": 200,
    "htsus_no": 130, "import_date": 100, "entry_date": 100,
    "total_entered_value": 120, "duty_total": 110, "mpf_total": 90,
    "grand_total": 110, "entered_value": 100, "duty_amount": 100,
    "mpf_amount": 85, "hmf_amount": 85, "invoice_value_amount": 110,
    "importing_carrier": 120, "country_of_origin_line": 90,
    "invoice_no": 130, "us_port_of_unlading": 80,
    # Vendor
    "dbt_dachser_dokument": 120, "dbt_entry_document": 130,
    "dbt_vendor_name": 180, "dbt_vendor": 80,
    "dbt_invoice_no_delivery_note": 140, "dbt_po_order_no": 120,
    "dbt_mat_no_sap": 140, "dbt_material_cat_sap": 200,
    "dbt_invoice_doc_value": 110, "dbt_local_value_sap": 110,
    "dbt_quantity": 90, "dbt_rate": 80,
    # ZuOrd-Mat-KD
    "kd_sap_customer_no": 90, "kd_sap_customer_name": 210, "kd_customer_part_no": 130,
    # Claim
    "dbt_claim_status": 100, "dbt_claim_value_1": 110, "dbt_claim_value_2": 110,
}


def human_size(n):
    if n is None:
        return ""
    n = float(n)
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} GB"


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.executescript(extractor.SCHEMA)
    conn.executescript(LOG_SCHEMA)
    conn.executescript(TARIFF_CLAIM_SCHEMA)
    conn.executescript(CLAIM_SETTINGS_SCHEMA)
    conn.executescript(CLAIM_SUPPLEMENT_SCHEMA)
    conn.executescript(CK_SELECTION_SCHEMA)
    conn.executescript(ZUORD_SCHEMA)
    conn.executescript(DB_TARIFF_SCHEMA)
    conn.executescript(PDF_VENDOR_SCHEMA)
    conn.executescript(HTSUS_CLAIM_SCHEMA)
    _migrate_ck_row_keys(conn)
    _migrate_processed_files(conn)
    return conn


def log_processed_file(conn, file_name, file_path, file_size, status, message,
                        archived_path=None):
    conn.execute(
        "INSERT INTO processed_files "
        "(file_name, file_path, file_size_bytes, processed_at, status, message, archived_path) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (file_name, file_path, file_size,
         datetime.now().isoformat(timespec="seconds"),
         status, message, archived_path),
    )
    conn.commit()


def _archive_pdf(src_path):
    """Kopiert die PDF in den Archiv-Ordner. Gibt den Zielpfad zurück.
    Bei Namenskollision wird ein Timestamp-Präfix verwendet."""
    import shutil as _sh
    fname = os.path.basename(src_path)
    dest  = os.path.join(PDF_ARCHIVE_DIR, fname)
    if os.path.abspath(src_path) == os.path.abspath(dest):
        return dest          # PDF ist bereits im Archiv
    if os.path.exists(dest):
        # Präfix mit Zeitstempel um Überschreiben zu vermeiden
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S_")
        dest  = os.path.join(PDF_ARCHIVE_DIR, ts + fname)
    try:
        _sh.copy2(src_path, dest)
    except Exception:
        dest = src_path      # Fallback: Original-Pfad behalten
    return dest


def process_one_pdf(conn, path):
    file_name = os.path.basename(path)
    try:
        file_size = os.path.getsize(path)
    except OSError:
        file_size = None

    if not path.lower().endswith(".pdf"):
        status, message = "Uebersprungen", "Keine PDF-Datei"
        log_processed_file(conn, file_name, path, file_size, status, message)
        return status, message

    # PDF ins Archiv kopieren (vor dem Verarbeiten, damit sie immer vorhanden ist)
    archived_path = _archive_pdf(path)

    try:
        import io, contextlib
        _buf = io.StringIO()
        with contextlib.redirect_stdout(_buf):
            summary = extractor.process_pdf(path, conn, verbose=True)
        _verbose_log = _buf.getvalue().strip()
    except Exception as exc:
        status, message = "Fehler", str(exc)
        log_processed_file(conn, file_name, path, file_size, status, message,
                            archived_path=archived_path)
        return status, message

    if not summary:
        ocr_ok  = getattr(extractor, "_OCR_AVAILABLE", False)
        tess_ok = getattr(extractor, "_TESS_OK",       False)
        img_ok  = getattr(extractor, "_PDF2IMG_OK", False) or getattr(extractor, "_FITZ_OK", False)
        if not tess_ok:
            ocr_hint = (" | ⚠ OCR inaktiv: Tesseract nicht installiert. "
                        "Bitte installieren: https://github.com/UB-Mannheim/tesseract/wiki  "
                        "Dann: pip install pytesseract pymupdf")
        elif not img_ok:
            ocr_hint = (" | ⚠ OCR inaktiv: PyMuPDF fehlt. Bitte installieren: "
                        "pip install pymupdf")
        else:
            # OCR war aktiv (incl. Bild-Fallback), konnte aber CBP-Seite nicht finden.
            # Mögliche Ursachen: sehr schlechte Scan-Qualität, gedrehte Seiten, zu klein.
            ocr_hint = (" | OCR-Bild-Fallback wurde ausgeführt, CBP Form 7501 aber nicht "
                        "erkannt. Mögliche Ursachen: Scan-Qualität zu niedrig, Seite gedreht "
                        "oder kein CBP Form 7501 in dieser Datei.")
        status  = "Keine CBP-Seite gefunden"
        message = "Kein Entry Summary (CBP Form 7501) in der PDF erkannt" + ocr_hint
    else:
        parts = []
        new_total = 0
        for entry_no, entry_id, found_lines, inserted_lines in summary:
            new_total += inserted_lines
            parts.append(f"{entry_no}: {found_lines} Position(en), {inserted_lines} neu gespeichert")
        status = "Importiert" if new_total > 0 else "Bereits vorhanden"
        diag = " | DETAIL: " + _verbose_log.replace("\n", " // ") if _verbose_log else ""
        message = "; ".join(parts) + diag

    log_processed_file(conn, file_name, path, file_size, status, message,
                        archived_path=archived_path)
    return status, message


def _esc(v):
    return html.escape("" if v is None else str(v))


def _fmt_ts_html(ts):
    if not ts:
        return ""
    try:
        return datetime.fromisoformat(ts).strftime("%d.%m.%Y %H:%M:%S")
    except (ValueError, TypeError):
        return ts


def _table_html(headers, rows, css_class="data-table"):
    head = "".join(f"<th>{_esc(h)}</th>" for h in headers)
    body = "".join(
        "<tr>" + "".join(f"<td>{_esc(v)}</td>" for v in row) + "</tr>"
        for row in rows
    )
    return (f"<div class='table-wrap'><table class='{css_class}'><thead><tr>{head}</tr></thead>"
            f"<tbody>{body}</tbody></table></div>")


def generate_html_report(conn, out_path):
    """Erzeugt EINEN gemeinsamen HTML-Report: Upload-Hinweis + Status +
    Liste ohne CBP-Seite + kompletter Inhalt der Tabellen 'entries' und 'entry_lines'."""
    conn.row_factory = sqlite3.Row

    status_rows = [
        [r["file_name"], _fmt_ts_html(r["processed_at"]), human_size(r["file_size_bytes"]),
         r["status"], r["message"] or ""]
        for r in conn.execute("SELECT * FROM processed_files ORDER BY id DESC")
    ]
    status_table = (_table_html(["Datei-Name", "Gelesen am", "Groesse", "Status", "Details"],
                                status_rows, "status-table")
                    if status_rows else "<p><em>Noch keine Dateien verarbeitet.</em></p>")

    no_cbp_rows = [
        [r["file_name"], _fmt_ts_html(r["processed_at"]), human_size(r["file_size_bytes"]), r["message"] or ""]
        for r in conn.execute(
            "SELECT * FROM processed_files WHERE status = 'Keine CBP-Seite gefunden' ORDER BY id DESC")
    ]
    no_cbp_table = (_table_html(["Datei-Name", "Gelesen am", "Groesse", "Details"], no_cbp_rows, "status-table")
                    if no_cbp_rows else "<p><em>Alle hochgeladenen PDFs enthielten eine CBP Form 7501.</em></p>")

    try:
        e_cols = [d[0] for d in conn.execute("SELECT * FROM entries LIMIT 1").description]
        e_rows = [list(r) for r in conn.execute(f"SELECT {','.join(e_cols)} FROM entries ORDER BY id")]
    except sqlite3.OperationalError:
        e_cols, e_rows = [], []
    entries_table = _table_html(e_cols, e_rows) if e_rows else "<p><em>Keine Daten.</em></p>"

    try:
        l_cols = [d[0] for d in conn.execute("SELECT * FROM entry_lines LIMIT 1").description]
        l_rows = [list(r) for r in conn.execute(f"SELECT {','.join(l_cols)} FROM entry_lines ORDER BY id")]
    except sqlite3.OperationalError:
        l_cols, l_rows = [], []
    lines_table = _table_html(l_cols, l_rows) if l_rows else "<p><em>Keine Daten.</em></p>"

    bat_uri = "file:///" + STARTER_BAT.replace("\\", "/")

    doc = """<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<title>Catensys Claim-Tariff - CBP Form 7501 Gesamtreport</title>
<style>
  :root { --accent:#1F4E79; --bg:#f7f9fb; --border:#d7dee5; }
  body { font-family: Arial, Helvetica, sans-serif; margin:0; background:var(--bg); color:#222; }
  header { background:var(--accent); color:#fff; padding:24px 32px; }
  header h1 { margin:0; font-size:1.6em; }
  header p { margin:6px 0 0; font-size:0.9em; opacity:.85; }
  main { max-width:1280px; margin:24px auto; padding:0 16px 60px; }
  section { background:#fff; border:1px solid var(--border); border-radius:8px;
            padding:20px 24px; margin-bottom:28px; box-shadow:0 1px 3px rgba(0,0,0,.06); }
  section h2 { color:var(--accent); margin-top:0; border-bottom:2px solid var(--border); padding-bottom:6px; }
  .note { background:#fff7e6; border:1px solid #f0d99a; border-radius:6px; padding:10px 14px; font-size:0.9em; }
  .table-wrap { overflow-x:auto; margin-top:10px; }
  table.data-table, table.status-table { border-collapse:collapse; width:100%; font-size:0.85em; white-space:nowrap; }
  table.data-table th, table.data-table td,
  table.status-table th, table.status-table td { border:1px solid var(--border); padding:6px 10px; text-align:left; }
  table.data-table th, table.status-table th { background:#1F4E79; color:#fff; position:sticky; top:0; }
  table.data-table tr:nth-child(even) td, table.status-table tr:nth-child(even) td { background:#f3f7fa; }
  .badge { display:inline-block; background:#eaf1f7; color:var(--accent); border-radius:12px;
           padding:2px 12px; font-size:0.8em; margin-left:8px; }
  .btn { display:inline-block; background:var(--accent); color:#fff; text-decoration:none;
         padding:10px 22px; border-radius:6px; font-weight:bold; }
  footer { text-align:center; color:#888; font-size:0.8em; padding:20px; }
</style>
</head>
<body>
<header>
  <h1>Catensys Claim-Tariff - CBP Form 7501 Gesamtreport (Import &amp; Datenbank-Inhalt)</h1>
  <p>Datenquelle: __DBNAME__ &nbsp;-&nbsp; erstellt am __NOW__</p>
</header>
<main>

  <section>
    <h2>PDF-Datei importieren</h2>
    <p class="note">
      Hinweis: Eine Web-Seite darf aus Sicherheitsgruenden keine Dateien lesen
      oder die Datenbank veraendern. Fuer den <strong>echten Import</strong> bitte
      die App-Verknuepfung unten per Doppelklick oeffnen und die PDF-Datei(en)
      per Drag &amp; Drop in das erscheinende Fenster ziehen - sie werden sofort
      gelesen, gespeichert und erscheinen danach automatisch in den Tabellen
      weiter unten sowie in diesem Report (nach erneutem Erzeugen).
    </p>
    <p>
      <a class="btn" href="__BATURI__" download="CBP7501-Import-starten.bat">Import-App oeffnen (Verknuepfung)</a>
      <br><span style="font-size:0.8em; color:#777;">
        Datei <code>CBP7501-Import-starten.bat</code> liegt auch direkt im Ordner
        <code>__APPDIR__</code> - dort doppelklicken, um die App ohne Umweg zu starten.
      </span>
    </p>
  </section>

  <section>
    <h2>Status der verarbeiteten PDF-Dateien <span class="badge">__NSTATUS__ Datei(en)</span></h2>
    __STATUSTABLE__
  </section>

  <section>
    <h2>Hochgeladene PDF-Dateien ohne erkannte CBP-Seite (Form 7501) <span class="badge">__NNOCBP__ Datei(en)</span></h2>
    <p>Diese Dateien wurden hochgeladen und gelesen, enthielten jedoch keine
       Seite mit einer CBP Form 7501 (Entry Summary) und wurden daher nicht in
       die Datenbank uebernommen.</p>
    __NOCBPTABLE__
  </section>

  <section>
    <h2>Datenbank-Tabelle: entries <span class="badge">__NENTRIES__ Datensatz/Datensaetze</span></h2>
    <p>Vollstaendige Kopfdaten (Felder 1-26 sowie 35-43) je Entry Summary-Dokument.</p>
    __ENTRIESTABLE__
  </section>

  <section>
    <h2>Datenbank-Tabelle: entry_lines <span class="badge">__NLINES__ Datensatz/Datensaetze</span></h2>
    <p>Vollstaendige Warenpositionen (Felder 27-34), verknuepft ueber <code>entry_id</code> mit <code>entries</code>.</p>
    __LINESTABLE__
  </section>

</main>
<footer>Dieser Report ist ein Schnappschuss des aktuellen Datenbankinhalts (__DBNAME__),
direkt aus der App erzeugt am __NOW__.</footer>
</body>
</html>"""

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    doc = (doc
           .replace("__DBNAME__", _esc(os.path.basename(DB_PATH)))
           .replace("__NOW__", now)
           .replace("__BATURI__", bat_uri)
           .replace("__APPDIR__", _esc(APP_DIR))
           .replace("__NSTATUS__", str(len(status_rows)))
           .replace("__STATUSTABLE__", status_table)
           .replace("__NNOCBP__", str(len(no_cbp_rows)))
           .replace("__NOCBPTABLE__", no_cbp_table)
           .replace("__NENTRIES__", str(len(e_rows)))
           .replace("__ENTRIESTABLE__", entries_table)
           .replace("__NLINES__", str(len(l_rows)))
           .replace("__LINESTABLE__", lines_table))

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(doc)


def generate_xlsx_export(conn, out_path):
    """Exportiert den kompletten Datenbank-Inhalt (Tabellen 'entries' und
    'entry_lines') als Excel-Arbeitsmappe (.xlsx) mit zwei Tabellenblaettern,
    inkl. lesbarer Spaltenueberschriften (wie in der App-Ansicht 'Datenbank-Inhalt')."""
    if openpyxl is None:
        raise ImportError(
            "Das Paket 'openpyxl' fehlt. Bitte installieren mit:\n    pip install openpyxl"
        )

    conn.row_factory = None
    wb = openpyxl.Workbook()

    def _fill_sheet(ws, title, columns, labels, query):
        ws.title = title
        headers = [labels.get(col, col) for col in columns]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
            cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
        for row in conn.execute(query):
            ws.append(["" if v is None else v for v in row])
        ws.freeze_panes = "A2"
        for idx, col in enumerate(columns, start=1):
            header_len = len(labels.get(col, col))
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(12, min(48, header_len + 4))

    entries_columns = list(ENTRIES_LABELS.keys())
    lines_columns = list(LINES_LABELS.keys())

    ws_entries = wb.active
    _fill_sheet(
        ws_entries, "entries", entries_columns, ENTRIES_LABELS,
        f"SELECT {', '.join(entries_columns)} FROM entries ORDER BY id",
    )

    ws_lines = wb.create_sheet("entry_lines")
    _fill_sheet(
        ws_lines, "entry_lines", lines_columns, LINES_LABELS,
        f"SELECT {', '.join(lines_columns)} FROM entry_lines ORDER BY id",
    )

    wb.save(out_path)


# Spalten des Kunden-Claim-Report-Templates, die direkt aus der CBP-7501-
# Datenbank befuellt werden koennen (Rest = Lieferanten-/Stammdaten-Felder,
# die im Customs-Dokument nicht enthalten sind).
CLAIM_AUTO_COLUMNS = ("J", "K", "L", "M", "N", "O", "P")
CLAIM_GAP_TEXT_COLUMNS = ("C", "D", "E", "F", "G", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC")
CLAIM_GAP_NUMERIC_COLUMNS = ("R", "S")
CLAIM_GAP_HINT = ("Datenfeld in CBP Form 7501 / Datenbank nicht vorhanden - "
                  "bitte vom Lieferanten / aus Stammdaten ergaenzen")
CLAIM_FORMULA_COLS = {
    "A": "=$D$6",
    "B": "=+$D$8",
    "Q": "=R{r}/P{r}",
    "AD": '=IF(OR(G{r}="Sec 232 Steel",G{r}="Sec 232 Aluminum"),(Q{r}*N{r}*Z{r}),(Q{r}*N{r}))',
    "AF": "=AD{r}*AE{r}",
    "AG": "=AF{r}*E{r}",
}
CLAIM_LABELS = {
    "C": "Vehicle Line", "D": "Production/FCSD parts", "E": "CMMS Actual Volume",
    "F": "Supplier Shipment Volume", "G": "Tariff Type",
    "R": "Invoice Value", "S": "Invoice Date",
    "T": "Ford Part Prefix", "U": "Ford Part Base Number", "V": "Ford Part Suffix",
    "W": "Supplier Part number", "X": "T2 part number", "Y": "Raw Material",
    "Z": "Weight (Raw material content)", "AA": "UoM", "AB": "AMI Linked (Yes/No)",
    "AC": "If Linked, Index Code",
    "O": "Supplier Invoice No",
}

# ---------------------------------------------------------------------------
# SAP-Konfiguration (MB51 Extrakt)
# ---------------------------------------------------------------------------
SAP_WORKER_DIR    = r"C:\WF\sap-robots\worker"
SAP_WORKER_PYTHON = os.path.join(SAP_WORKER_DIR, ".venv", "Scripts", "python.exe")
SAP_MB51_SCRIPT   = os.path.join(SAP_WORKER_DIR, "mb51_extract.py")
SAP_IV_SCRIPT     = os.path.join(SAP_WORKER_DIR, "iv_extract.py")

SAP_IV_COLUMNS = [
    "Rechnungsnummer", "Jahr", "Belegdatum", "Buchungsdatum",
    "Lieferant", "Ext_Referenz", "Waehrung", "Wechselkurs",
    "Position", "Material", "Werk", "Menge", "Mengeneinheit",
    "Betrag_BelegWaehrung", "Betrag_HausWaehrung",
    "Bestellnummer", "Bestellposition", "Positionstext",
]
# Projektlokale Einstellungsdatei (liegt im selben Ordner wie die App)
SAP_CONFIG_PATH   = os.path.join(APP_DIR, "sap_settings.ini")

SAP_MB51_COLUMNS = [
    "Belegnummer", "Jahr", "Buchungsdatum", "Belegdatum", "Erfassungsdatum",
    "Benutzer", "Transaktion", "Referenzbeleg", "Belegkopftext",
    "Position", "Bewegungsart", "Material", "Werk", "Lagerort",
    "Charge", "Menge", "Mengeneinheit", "Betrag_HW", "Waehrung",
    "Bestellnummer", "Bestellposition", "Lieferant", "Kunde",
    "Kostenstelle", "Auftrag", "Positionstext",
]

# ---------------------------------------------------------------------------
# US.Customs-Import (Entry Summary Line Tariff Details)
# ---------------------------------------------------------------------------
TARIFF_CLAIM_SCHEMA = """
CREATE TABLE IF NOT EXISTS tariff_claim_lines (
    id                     INTEGER PRIMARY KEY AUTOINCREMENT,
    source_file            TEXT,
    imported_at            TEXT NOT NULL,
    entry_summary_number   TEXT,
    entry_type_code        TEXT,
    importer_number        TEXT,
    port_of_entry_code     TEXT,
    entry_date             TEXT,
    entry_summary_date     TEXT,
    initial_create_date    TEXT,
    line_number            INTEGER,
    review_team_number     TEXT,
    country_of_origin      TEXT,
    country_of_export      TEXT,
    manufacturer_id        TEXT,
    foreign_exporter_id    TEXT,
    line_spi_code          TEXT,
    standard_visa_number   TEXT,
    textile_category_code  TEXT,
    textile_category       TEXT,
    tariff_ordinal_number  INTEGER,
    hts_number             TEXT,
    quantity_1             REAL,
    uom_1                  TEXT,
    quantity_2             REAL,
    uom_2                  TEXT,
    quantity_3             REAL,
    uom_3                  TEXT,
    goods_value            REAL,
    duty_amount            REAL
);
"""

TARIFF_CLAIM_COLUMNS = [
    "id", "source_file", "imported_at",
    "entry_summary_number", "entry_type_code", "importer_number",
    "port_of_entry_code", "entry_date", "entry_summary_date",
    "initial_create_date", "line_number", "review_team_number",
    "country_of_origin", "country_of_export", "manufacturer_id",
    "foreign_exporter_id", "line_spi_code", "standard_visa_number",
    "textile_category_code", "textile_category", "tariff_ordinal_number",
    "hts_number", "quantity_1", "uom_1", "quantity_2", "uom_2",
    "quantity_3", "uom_3", "goods_value", "duty_amount",
]

TARIFF_CLAIM_LABELS = {
    "id":                    "ID",
    "source_file":           "Quelldatei",
    "imported_at":           "Importiert am",
    "entry_summary_number":  "Entry Summary Nr.",
    "entry_type_code":       "Entry Type",
    "importer_number":       "Importer Nr.",
    "port_of_entry_code":    "Port Code",
    "entry_date":            "Entry Date",
    "entry_summary_date":    "Summary Date",
    "initial_create_date":   "Create Date",
    "line_number":           "Zeile Nr.",
    "review_team_number":    "Review Team",
    "country_of_origin":     "Ursprungsland",
    "country_of_export":     "Exportland",
    "manufacturer_id":       "Hersteller-ID",
    "foreign_exporter_id":   "Exporter-ID",
    "line_spi_code":         "SPI Code",
    "standard_visa_number":  "Visa-Nr.",
    "textile_category_code": "Textil-Code",
    "textile_category":      "Textil-Kategorie",
    "tariff_ordinal_number": "Tariff Ordinal",
    "hts_number":            "HTS-Nr.",
    "quantity_1":            "Menge 1",
    "uom_1":                 "Einheit 1",
    "quantity_2":            "Menge 2",
    "uom_2":                 "Einheit 2",
    "quantity_3":            "Menge 3",
    "uom_3":                 "Einheit 3",
    "goods_value":           "Warenwert ($)",
    "duty_amount":           "Zollbetrag ($)",
}

# ---------------------------------------------------------------------------
# Claim-Kunde Report (Antrag-Template für Kunden)
# ---------------------------------------------------------------------------
CLAIM_SETTINGS_SCHEMA = """
CREATE TABLE IF NOT EXISTS claim_kunde_settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);"""

CLAIM_SUPPLEMENT_SCHEMA = """
CREATE TABLE IF NOT EXISTS claim_kunde_supplement (
    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
    tcl_id               INTEGER UNIQUE,
    supplier_invoice_no  TEXT,
    invoice_date         TEXT,
    ford_part_base       TEXT,
    supplier_part_number TEXT,
    t2_part_number       TEXT,
    weight               REAL
);"""

CK_SELECTION_SCHEMA = """
CREATE TABLE IF NOT EXISTS ck_row_selection (
    row_key  TEXT PRIMARY KEY,
    selected INTEGER DEFAULT 1
);
"""

CLAIM_PREVIEW_COLS = [
    "vhm", "commodity_code", "vehicle_line", "production_fcsd",
    "cmms_volume", "ship_volume", "tariff_type",
    "claim_start", "claim_end",
    "exporting_country", "entry_no", "hts_code",
    "import_date", "tariff_pct",
    "supplier_invoice_no", "quantity", "unit_price", "invoice_value",
    "invoice_date",
    "ford_prefix", "ford_base", "ford_suffix",
    "supplier_part", "t2_part", "raw_material", "weight", "uom",
    "ami_linked", "index_code",
    "tariff_ppu", "components_per_part", "tariff_per_part", "tariff_claim",
    "material_nr", "kunde_nr", "kunde_name",
]

CLAIM_PREVIEW_LABELS = {
    "vhm":                "VHM",
    "commodity_code":     "Commodity Code",
    "vehicle_line":       "Vehicle Line",
    "production_fcsd":    "Prod./FCSD",
    "cmms_volume":        "CMMS Actual Vol.",
    "ship_volume":        "Supplier Ship. Vol.",
    "tariff_type":        "Tariff Type",
    "claim_start":        "Claim Start",
    "claim_end":          "Claim End",
    "exporting_country":  "Exporting Country (14)",
    "entry_no":           "Filer Code / Entry No (1)",
    "hts_code":           "HTS Code",
    "import_date":        "Import Date (11)",
    "tariff_pct":         "Tariff % (33)",
    "supplier_invoice_no":"Supplier Invoice No",
    "quantity":           "Quantity (30)",
    "unit_price":         "Unit Price",
    "invoice_value":      "Invoice Value",
    "invoice_date":       "Invoice Date",
    "ford_prefix":        "Ford Part Prefix",
    "ford_base":          "Ford Part Base Nr.",
    "ford_suffix":        "Ford Part Suffix",
    "supplier_part":      "Supplier Part Nr.",
    "t2_part":            "T2 Part Nr.",
    "raw_material":       "Raw Material",
    "weight":             "Weight",
    "uom":                "UoM",
    "ami_linked":         "AMI Linked",
    "index_code":         "Index Code",
    "tariff_ppu":         "Tariff Price/Unit",
    "components_per_part":"Components/Part",
    "tariff_per_part":    "Tariff $/Part",
    "tariff_claim":       "Tariff Claim ($)",
    "material_nr":        "Material-Nr. (SAP)",
    "kunde_nr":           "Kunden-Nr.",
    "kunde_name":         "Kundenname",
}

# ---------------------------------------------------------------------------
# UI-Farbkonstanten (Vista-Mar-Stil)
# ---------------------------------------------------------------------------
SIDEBAR_BG  = "#0B1929"   # dunkelblau Sidebar
SIDEBAR_FG  = "#7A94AD"   # gedaempftes Blaugrau (inaktive Nav-Items)
SIDEBAR_ACT = "#D4A843"   # Gold-Akzent (aktives Nav-Item / Logo)
MAIN_BG     = "#F0F2F5"   # helles Grau Hauptflaeche
CARD_BG     = "#FFFFFF"   # weiss fuer Karten

# INPUT-DATA Gruppe (Haupt-Arbeitsfläche)
_NAV_ITEMS_INPUT = [
    ("import",      "⌂", "Dashboard",     "Import & Status"),
    ("db",          "≡", "Datenbank",     "Datenbank-Inhalt"),
    ("claim_kunde", "✎", "Claim-Kunde",   "Claim-Report für Kunden"),
    ("auswertung",  "📈", "Auswertung",    "Auswertung – entry_lines Datenbank-Inhalt"),
]

# SAP-Stammdaten Gruppe (ehemals US-Customs)
_NAV_ITEMS_CUSTOMS = [
    ("sapconn", "⚙", "SAP-Verbindg",  "SAP-Verbindung"),
    ("sap",     "↑", "SAP-Daten",     "SAP-Daten (MB51)"),
    ("sap_iv",  "📄", "Eingangsrechn.", "Eingangsrechnungen (IV)"),
]

# CBP-Report Gruppe (separate Liste fuer eigene Sidebar-Sektion)
_NAV_ITEMS_CBP = [
    ("tariff", "⊕", "US.Customs",    "US.Customs Import (Entry Summary)"),
    ("cape",   "📊", "CAPE Summary",  "CAPE Entry Summary Report"),
    ("liq",    "💰", "Liquidation",   "CBP Liquidation Refund Report"),
]

# Plausibilitaet Gruppe
_NAV_ITEMS_PLAUSI = [
    ("abgleich_manual", "🔍", "Abgleich-Manual", "Plausibilität – Abgleich Manual"),
    ("abgleich_db",     "⚖",  "Abgleich-DB",     "Plausibilität – Abgleich DB vs. Manual"),
]

# Stammdaten Gruppe
_NAV_ITEMS_STAMM = [
    ("zuord_mat_kd",      "🗂", "ZuOrd-MAT-KD",  "Stammdaten – Zuordnung Material / Kunde"),
    ("db_tariff",         "📋", "db_Tariff",      "Stammdaten – Tariff Datenbank"),
    ("htsus_claim_cfg",   "⚙️", "HTSUS Claim",   "HTSUS Claim Konfiguration – Erstattungsfähig?"),
]

# db_Tariff – Spalten für die Summenzeile (alle numerischen Betragsfelder)
_DBT_SUM_COLS = {
    "tax_vol_total_doc", "invoice_value",
    "merchandise_processing_fee", "harbour_fee",
    "hts_3926_90_9989", "hts_3926_90_9905",
    "hts_8409_91_9990", "hts_8483_90_1050",
    "hts_8483_90_8080", "hts_8483_90_9990",
    "hts_8708_99_6805", "hts_8708_99_8180",
    "hts_9903_01_10_25", "hts_9903_01_10_35",
    "hts_9903_01_24a", "hts_9903_01_24b",
    "hts_9903_01_25", "hts_9903_01_33",
    "hts_9903_01_63", "hts_9903_01_77",
    "hts_9903_01_84", "hts_9903_02_09",
    "hts_9903_02_20", "hts_9903_02_58",
    "hts_9903_03_01", "hts_9903_81_90_25",
    "hts_9903_81_90_50", "hts_9903_81_91",
    "hts_9903_82_02", "hts_9903_82_09",
    "hts_9903_88_01", "hts_9903_88_03",
    "hts_9903_88_15", "hts_9903_94_05",
    "ascertained_other", "summe_tax",
    "quantity", "rate",
    "invoice_doc_value", "local_value_sap",
    "claim_value_1", "claim_value_2",
}

# ZuOrd-MAT-KD Excel-Spalten (Sheet "StammdatenKundenMAT", Header Zeile 1, Daten ab Zeile 2)
ZUORD_COLUMNS = ["mat_no_sap", "material_cat_sap", "kunden_nr", "sap_customer_name"]
ZUORD_HEADERS = ["Mat-No. SAP", "Material CAT-SAP", "Kunden-Nr", "SAP Customer Name"]

ZUORD_SCHEMA = """
CREATE TABLE IF NOT EXISTS stammdaten_zuord_mat_kd (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    mat_no_sap      TEXT,
    material_cat_sap TEXT,
    kunden_nr       TEXT,
    sap_customer_name TEXT
);"""

# db_Tariff Excel-Spalten (Sheet "Data", Header Zeile 1, Daten ab Zeile 2)
DB_TARIFF_COLUMNS = [
    "year_month", "dachser_duty_doc_no", "entry_no_new", "line_item",
    "tax_vol_total_doc", "tax_vol_total_doc_curr",
    "invoice_value", "invoice_value_curr",
    "merchandise_processing_fee", "harbour_fee",
    "hts_3926_90_9989", "hts_3926_90_9905",
    "hts_8409_91_9990", "hts_8483_90_1050",
    "hts_8483_90_8080", "hts_8483_90_9990",
    "hts_8708_99_6805", "hts_8708_99_8180",
    "hts_9903_01_10_25", "hts_9903_01_10_35",
    "hts_9903_01_24a", "hts_9903_01_24b",
    "hts_9903_01_25", "hts_9903_01_33",
    "hts_9903_01_63", "hts_9903_01_77",
    "hts_9903_01_84", "hts_9903_02_09",
    "hts_9903_02_20", "hts_9903_02_58",
    "hts_9903_03_01", "hts_9903_81_90_25",
    "hts_9903_81_90_50", "hts_9903_81_91",
    "hts_9903_82_02", "hts_9903_82_09",
    "hts_9903_88_01", "hts_9903_88_03",
    "hts_9903_88_15", "hts_9903_94_05",
    "ascertained_other", "summe_tax",
    "dachser_dokument", "entry_document",
    "vendor", "vendor_name",
    "vendor_origin_o", "vendor_origin_e",
    "invoice_no_delivery_note", "po_order_no",
    "mat_no_sap", "material_cat_sap",
    "quantity", "rate",
    "invoice_doc_value", "doc_currency",
    "local_value_sap", "local_currency",
    "sap_customer_no", "sap_customer_name",
    "customer_part_no", "claim_status",
    "claim_value_1", "claim_value_2",
]

DB_TARIFF_HEADERS = [
    "Year/Month", "Dachser Doc-No", "Entry-No New", "Line Item",
    "Tax Vol. Total", "Tax Vol. Curr.",
    "Invoice Value", "Inv. Curr.",
    "Merch. Proc. Fee", "Harbour Fee",
    "3926.90.9989", "3926.90.9905",
    "8409.91.9990", "8483.90.1050",
    "8483.90.8080", "8483.90.9990",
    "8708.99.6805", "8708.99.8180",
    "9903.01.10 (25%)", "9903.01.10 (35%)",
    "9903.01.24a", "9903.01.24b",
    "9903.01.25", "9903.01.33",
    "9903.01.63", "9903.01.77",
    "9903.01.84", "9903.02.09",
    "9903.02.20", "9903.02.58",
    "9903.03.01", "9903.81.90 (25%)",
    "9903.81.90 (50%)", "9903.81.91",
    "9903.82.02", "9903.82.09",
    "9903.88.01", "9903.88.03",
    "9903.88.15", "9903.94.05",
    "Ascertained Other", "Summe TAX",
    "Dachser Dokument", "Entry Document",
    "Vendor", "Vendor Name",
    "Vendor Origin (O)", "Vendor Origin (E)",
    "Invoice-No/Del. Note", "PO/Order-No.",
    "Mat-No. SAP", "Material CAT-SAP",
    "Quantity", "Rate",
    "Invoice Doc Value", "Doc Currency",
    "Local Value SAP", "Local Currency",
    "SAP Customer No", "SAP Customer Name",
    "Customer Part-No.", "Claim Status",
    "Claim Value 1", "Claim Value 2",
]

HTSUS_CLAIM_SCHEMA = """
CREATE TABLE IF NOT EXISTS htsus_claim_config (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    htsus_no          TEXT NOT NULL UNIQUE,
    htsus_rate        TEXT,
    erstattungsfaehig TEXT CHECK(erstattungsfaehig IN ('ja','nein')) DEFAULT 'nein',
    comment_1         TEXT,
    comment_2         TEXT,
    comment_3         TEXT
);
"""

PDF_VENDOR_SCHEMA = """
CREATE TABLE IF NOT EXISTS pdf_vendor_extract (
    id                       INTEGER PRIMARY KEY AUTOINCREMENT,
    pdf_filename             TEXT NOT NULL,
    dachser_freight_no       TEXT,
    dhl_doc_no               TEXT,
    entry_no                 TEXT,
    vendor_name              TEXT,
    vendor_invoice_no        TEXT,
    vendor_invoice_date      TEXT,
    vendor_po_no             TEXT,
    vendor_catalogue_no      TEXT,
    vendor_tariff_code       TEXT,
    vendor_quantity          TEXT,
    vendor_net_weight        TEXT,
    vendor_value             TEXT,
    vendor_currency          TEXT,
    raw_json                 TEXT,
    processed_at             TEXT DEFAULT (datetime('now')),
    UNIQUE(pdf_filename)
);
"""

DB_TARIFF_SCHEMA = """
CREATE TABLE IF NOT EXISTS stammdaten_db_tariff (
    id                       INTEGER PRIMARY KEY AUTOINCREMENT,
    year_month               TEXT,
    dachser_duty_doc_no      TEXT,
    entry_no_new             TEXT,
    line_item                TEXT,
    tax_vol_total_doc        TEXT,
    tax_vol_total_doc_curr   TEXT,
    invoice_value            TEXT,
    invoice_value_curr       TEXT,
    merchandise_processing_fee TEXT,
    harbour_fee              TEXT,
    hts_3926_90_9989         TEXT,
    hts_3926_90_9905         TEXT,
    hts_8409_91_9990         TEXT,
    hts_8483_90_1050         TEXT,
    hts_8483_90_8080         TEXT,
    hts_8483_90_9990         TEXT,
    hts_8708_99_6805         TEXT,
    hts_8708_99_8180         TEXT,
    hts_9903_01_10_25        TEXT,
    hts_9903_01_10_35        TEXT,
    hts_9903_01_24a          TEXT,
    hts_9903_01_24b          TEXT,
    hts_9903_01_25           TEXT,
    hts_9903_01_33           TEXT,
    hts_9903_01_63           TEXT,
    hts_9903_01_77           TEXT,
    hts_9903_01_84           TEXT,
    hts_9903_02_09           TEXT,
    hts_9903_02_20           TEXT,
    hts_9903_02_58           TEXT,
    hts_9903_03_01           TEXT,
    hts_9903_81_90_25        TEXT,
    hts_9903_81_90_50        TEXT,
    hts_9903_81_91           TEXT,
    hts_9903_82_02           TEXT,
    hts_9903_82_09           TEXT,
    hts_9903_88_01           TEXT,
    hts_9903_88_03           TEXT,
    hts_9903_88_15           TEXT,
    hts_9903_94_05           TEXT,
    ascertained_other        TEXT,
    summe_tax                TEXT,
    dachser_dokument         TEXT,
    entry_document           TEXT,
    vendor                   TEXT,
    vendor_name              TEXT,
    vendor_origin_o          TEXT,
    vendor_origin_e          TEXT,
    invoice_no_delivery_note TEXT,
    po_order_no              TEXT,
    mat_no_sap               TEXT,
    material_cat_sap         TEXT,
    quantity                 TEXT,
    rate                     TEXT,
    invoice_doc_value        TEXT,
    doc_currency             TEXT,
    local_value_sap          TEXT,
    local_currency           TEXT,
    sap_customer_no          TEXT,
    sap_customer_name        TEXT,
    customer_part_no         TEXT,
    claim_status             TEXT,
    claim_value_1            TEXT,
    claim_value_2            TEXT
);"""

# Plausi-Abgleich Excel-Spalten (Sheet "Plausi", Header Zeile 5, Daten ab Zeile 6)
PLAUSI_COLUMNS = [
    "year_month", "dachser_duty_doc_no", "entry_no",
    "tax_volume_total_doc", "tax_volume_total_doc_currency",
    "invoices_base_for_tax", "invoices_currency_base_for_tax",
    "merchandise_processing_fee", "harbour_fee",
    "hts_3926_90_9989", "hts_8483_90_1050", "hts_8483_90_8080",
    "hts_8708_99_6805", "hts_8708_99_8180", "hts_9903_81_90",
    "hts_9903_01_10", "hts_9903_01_24", "hts_9903_88_01",
    "summe_tax", "vendor", "vendor_name",
    "vendor_origin_o", "vendor_origin_e",
    "invoice_no_from_vendor", "po_order_no",
    "catensys_mat_no_sap", "catensys_mat_desc_sap",
    "quantity", "rate", "invoice_doc_value", "doc_currency",
    "local_value_sap", "local_currency",
    "reference_doc", "material_doc", "batch",
    "air_see_truck", "customer_partno",
    "sap_customer_no", "sap_customer_descr", "project_contingent",
]
PLAUSI_HEADERS = [
    "Year/Month", "Dachser Doc-No", "Entry-No",
    "Tax Vol. Total", "Currency",
    "Invoice Base", "Invoice Currency",
    "MPF", "Harbour Fee",
    "3926.90.9989", "8483.90.1050", "8483.90.8080",
    "8708.99.6805", "8708.99.8180", "9903.81.90",
    "9903.01.10", "9903.01.24", "9903.88.01",
    "Summe TAX", "Vendor", "Vendor Name",
    "Origin (O)", "Origin (E)",
    "Invoice-No Vendor", "PO/Order-No",
    "Mat-No SAP", "Mat-Desc SAP",
    "Quantity", "Rate", "Invoice Doc Value", "Doc Currency",
    "Local Value SAP", "Local Currency",
    "Reference Doc", "Material Doc", "Batch",
    "Transport", "Customer Part-No",
    "SAP Customer No", "SAP Customer Descr", "Project",
]

# CAPE Excel-Spalten (Reihenfolge wie im Excel-Sheet "Main Report")
CAPE_COLUMNS = [
    "claim_number", "claim_status", "filer_code", "importer_number",
    "cf4811_number", "cf4811_name", "entry_summary_number",
    "entry_date", "liquidation_date", "control_team_number",
    "center_id_code", "center_id_name", "refund_amount", "interest_amount",
    "refund_number", "refund_date", "refund_status",
]
CAPE_HEADERS = [
    "Claim Nr.", "Status", "Filer", "Importer Nr.",
    "CF4811 Nr.", "CF4811 Name", "Entry Summary Nr.",
    "Entry Date", "Liquidation Date", "Control Team",
    "Center Code", "Center Name", "Refund Amt.", "Interest Amt.",
    "Refund Nr.", "Refund Date", "Refund Status",
]

# Liquidation Excel-Spalten (Sheet "Report", Header Zeile 6, Daten ab Zeile 7)
LIQ_COLUMNS = [
    "refund_id", "payee_id", "company_name", "co", "address",
    "refund_date", "refund_status", "refund_secondary_status",
    "refund_type", "document_number", "total_refund_amount",
    "check_ach_trace_number", "check_or_ach",
    "center_id", "team_number", "port_code",
]
LIQ_HEADERS = [
    "Refund ID", "Payee ID", "Company Name", "C/O", "Address",
    "Refund Date", "Status", "Secondary Status",
    "Refund Type", "Document Nr.", "Total Refund Amt.",
    "Check / ACH Trace", "Check/ACH",
    "Center ID", "Team Nr.", "Port Code",
]


def generate_claim_report(conn, template_path, out_path, period_mm, period_yyyy):
    """Erzeugt den Kunden-Claim-Report fuer einen Monat (period_mm/period_yyyy,
    z. B. '03'/'2025') aus der Vorlage 'template_Claim-Report_Kunde.xlsx',
    befuellt mit den Daten der CBP-7501-Datenbank (Entry Summaries mit
    Importdatum in diesem Monat).

    Liefert (anzahl_zeilen, fehlende_felder) zurueck, wobei fehlende_felder
    eine sortierte Liste von (Spaltenbuchstabe, Feldname, betroffene_zeilen)
    ist - fuer die Rueckmeldung an den Anwender, welche Angaben noch fehlen."""
    if openpyxl is None:
        raise ImportError(
            "Das Paket 'openpyxl' fehlt. Bitte installieren mit:\n    pip install openpyxl"
        )
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Vorlage nicht gefunden:\n{template_path}\n\n"
            "Bitte 'template_Claim-Report_Kunde.xlsx' in den App-Ordner legen."
        )

    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment
    yellow = PatternFill("solid", fgColor="FFFF00")

    conn.row_factory = sqlite3.Row
    like_pattern = f"{period_mm}/%/{period_yyyy[2:]}"
    rows = list(conn.execute(
        """SELECT e.exporting_country, e.filer_code_entry_no, e.import_date,
                  l.htsus_no, l.htsus_rate, l.invoice_no, l.net_quantity,
                  l.invoice_value_amount, l.entered_value, l.line_no
           FROM entry_lines l JOIN entries e ON e.id = l.entry_id
           WHERE e.import_date LIKE ?
           ORDER BY l.id""",
        (like_pattern,),
    ).fetchall())

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Tabelle1"]

    first_data_row, last_template_row = 4, ws.max_row
    for r in range(first_data_row, last_template_row + 1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=r, column=col)
            c.value = None
            c.fill = PatternFill(fill_type=None)
            c.comment = None

    def set_gap(cell):
        cell.fill = yellow
        cell.comment = Comment(CLAIM_GAP_HINT, "Catensys Tariff-Tool")

    missing_rows = {}  # col -> set of row numbers with a gap

    period_start = f"{period_mm}/01/{period_yyyy}"
    last_day = {"01": 31, "02": 28, "03": 31, "04": 30, "05": 31, "06": 30,
                "07": 31, "08": 31, "09": 30, "10": 31, "11": 30, "12": 31}.get(period_mm, 31)
    period_end = f"{period_mm}/{last_day:02d}/{period_yyyy}"

    start = first_data_row
    for i, r in enumerate(rows):
        row = start + i
        rate = r["htsus_rate"]
        try:
            rate_val = float(str(rate).replace("%", "").strip()) / 100
        except Exception:
            rate_val = None
        try:
            qty_val = float(r["net_quantity"])
        except Exception:
            qty_val = r["net_quantity"]

        for col, formula in CLAIM_FORMULA_COLS.items():
            ws[f"{col}{row}"] = formula.format(r=row)

        ws[f"H{row}"] = period_start
        ws[f"I{row}"] = period_end
        ws[f"J{row}"] = r["exporting_country"]
        ws[f"K{row}"] = r["filer_code_entry_no"]
        ws[f"L{row}"] = r["htsus_no"]
        ws[f"M{row}"] = r["import_date"]
        ws[f"N{row}"] = rate_val
        ws[f"P{row}"] = qty_val

        if r["invoice_no"]:
            ws[f"O{row}"] = r["invoice_no"]
        else:
            set_gap(ws[f"O{row}"])
            missing_rows.setdefault("O", set()).add(row)

        ws[f"AE{row}"] = 1
        ws[f"AE{row}"].fill = yellow
        ws[f"AE{row}"].comment = Comment(
            "Default-Annahme It. Template = 1 Bauteil pro Teil - "
            "bitte mit Lieferant/Stammdaten bestaetigen", "Catensys Tariff-Tool")
        missing_rows.setdefault("AE", set()).add(row)

        for col in CLAIM_GAP_NUMERIC_COLUMNS + CLAIM_GAP_TEXT_COLUMNS:
            set_gap(ws[f"{col}{row}"])
            missing_rows.setdefault(col, set()).add(row)

    wb.save(out_path)

    missing_summary = []
    for col in ("AE", "O") + CLAIM_GAP_NUMERIC_COLUMNS + CLAIM_GAP_TEXT_COLUMNS:
        if col in missing_rows:
            label = CLAIM_LABELS.get(col, col)
            if col == "AE":
                label = "Components Per Part (Annahme = 1, bitte bestaetigen)"
            missing_summary.append((col, label, sorted(missing_rows[col])))

    return len(rows), missing_summary


# Schluesselwoerter fuer Betragsspalten (genutzt in _insert_grand_total)
_SUM_KEYWORDS = frozenset({
    "amount", "total", "value", "fee", "betrag", "wert",
    "menge", "qty", "quantity", "tariff", "price", "weight",
    "duty", "mpf", "hmf", "tax", "grand", "summe",
    "invoice_value", "tariff_claim", "tariff_per", "tariff_ppu",
    "refund", "interest",
})

def _is_sum_col(col_name):
    """True wenn Spaltenname auf einen Betragswert hindeutet."""
    col_lower = str(col_name).lower()
    return any(kw in col_lower for kw in _SUM_KEYWORDS)


class ScrollableTable(ttk.Frame):
    def __init__(self, parent, columns, labels, col_width=130,
                 subtotal_group_col=None, subtotal_sum_col=None,
                 subtotal_label_col=None):
        super().__init__(parent)
        self.columns = columns
        self._all_rows = []
        self._subtotal_group_col  = subtotal_group_col   # column index to group by
        self._subtotal_sum_col    = subtotal_sum_col     # column index to sum
        self._subtotal_label_col  = subtotal_label_col   # column index for label text
        self._subtotal_id_labels  = {}                   # {group_id: label_str}
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=12)
        for col in columns:
            self.tree.heading(col, text=labels.get(col, col))
            self.tree.column(col, width=col_width, minwidth=60, anchor="w", stretch=False)
        # Light-green subtotal row style
        # background= works on Windows/tk8.6+; foreground+font work everywhere
        self.tree.tag_configure(
            "subtotal",
            background="#D1FAE5",
            foreground="#065F46",
            font=("Segoe UI", 9, "bold"),
        )
        self.tree.tag_configure(
            "grand_total",
            background="#1F4E79",
            foreground="white",
            font=("Segoe UI", 9, "bold"),
        )
        # Workaround: force tk to honour tag background in clam theme
        try:
            self.tree.tk.call("ttk::style", "map", "Treeview",
                              "-background", [])
        except Exception:
            pass

        vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

    def set_rows(self, rows):
        self._all_rows = list(rows)
        self._show(self._all_rows)

    def _show(self, rows):
        for item in self.tree.get_children():
            self.tree.delete(item)
        if (self._subtotal_group_col is not None
                and self._subtotal_sum_col is not None):
            first       = True
            current_grp = None
            group_total = 0.0
            for row in rows:
                values = ["" if v is None else v for v in row]
                grp = (row[self._subtotal_group_col]
                       if len(row) > self._subtotal_group_col else None)
                if not first and grp != current_grp:
                    self._insert_subtotal_row(current_grp, group_total)
                    group_total = 0.0
                first       = False
                current_grp = grp
                raw = (row[self._subtotal_sum_col]
                       if len(row) > self._subtotal_sum_col else None)
                if raw not in (None, ""):
                    try:
                        group_total += float(str(raw).replace(",", "."))
                    except ValueError:
                        pass
                self.tree.insert("", "end", values=values)
            if not first:
                self._insert_subtotal_row(current_grp, group_total)
        else:
            for row in rows:
                values = ["" if v is None else v for v in row]
                self.tree.insert("", "end", values=values)
        self._insert_grand_total(rows)
        return len(rows)

    def _insert_subtotal_row(self, grp, total):
        """Insert a light-green subtotal row after a group of rows."""
        sub = [""] * len(self.columns)
        if self._subtotal_group_col is not None:
            sub[self._subtotal_group_col] = grp if grp is not None else ""
        # Put "∑ ZWISCHENSUMME" marker in line_no column (index 3) if available
        _line_no_col = 3 if len(sub) > 3 else None
        if _line_no_col is not None:
            sub[_line_no_col] = "∑ SUMME"
        if self._subtotal_label_col is not None:
            lbl = self._subtotal_id_labels.get(grp, f"Entry {grp}")
            sub[self._subtotal_label_col] = f"{lbl}"
        sub[self._subtotal_sum_col] = _de_num(total)
        self.tree.insert("", "end", values=sub, tags=("subtotal",))

    def _insert_grand_total(self, rows):
        """Dunkelblauer Gesamt-Footer am Ende (auto-detect Betragsspalten)."""
        if not rows:
            return
        ncols = len(self.columns)
        sums = {}
        for ci, col in enumerate(self.columns):
            if not _is_sum_col(col):
                continue
            total = 0.0
            found = False
            for row in rows:
                v = row[ci] if ci < len(row) else None
                parsed = _parse_de(v)
                if parsed is not None:
                    total += parsed
                    found = True
            if found:
                sums[ci] = total
        if not sums:
            return
        footer = [""] * ncols
        footer[0] = "\u2211 GESAMT"
        for ci, total in sums.items():
            footer[ci] = _de_num(total)
        self.tree.insert("", "end", values=footer, tags=("grand_total",))

    def apply_filter(self, text, col_idx=None):
        if not text.strip():
            return self._show(self._all_rows)
        txt = text.strip().lower()
        if col_idx is not None and col_idx >= 0:
            filtered = [r for r in self._all_rows
                        if txt in str(r[col_idx] if r[col_idx] is not None else "").lower()]
        else:
            filtered = [r for r in self._all_rows
                        if any(txt in str(v if v is not None else "").lower() for v in r)]
        return self._show(filtered)

    def reset_filter(self):
        return self._show(self._all_rows)

    def total_count(self):
        return len(self._all_rows)


def _setup_filter_bar(parent, table, col_names, col_labels):
    """
    Baut eine Live-Suchleiste + klappbaren Mehrfach-Spaltenfilter in 'parent' ein.
    Gibt eine reapply()-Funktion zurueck (nach set_rows() aufrufen).
    Alle aktiven Filter werden mit AND-Logik kombiniert.
    """
    col_headers = [col_labels.get(c, c) if isinstance(col_labels, dict) else str(c)
                   for c in col_names]

    # ── Toolbar-Zeile ──────────────────────────────────────────────────────
    try:
        _tbg = parent.cget("background")
    except Exception:
        _tbg = "white"
    toolbar = tk.Frame(parent, bg=_tbg)
    toolbar.pack(fill="x")

    search_var = tk.StringVar()
    count_var  = tk.StringVar(value="")
    _panel_open = [False]
    _get_col_filters = [lambda: []]
    _reset_col = [lambda: None]
    _panel_ref = [None]

    ttk.Label(toolbar, text="Suche:").pack(side="left")
    search_entry = ttk.Entry(toolbar, textvariable=search_var, width=22)
    search_entry.pack(side="left", padx=(4, 6))

    toggle_lbl = tk.StringVar(value="▶ Spaltenfilter")
    toggle_btn = ttk.Button(toolbar, textvariable=toggle_lbl,
                            command=lambda: _toggle())
    toggle_btn.pack(side="left", padx=(0, 6))

    def _do_filter(*_):
        txt = search_var.get().lower().strip()
        col_filters = _get_col_filters[0]()
        if not txt and not col_filters:
            n = table.reset_filter()
        else:
            filtered = [r for r in table._all_rows
                        if _row_passes(r, col_filters, txt)]
            n = table._show(filtered)
        total = table.total_count()
        active = len(col_filters)
        badge = f"  [{active} Filter]" if active else ""
        count_var.set(f"{n} von {total} Zeilen{badge}" if (txt or active) else f"{total} Zeilen")

    def _reset(*_):
        search_var.set("")
        _reset_col[0]()
        n = table.reset_filter()
        count_var.set(f"{n} Zeilen")
        if _panel_open[0]:
            _toggle()  # panel schliessen

    def _toggle():
        p = _panel_ref[0]
        if p is None:
            return
        if _panel_open[0]:
            p.pack_forget()
            _panel_open[0] = False
            toggle_lbl.set("▶ Spaltenfilter")
        else:
            p.pack(fill="x", pady=(2, 0))
            _panel_open[0] = True
            toggle_lbl.set("▼ Spaltenfilter")

    search_entry.bind("<Return>",     _do_filter)
    search_entry.bind("<KeyRelease>", _do_filter)

    ttk.Button(toolbar, text="Zuruecksetzen", command=_reset).pack(side="left", padx=(0, 6))
    ttk.Label(toolbar, textvariable=count_var,
              foreground="#1F4E79", width=24, anchor="w").pack(side="left", padx=(8, 0))

    # ── Klappbares Spalten-Filter-Panel ────────────────────────────────────
    panel, get_fn, reset_fn = _build_col_filter_panel(
        parent, col_headers, _do_filter, cols_per_row=5)
    _panel_ref[0] = panel
    _get_col_filters[0] = get_fn
    _reset_col[0] = reset_fn
    # Panel ist standardmaessig ausgeblendet

    return _do_filter   # caller invokes this after set_rows() to refresh the count label



def _build_col_filter_panel(parent, headers, on_change, bg=MAIN_BG, cols_per_row=5):
    """
    Klappbares Raster-Filter-Panel: ein Entry-Feld pro Spalte (AND-Logik).
    Rueckgabe: (panel_frame, get_filters_fn, reset_fn)
      get_filters_fn() -> list[(col_idx, lower_term)]  – nur nicht-leere Felder
      reset_fn()       -> leert alle Eingabefelder
    """
    panel = tk.Frame(parent, bg=bg, padx=4, pady=6)
    _entries = []
    for flat_i, hdr in enumerate(headers):
        r = flat_i // cols_per_row
        c = flat_i % cols_per_row
        cell = tk.Frame(panel, bg=bg)
        cell.grid(row=r, column=c, padx=5, pady=2, sticky="ew")
        tk.Label(cell, text=hdr, bg=bg, fg="#5A6A80",
                 font=("Segoe UI", 8)).pack(anchor="w")
        v = tk.StringVar()
        tk.Entry(cell, textvariable=v, width=13, relief="flat",
                 bg="white", fg="#1A2742", insertbackground="#1A2742",
                 font=("Segoe UI", 9)).pack(ipady=3, fill="x")
        v.trace_add("write", lambda *_, fn=on_change: fn())
        _entries.append(v)

    def get_filters():
        return [(i, v.get().strip().lower())
                for i, v in enumerate(_entries) if v.get().strip()]

    def reset():
        for v in _entries:
            v.set("")

    return panel, get_filters, reset


def _row_passes(row, col_filters, global_term=""):
    """True wenn Zeile ALLE aktiven Filter erfuellt (AND-Logik)."""
    if global_term and not any(
            global_term in str(v).lower() for v in row if v is not None):
        return False
    for col_idx, term in col_filters:
        val = (str(row[col_idx]).lower()
               if col_idx < len(row) and row[col_idx] is not None else "")
        if term not in val:
            return False
    return True


def _de_num(v, dec=2):
    """Zahl in deutsches Zahlenformat: 1234.56 -> '1.234,56'."""
    if v is None or v == "":
        return ""
    try:
        return f"{float(str(v).replace(',','.').replace(' ','')):,.{dec}f}"               .replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(v)


def _parse_de(v):
    """Deutsches oder englisches Zahlen-String -> float.
    Versteht: '5.101,30'  '311,30'  '1,234.56'  '311.30'  '5101.30'
    """
    if v is None or v == "" or v == "–":
        return None
    s = str(v).strip().replace(" ", "").replace(" ", "")
    if not s:
        return None
    # Deutsches Format: Punkt als Tausender, Komma als Dezimal
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):          # z.B. "5.101,30"
            s = s.replace(".", "").replace(",", ".")
        else:                                     # z.B. "1,234.56"
            s = s.replace(",", "")
    elif "," in s:                                # nur Komma -> Dezimal
        s = s.replace(",", ".")
    # else: nur Punkt oder nichts -> unveraendert
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _to_de(v):
    """Amerikanische Zahl -> deutsches Format: '8,151.85' -> '8.151,85'."""
    if v is None:
        return ""
    s = str(v).strip().replace(",", "")   # Tausender-Komma weg
    try:
        return _de_num(float(s))
    except ValueError:
        return s


class App:
    STATUS_COLORS = {
        "Importiert": "#1c7c33",
        "Bereits vorhanden": "#8a7000",
        "Keine CBP-Seite gefunden": "#a04a00",
        "Fehler": "#b00020",
        "Uebersprungen": "#777777",
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Catensys Claim-Tariff – CBP Form 7501")
        self.root.geometry("1280x820")
        self.root.minsize(1000, 640)
        self.root.configure(bg=SIDEBAR_BG)

        self.conn = get_db_connection()

        # ttk-Styles konfigurieren
        self._apply_styles()

        # ── Aeusserer Shell: Sidebar links + Hauptflaeche rechts ──
        shell = tk.Frame(root, bg=SIDEBAR_BG)
        shell.pack(fill="both", expand=True)

        # Sidebar
        self._sidebar = tk.Frame(shell, bg=SIDEBAR_BG, width=218)
        self._sidebar.pack(side="left", fill="y")
        self._sidebar.pack_propagate(False)
        self._build_sidebar()

        # Trennlinie
        tk.Frame(shell, bg="#1E3A5F", width=1).pack(side="left", fill="y")

        # Hauptflaeche
        main = tk.Frame(shell, bg=MAIN_BG)
        main.pack(side="left", fill="both", expand=True)

        # Header-Leiste
        self._header_frame = tk.Frame(main, bg=CARD_BG, height=64)
        self._header_frame.pack(fill="x")
        self._header_frame.pack_propagate(False)
        self._build_header()
        tk.Frame(main, bg="#E4E8EF", height=1).pack(fill="x")

        # Seiten-Container (eine Seite gleichzeitig sichtbar)
        self._pages_outer = tk.Frame(main, bg=MAIN_BG)
        self._pages_outer.pack(fill="both", expand=True)

        # Seiten-Frames anlegen
        self.tab_import      = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_db          = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_tariff      = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_claim_kunde = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_sap         = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_sapconn     = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_sap_iv      = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_cape           = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_liq            = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_abgleich_manual = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_abgleich_db     = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_zuord_mat_kd    = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_db_tariff       = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_htsus_claim_cfg = tk.Frame(self._pages_outer, bg=MAIN_BG)
        self.tab_auswertung      = tk.Frame(self._pages_outer, bg=MAIN_BG)

        self._build_import_tab()
        self._build_db_tab()
        self._build_tariff_tab()
        self._build_claim_kunde_tab()
        self._build_sap_tab()
        self._build_sap_conn_tab()
        self._build_sap_iv_tab()
        self._build_cape_tab()
        self._build_liq_tab()
        self._build_abgleich_manual_tab()
        self._build_abgleich_db_tab()
        self._build_zuord_mat_kd_tab()
        self._build_db_tariff_tab()
        self._build_htsus_claim_cfg_tab()
        self._build_auswertung_tab()

        # Startseite anzeigen
        self._current_page = None
        self._nav_select("import")

        self._load_history()
        self._load_db_tables()
        self._load_tariff_data()
        self._load_sap_data()

    # -----------------------------------------------------------------------
    # Styles
    # -----------------------------------------------------------------------
    def _apply_styles(self):
        s = ttk.Style()
        try:
            s.theme_use("clam")
        except Exception:
            pass

        # Treeview
        s.configure("Treeview",
            background=CARD_BG, foreground="#1A2742",
            rowheight=25, fieldbackground=CARD_BG,
            borderwidth=0, font=("Segoe UI", 9),
        )
        s.configure("Treeview.Heading",
            background="#1A2742", foreground="white",
            font=("Segoe UI", 9, "bold"),
            borderwidth=0, relief="flat",
        )
        s.map("Treeview.Heading",
            background=[("active", "#243659")],
        )
        s.map("Treeview",
            background=[("selected", "#DBEAFE")],
            foreground=[("selected", "#1E3A8A")],
        )

        # Progressbar grün
        s.configure(
            "green.Horizontal.TProgressbar",
            troughcolor="#D1FAE5",
            background="#16A34A",
            bordercolor="#D1FAE5",
            lightcolor="#22C55E",
            darkcolor="#15803D",
        )

        # Buttons
        s.configure("TButton",
            background=CARD_BG, foreground="#1A2742",
            font=("Segoe UI", 9), relief="solid",
            borderwidth=1, padding=(10, 5),
        )
        s.map("TButton",
            background=[("active", "#F0F2F5"), ("pressed", "#E4E8EF")],
        )
        s.configure("Primary.TButton",
            background="#1A2742", foreground="white",
            font=("Segoe UI", 9, "bold"),
        )
        s.map("Primary.TButton",
            background=[("active", "#243659"), ("pressed", "#0F1929")],
        )

        # Labels
        s.configure("TLabel", background=MAIN_BG, font=("Segoe UI", 10))

        # LabelFrame  (Karten-Stil)
        s.configure("TLabelframe",
            background=CARD_BG, bordercolor="#E4E8EF",
            relief="solid", padding=10,
        )
        s.configure("TLabelframe.Label",
            background=CARD_BG, foreground="#1A2742",
            font=("Segoe UI", 10, "bold"),
        )

        # Frame
        s.configure("TFrame", background=MAIN_BG)

        # Scrollbar
        s.configure("Vertical.TScrollbar",
            background="#E4E8EF", troughcolor=MAIN_BG,
            arrowcolor="#8A9EB5", borderwidth=0, relief="flat",
        )
        s.configure("Horizontal.TScrollbar",
            background="#E4E8EF", troughcolor=MAIN_BG,
            arrowcolor="#8A9EB5", borderwidth=0, relief="flat",
        )

        # Entry / Combobox
        s.configure("TEntry",
            fieldbackground=CARD_BG, bordercolor="#E4E8EF",
            font=("Segoe UI", 10),
        )
        s.configure("TCombobox",
            fieldbackground=CARD_BG, background=CARD_BG,
            font=("Segoe UI", 10),
        )

    # -----------------------------------------------------------------------
    # Sidebar
    # -----------------------------------------------------------------------
    def _build_collapsible_section(self, sb, title, open_default=False):
        """Baut eine aufklappbare Sidebar-Sektion.
        Gibt den Content-Frame zurück, in den Nav-Items eingefügt werden."""
        section_f = tk.Frame(sb, bg=SIDEBAR_BG)
        section_f.pack(fill="x")

        tk.Frame(section_f, bg="#1E3A5F", height=1).pack(fill="x", pady=(4, 0))

        hdr_text = ("▼  " if open_default else "▶  ") + title
        hdr_lbl = tk.Label(
            section_f,
            text=hdr_text,
            bg=SIDEBAR_BG, fg=SIDEBAR_ACT,
            font=("Segoe UI", 8, "bold"),
            anchor="w", cursor="hand2",
            padx=8, pady=5,
        )
        hdr_lbl.pack(fill="x")

        content_f = tk.Frame(section_f, bg=SIDEBAR_BG)
        if open_default:
            content_f.pack(fill="x")

        def _toggle(_e=None, _t=title, _lbl=hdr_lbl, _f=content_f):
            if _f.winfo_ismapped():
                _f.pack_forget()
                _lbl.config(text="▶  " + _t)
            else:
                _f.pack(fill="x")
                _lbl.config(text="▼  " + _t)

        hdr_lbl.bind("<Button-1>", _toggle)
        # Hover-Effekt auf Header
        hdr_lbl.bind("<Enter>", lambda e, l=hdr_lbl: l.config(bg="#112438"))
        hdr_lbl.bind("<Leave>", lambda e, l=hdr_lbl: l.config(bg=SIDEBAR_BG))

        return content_f

    def _build_sidebar(self):
        sb = self._sidebar

        # Logo
        logo_f = tk.Frame(sb, bg=SIDEBAR_BG)
        logo_f.pack(fill="x", padx=16, pady=(20, 16))
        crown = tk.Label(logo_f, text="✦", bg=SIDEBAR_BG, fg=SIDEBAR_ACT,
                         font=("Segoe UI", 20))
        crown.pack(side="left", padx=(0, 8))
        lt = tk.Frame(logo_f, bg=SIDEBAR_BG)
        lt.pack(side="left")
        tk.Label(lt, text="CATENSYS", bg=SIDEBAR_BG, fg="white",
                 font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(lt, text="Claim-Tariff", bg=SIDEBAR_BG, fg=SIDEBAR_ACT,
                 font=("Segoe UI", 7)).pack(anchor="w")

        tk.Frame(sb, bg="#1E3A5F", height=1).pack(fill="x")

        self._nav_refs = {}

        # --- INPUT-DATA (offen beim Start) ---
        inp_f = self._build_collapsible_section(sb, "INPUT-DATA", open_default=True)
        for key, icon, short, _ in _NAV_ITEMS_INPUT:
            self._build_nav_item(inp_f, key, icon, short, indent=12)

        # --- SAP-STAMMDATEN ---
        cust_f = self._build_collapsible_section(sb, "SAP-STAMMDATEN", open_default=False)
        for key, icon, short, _ in _NAV_ITEMS_CUSTOMS:
            self._build_nav_item(cust_f, key, icon, short, indent=12)

        # --- CBP-REPORT ---
        cbp_f = self._build_collapsible_section(sb, "CBP-REPORT", open_default=False)
        for key, icon, short, _ in _NAV_ITEMS_CBP:
            self._build_nav_item(cbp_f, key, icon, short, indent=12)

        # --- PLAUSIBILITÄT ---
        plausi_f = self._build_collapsible_section(sb, "PLAUSIBILITÄT", open_default=False)
        for key, icon, short, _ in _NAV_ITEMS_PLAUSI:
            self._build_nav_item(plausi_f, key, icon, short, indent=12)

        # --- STAMMDATEN ---
        stamm_f = self._build_collapsible_section(sb, "STAMMDATEN", open_default=False)
        for key, icon, short, _ in _NAV_ITEMS_STAMM:
            self._build_nav_item(stamm_f, key, icon, short, indent=12)

        # Abstand
        spacer = tk.Frame(sb, bg=SIDEBAR_BG)
        spacer.pack(fill="both", expand=True)

        # Untere Info-Leiste
        tk.Frame(sb, bg="#1E3A5F", height=1).pack(fill="x")
        bottom = tk.Frame(sb, bg=SIDEBAR_BG)
        bottom.pack(fill="x", padx=16, pady=12)
        tk.Label(bottom, text="●  Verbunden", bg=SIDEBAR_BG, fg="#4CAF50",
                 font=("Segoe UI", 10)).pack(anchor="w")
        tk.Label(bottom, text=os.path.basename(DB_PATH), bg=SIDEBAR_BG, fg=SIDEBAR_FG,
                 font=("Segoe UI", 9), wraplength=180, justify="left").pack(anchor="w", pady=(2, 0))

    def _build_nav_item(self, parent, key, icon, label, indent=0):
        f = tk.Frame(parent, bg=SIDEBAR_BG, cursor="hand2")
        f.pack(fill="x")

        accent_bar = tk.Frame(f, bg=SIDEBAR_BG, width=3)
        accent_bar.pack(side="left", fill="y")

        inner = tk.Frame(f, bg=SIDEBAR_BG)
        inner.pack(side="left", fill="x", expand=True, padx=(10 + indent, 12), pady=8)

        icon_lbl = tk.Label(inner, text=icon, bg=SIDEBAR_BG, fg=SIDEBAR_FG,
                            font=("Segoe UI", 14), width=2, anchor="w")
        icon_lbl.pack(side="left", padx=(0, 8))

        text_lbl = tk.Label(inner, text=label, bg=SIDEBAR_BG, fg=SIDEBAR_FG,
                            font=("Segoe UI", 11), anchor="w")
        text_lbl.pack(side="left", fill="x", expand=True)

        self._nav_refs[key] = {
            "frame": f, "inner": inner,
            "accent": accent_bar, "icon": icon_lbl, "text": text_lbl,
        }

        def _click(e, k=key): self._nav_select(k)
        def _enter(e, k=key):
            if self._current_page != k:
                for w in (f, inner, icon_lbl, text_lbl):
                    w.config(bg="#112438")
        def _leave(e, k=key):
            if self._current_page != k:
                for w in (f, inner, icon_lbl, text_lbl):
                    w.config(bg=SIDEBAR_BG)

        for w in (f, inner, icon_lbl, text_lbl, accent_bar):
            w.bind("<Button-1>", _click)
            w.bind("<Enter>", _enter)
            w.bind("<Leave>", _leave)

    def _nav_select(self, key):
        # Alle Items deaktivieren
        for k, ref in self._nav_refs.items():
            for w in (ref["frame"], ref["inner"], ref["icon"], ref["text"]):
                w.config(bg=SIDEBAR_BG)
            ref["accent"].config(bg=SIDEBAR_BG)
            ref["text"].config(fg=SIDEBAR_FG, font=("Segoe UI", 11))
            ref["icon"].config(fg=SIDEBAR_FG)

        # Aktives Item hervorheben
        r = self._nav_refs[key]
        for w in (r["frame"], r["inner"], r["icon"], r["text"]):
            w.config(bg="#0F2337")
        r["accent"].config(bg=SIDEBAR_ACT)
        r["text"].config(fg="white", font=("Segoe UI", 11, "bold"))
        r["icon"].config(fg=SIDEBAR_ACT)

        # Seite wechseln
        _page_map = {
            "import":      self.tab_import,
            "db":          self.tab_db,
            "tariff":      self.tab_tariff,
            "claim_kunde": self.tab_claim_kunde,
            "sap":         self.tab_sap,
            "sapconn":     self.tab_sapconn,
            "sap_iv":      self.tab_sap_iv,
            "cape":            self.tab_cape,
            "liq":             self.tab_liq,
            "abgleich_manual": self.tab_abgleich_manual,
            "abgleich_db":     self.tab_abgleich_db,
            "zuord_mat_kd":    self.tab_zuord_mat_kd,
            "db_tariff":       self.tab_db_tariff,
            "htsus_claim_cfg": self.tab_htsus_claim_cfg,
            "auswertung":      self.tab_auswertung,
        }
        if self._current_page:
            _page_map[self._current_page].pack_forget()
        _page_map[key].pack(fill="both", expand=True)
        self._current_page = key

        # Header aktualisieren
        titles = {
            "import":      ("Dashboard",            "Import & Status"),
            "db":          ("Datenbank-Inhalt",     "SQLite-Datenbank"),
            "tariff":      ("US.Customs Import",  "Entry Summary Line Tariff Details"),
            "claim_kunde": ("Claim-Report Kunde",   "Antrag-Template für Kunden erstellen & exportieren"),
            "sap":         ("SAP-Daten (MB51)",     "Warenbewegungen aus SAP ASE"),
            "sapconn":     ("SAP-Verbindung",       "Verbindungseinstellungen & Tests"),
            "sap_iv":      ("Eingangsrechnungen (IV)",   "SAP Lieferantenrechnungen abrufen & speichern"),
            "cape":            ("CAPE Entry Summary",   "CBP Report – CAPE Entry Summary Report Import"),
            "liq":             ("CBP Liquidation",      "CBP Report – Liquidation Refund Report Import"),
            "abgleich_manual": ("Abgleich-Manual",       "Plausibilität – Zollbetrag Abgleich manuell"),
            "abgleich_db":     ("Abgleich DB vs. Manual","Plausibilität – Vergleich DB entry_lines vs. Manual-Excel"),
            "zuord_mat_kd":    ("ZuOrd-MAT-KD",            "Stammdaten – Zuordnung Material / Kunde"),
            "db_tariff":       ("db_Tariff",                "Stammdaten – Tariff Datenbank"),
            "auswertung":      ("Auswertung",               "Datenbank-Inhalt – entry_lines (alle Felder)"),
        }
        self._header_title.config(text=titles[key][0])
        self._header_sub.config(text=titles[key][1])

    # -----------------------------------------------------------------------
    # Header-Leiste
    # -----------------------------------------------------------------------
    def _build_header(self):
        hf = self._header_frame
        left = tk.Frame(hf, bg=CARD_BG)
        left.pack(side="left", fill="y", padx=(22, 0))
        self._header_title = tk.Label(
            left, text="Dashboard", bg=CARD_BG, fg="#1A2742",
            font=("Segoe UI", 16, "bold"))
        self._header_title.pack(anchor="w", pady=(13, 0))
        self._header_sub = tk.Label(
            left, text="Import & Status", bg=CARD_BG, fg="#8A9EB5",
            font=("Segoe UI", 10))
        self._header_sub.pack(anchor="w")

    def _build_import_tab(self):
        # ── Scrollbarer Tab-Inhalt ──────────────────────────────────────────
        # Scrollbares Canvas mit korrektem Parent-Kind-Verhältnis:
        # outer-Frame ist Kind von tab_import (nicht vom Canvas), damit
        # der Frame NICHT doppelt gerendert wird (Windows-Tk-Eigenheit).
        _canvas = tk.Canvas(self.tab_import, bg=MAIN_BG, highlightthickness=0)
        _vsb    = ttk.Scrollbar(self.tab_import, orient="vertical",
                                command=_canvas.yview)
        _canvas.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side="right", fill="y")
        _canvas.pack(side="left", fill="both", expand=True)
        # outer ist Kind von tab_import, nicht von _canvas – verhindert
        # die Doppel-Darstellung auf Windows (nativer Fenster-Platz (0,0)
        # plus create_window (0,0) → zwei Kopien desselben Frames)
        outer = tk.Frame(_canvas, bg=MAIN_BG)
        _cwin = _canvas.create_window((0, 0), window=outer, anchor="nw")
        # Innenabstand per Pack-pady/-padx simulieren
        _inner = ttk.Frame(outer, padding=14)
        _inner.pack(fill="both", expand=True)
        def _on_outer_cfg(e):
            _canvas.configure(scrollregion=_canvas.bbox("all"))
        def _on_canvas_cfg(e):
            _canvas.itemconfig(_cwin, width=e.width)
        outer.bind("<Configure>", _on_outer_cfg)
        _canvas.bind("<Configure>", _on_canvas_cfg)
        # Mausrad scrollen
        def _on_mousewheel(e):
            _canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        _canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # Ab hier: outer → _inner als eigentlicher Content-Container
        outer = _inner
        # ────────────────────────────────────────────────────────────────────

        # ── KPI-Karten ──
        kpi_row = tk.Frame(outer, bg=MAIN_BG)
        kpi_row.pack(fill="x", pady=(0, 10))
        kpi_row.columnconfigure((0, 1, 2, 3), weight=1, uniform="kpi")

        self._kpi_files_var   = tk.StringVar(value="–")
        self._kpi_entries_var = tk.StringVar(value="–")
        self._kpi_lines_var   = tk.StringVar(value="–")
        self._kpi_errors_var  = tk.StringVar(value="–")

        for col, (var, label, fgcolor) in enumerate([
            (self._kpi_files_var,   "Dateien verarbeitet",      "#7C3AED"),
            (self._kpi_entries_var, "Entry Summaries (DB)",     "#D97706"),
            (self._kpi_lines_var,   "Warenpositionen (DB)",     "#059669"),
            (self._kpi_errors_var,  "Fehler / keine CBP-Seite", "#DC2626"),
        ]):
            card = tk.Frame(kpi_row, bg=CARD_BG,
                            highlightthickness=1, highlightbackground="#E4E8EF")
            card.grid(row=0, column=col, sticky="nsew",
                      padx=(0, 12 if col < 3 else 0))
            tk.Label(card, textvariable=var, bg=CARD_BG, fg=fgcolor,
                     font=("Segoe UI", 28, "bold")).pack(
                         padx=16, pady=(14, 2), anchor="w")
            tk.Label(card, text=label, bg=CARD_BG, fg="#8A9EB5",
                     font=("Segoe UI", 9)).pack(
                         padx=16, pady=(0, 14), anchor="w")

        # ── Fehler-Liste: PDFs OHNE erkannte CBP-Seite ─────────────────────────
        # Direkt unter den KPI-Karten, sofort sichtbar beim Öffnen des Tabs.
        _err_header = tk.Frame(outer, bg=MAIN_BG)
        _err_header.pack(fill="x", pady=(4, 2))
        tk.Label(
            _err_header,
            text="⚠  Dateien ohne erkannte CBP Form 7501 (Fehler / kein Import)",
            font=("Segoe UI", 10, "bold"),
            bg=MAIN_BG, fg="#DC2626",
        ).pack(side="left")

        no_cbp_columns = ("file_name", "processed_at", "size", "message")
        _no_cbp_wrap = ttk.Frame(outer)
        _no_cbp_wrap.pack(fill="x", pady=(0, 10))
        self.no_cbp_tree = ttk.Treeview(
            _no_cbp_wrap, columns=no_cbp_columns, show="headings", height=3)
        self.no_cbp_tree.heading("file_name",    text="Datei-Name")
        self.no_cbp_tree.heading("processed_at", text="Hochgeladen am")
        self.no_cbp_tree.heading("size",         text="Grösse")
        self.no_cbp_tree.heading("message",      text="Fehlergrund")
        self.no_cbp_tree.column("file_name",    width=280, anchor="w")
        self.no_cbp_tree.column("processed_at", width=155, anchor="w")
        self.no_cbp_tree.column("size",         width=75,  anchor="e")
        self.no_cbp_tree.column("message",      width=500, anchor="w")
        self.no_cbp_tree.tag_configure(
            "Keine CBP-Seite gefunden",
            foreground=self.STATUS_COLORS["Keine CBP-Seite gefunden"])
        _no_cbp_vsb = ttk.Scrollbar(
            _no_cbp_wrap, orient="vertical", command=self.no_cbp_tree.yview)
        self.no_cbp_tree.configure(yscrollcommand=_no_cbp_vsb.set)
        self.no_cbp_tree.pack(side="left", fill="x", expand=True)
        _no_cbp_vsb.pack(side="right", fill="y")
        # ────────────────────────────────────────────────────────────────────

        # Hinweis-Text
        ttk.Label(
            outer,
            text="Ziehen Sie eine oder mehrere PDF-Dateien hierher (Drag & Drop). "
                 "Jede Datei wird sofort gelesen und in die Datenbank uebernommen.",
            wraplength=1100, foreground="#555",
        ).pack(anchor="w", pady=(0, 8))

        self.drop_zone = tk.Label(
            outer, text="PDF-Dateien hier ablegen (hochladen)",
            font=("Segoe UI", 13), relief="flat", borderwidth=0,
            background="#EFF4FA", foreground="#1A2742", height=4,
            highlightthickness=2, highlightbackground="#B8C9DC",
        )
        self.drop_zone.pack(fill="x", pady=(0, 6))
        self.drop_zone.drop_target_register(DND_FILES)
        self.drop_zone.dnd_bind("<<Drop>>", self._on_drop)
        self.drop_zone.dnd_bind("<<DragEnter>>", lambda e: self.drop_zone.configure(
            background="#D4E8FA", highlightbackground=SIDEBAR_ACT))
        self.drop_zone.dnd_bind("<<DragLeave>>", lambda e: self.drop_zone.configure(
            background="#EFF4FA", highlightbackground="#B8C9DC"))

        # ── OCR-Status ──────────────────────────────────────────────────────
        ocr_ok   = getattr(extractor, "_OCR_AVAILABLE", False)
        tess_ok  = getattr(extractor, "_TESS_OK",       False)
        img_ok   = getattr(extractor, "_PDF2IMG_OK", False) or getattr(extractor, "_FITZ_OK", False)
        if ocr_ok:
            ocr_msg = "OCR aktiv (Bild-PDFs werden automatisch per Tesseract gelesen)"
            ocr_fg  = "#166534"
            ocr_bg  = "#DCFCE7"
        elif not tess_ok:
            ocr_msg = ("OCR nicht verfuegbar – Tesseract Binary fehlt.  "
                       "Bitte installieren: https://github.com/UB-Mannheim/tesseract/wiki")
            ocr_fg  = "#991B1B"
            ocr_bg  = "#FEE2E2"
        elif not img_ok:
            ocr_msg = ("OCR nicht verfuegbar – pdf2image oder pymupdf fehlt.  "
                       "Bitte: pip install pdf2image")
            ocr_fg  = "#92400E"
            ocr_bg  = "#FEF3C7"
        else:
            ocr_msg = "OCR-Status unbekannt"
            ocr_fg  = "#555"
            ocr_bg  = "#F3F4F6"
        tk.Label(outer, text=ocr_msg, font=("Segoe UI", 9),
                 foreground=ocr_fg, background=ocr_bg,
                 padx=8, pady=4, anchor="w").pack(fill="x", pady=(0, 6))
        # ────────────────────────────────────────────────────────────────────

        # ── Upload-Fortschrittsanzeige (nur sichtbar während Verarbeitung) ──
        self._upload_progress_frame = ttk.LabelFrame(
            outer, text="Upload-Fortschritt", padding=(10, 6))
        # Wird erst beim Drag&Drop eingeblendet

        # Zeile 1: aktuelle Datei
        _prow1 = ttk.Frame(self._upload_progress_frame)
        _prow1.pack(fill="x", pady=(0, 4))
        ttk.Label(_prow1, text="Aktuelle Datei:").pack(side="left")
        self._upload_file_var = tk.StringVar(value="")
        ttk.Label(_prow1, textvariable=self._upload_file_var,
                  font=("Segoe UI", 9, "bold"), foreground="#1A2742").pack(
                      side="left", padx=(6, 0))
        self._upload_count_var = tk.StringVar(value="")
        ttk.Label(_prow1, textvariable=self._upload_count_var,
                  foreground="#888").pack(side="right")

        # Zeile 2: Fortschrittsbalken
        self._upload_pb = ttk.Progressbar(
            self._upload_progress_frame, orient="horizontal",
            mode="determinate", length=400,
            style="green.Horizontal.TProgressbar")
        self._upload_pb.pack(fill="x", pady=(0, 4))

        # Zeile 3: Log-Tabelle (aktuelle Batch-Ergebnisse)
        log_cols = ("file", "status", "detail")
        self._upload_log_tree = ttk.Treeview(
            self._upload_progress_frame,
            columns=log_cols, show="headings", height=5)
        self._upload_log_tree.heading("file",   text="Datei")
        self._upload_log_tree.heading("status", text="Status")
        self._upload_log_tree.heading("detail", text="Details")
        self._upload_log_tree.column("file",   width=260, anchor="w")
        self._upload_log_tree.column("status", width=160, anchor="w")
        self._upload_log_tree.column("detail", width=480, anchor="w")
        for status, color in self.STATUS_COLORS.items():
            self._upload_log_tree.tag_configure(status, foreground=color)
        self._upload_log_tree.tag_configure("running",
            foreground="#D97706", font=("Segoe UI", 9, "bold"))
        _log_sb = ttk.Scrollbar(self._upload_progress_frame,
                                orient="vertical",
                                command=self._upload_log_tree.yview)
        self._upload_log_tree.configure(yscrollcommand=_log_sb.set)
        _log_tree_frame = ttk.Frame(self._upload_progress_frame)
        _log_tree_frame.pack(fill="x")
        self._upload_log_tree.pack(in_=_log_tree_frame, side="left",
                                   fill="x", expand=True)
        _log_sb.pack(in_=_log_tree_frame, side="right", fill="y")

        self.status_var = tk.StringVar(value="Bereit. Datenbank: " + DB_PATH)
        status_row = ttk.Frame(outer)
        status_row.pack(fill="x", pady=(0, 8))
        ttk.Label(status_row, textvariable=self.status_var, foreground="#1F4E79").pack(side="left")
        ttk.Button(status_row, text="Excel-Export (XLSX) erzeugen und oeffnen",
                   command=self._generate_xlsx_export).pack(side="right")
        ttk.Button(status_row, text="Kunden-Claim-Report erzeugen...",
                   command=self._generate_claim_report).pack(side="right", padx=(0, 8))
        ttk.Button(status_row, text="Datenbank leeren...",
                   command=self._clear_database).pack(side="right", padx=(0, 8))

        # ── PDF-Archiv Aktionsleiste ─────────────────────────────────────────
        archive_row = tk.Frame(outer, bg=MAIN_BG)
        archive_row.pack(fill="x", pady=(0, 6))
        tk.Label(archive_row,
                 text=f"📁  PDF-Archiv: {PDF_ARCHIVE_DIR}",
                 bg=MAIN_BG, fg="#5A6A80", font=("Segoe UI", 9)).pack(side="left")
        tk.Button(
            archive_row,
            text="📂  Archiv-Ordner öffnen",
            command=lambda: os.startfile(PDF_ARCHIVE_DIR),
            bg="#E8F0FE", fg="#1A5B8A", relief="flat", padx=10, pady=3,
            font=("Segoe UI", 9), cursor="hand2",
        ).pack(side="right", padx=(0, 0))
        tk.Button(
            archive_row,
            text="🔄  Auswahl erneut lesen",
            command=self._reprocess_selected_pdfs,
            bg=SIDEBAR_ACT, fg="white", relief="flat", padx=10, pady=3,
            font=("Segoe UI", 9, "bold"), cursor="hand2",
        ).pack(side="right", padx=(0, 8))
        tk.Label(archive_row,
                 text="← Zeile(n) markieren und",
                 bg=MAIN_BG, fg="#8A9EB5", font=("Segoe UI", 9)).pack(side="right", padx=(0, 4))
        # ────────────────────────────────────────────────────────────────────

        # ── PDF-Datei Übersicht (eine Zeile pro PDF, mit Einzel-Button) ─────
        hdr_row = tk.Frame(outer, bg=MAIN_BG)
        hdr_row.pack(fill="x", pady=(4, 2))
        tk.Label(hdr_row, text="PDF-Dateien – Status & Aktionen",
                 bg=MAIN_BG, fg="#1A2742",
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        tk.Button(
            hdr_row, text="↻  Alle aktualisieren",
            command=self._load_pdf_overview,
            bg="#E8F0FE", fg="#1A5B8A", relief="flat",
            padx=10, pady=3, font=("Segoe UI", 9), cursor="hand2",
        ).pack(side="right")

        # Spalten-Header
        HDR_BG = "#1A2742"
        hdr = tk.Frame(outer, bg=HDR_BG)
        hdr.pack(fill="x")
        for txt, w, anc in [
            ("#",           36,  "center"),
            ("Datei-Name", 260,  "w"),
            ("Größe",       68,  "e"),
            ("Importiert",  140, "w"),
            ("Status",      160, "center"),
            ("Positionen",  80,  "center"),
            ("Archiv",      70,  "center"),
            ("Aktion",      110, "center"),
        ]:
            tk.Label(hdr, text=txt, bg=HDR_BG, fg="white",
                     font=("Segoe UI", 9, "bold"),
                     width=0, anchor=anc).pack(
                side="left", padx=(6 if anc == "w" else 2, 2),
                pady=5, ipadx=4)

        # Scrollbarer Körper
        _pdf_ov_outer = tk.Frame(outer, bg=MAIN_BG,
                                  highlightthickness=1,
                                  highlightbackground="#D0D8E4")
        _pdf_ov_outer.pack(fill="both", expand=True, pady=(0, 4))

        _pdf_canvas = tk.Canvas(_pdf_ov_outer, bg=MAIN_BG,
                                 highlightthickness=0)
        _pdf_vsb    = ttk.Scrollbar(_pdf_ov_outer, orient="vertical",
                                     command=_pdf_canvas.yview)
        _pdf_canvas.configure(yscrollcommand=_pdf_vsb.set)
        _pdf_vsb.pack(side="right", fill="y")
        _pdf_canvas.pack(side="left", fill="both", expand=True)

        self._pdf_ov_frame = tk.Frame(_pdf_canvas, bg=MAIN_BG)
        _pdf_cwin = _pdf_canvas.create_window(
            (0, 0), window=self._pdf_ov_frame, anchor="nw")

        def _ov_cfg(e):
            _pdf_canvas.configure(scrollregion=_pdf_canvas.bbox("all"))
        def _ov_can_cfg(e):
            _pdf_canvas.itemconfig(_pdf_cwin, width=e.width)
        self._pdf_ov_frame.bind("<Configure>", _ov_cfg)
        _pdf_canvas.bind("<Configure>", _ov_can_cfg)
        _pdf_canvas.bind("<MouseWheel>",
            lambda e: _pdf_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # "Keine Dateien"-Platzhalter (wird bei Bedarf ein-/ausgeblendet)
        self._pdf_ov_empty = tk.Label(
            self._pdf_ov_frame,
            text="Noch keine PDF-Dateien importiert.\nDateien per Drag & Drop ablegen.",
            bg=MAIN_BG, fg="#8A9EB5", font=("Segoe UI", 10))

        # Historien-Tabelle (kompakt, ausklappbar) ────────────────────────────
        hist_hdr = tk.Frame(outer, bg=MAIN_BG)
        hist_hdr.pack(fill="x", pady=(8, 2))
        tk.Label(hist_hdr, text="Import-Protokoll (alle Versuche)",
                 bg=MAIN_BG, fg="#1A2742",
                 font=("Segoe UI", 10, "bold")).pack(side="left")

        columns = ("file_name", "processed_at", "size", "status", "message", "archived")
        self.tree = ttk.Treeview(outer, columns=columns, show="headings", height=5)
        self.tree.heading("file_name",    text="Datei-Name")
        self.tree.heading("processed_at", text="Datum / Uhrzeit")
        self.tree.heading("size",         text="Größe")
        self.tree.heading("status",       text="Status")
        self.tree.heading("message",      text="Details")
        self.tree.heading("archived",     text="📁 Archiv")
        self.tree.column("file_name",    width=200, anchor="w")
        self.tree.column("processed_at", width=135, anchor="w")
        self.tree.column("size",         width=65,  anchor="e")
        self.tree.column("status",       width=145, anchor="w")
        self.tree.column("message",      width=310, anchor="w")
        self.tree.column("archived",     width=70,  anchor="center")

        for status, color in self.STATUS_COLORS.items():
            self.tree.tag_configure(status, foreground=color)

        _hist_vsb = ttk.Scrollbar(outer, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=_hist_vsb.set)
        _hist_frame = ttk.Frame(outer)
        _hist_frame.pack(fill="x", pady=(0, 4))
        self.tree.pack(in_=_hist_frame, side="left", fill="x", expand=True)
        _hist_vsb.pack(in_=_hist_frame, side="right", fill="y")

    def _on_drop(self, event):
        """Handle PDF drag-and-drop onto the import drop zone."""
        self.drop_zone.configure(
            background="#EFF4FA", highlightbackground="#B8C9DC")
        paths = self.root.tk.splitlist(event.data)
        pdf_files = [p for p in paths
                     if os.path.isfile(p) and p.lower().endswith(".pdf")]
        if not pdf_files:
            messagebox.showwarning(
                "Keine PDF", "Bitte mindestens eine PDF-Datei ablegen.")
            return
        self._process_pdf_batch(pdf_files)

    def _process_pdf_batch(self, pdf_files):
        """Process a list of PDF files in a background thread with progress UI."""
        total = len(pdf_files)

        # Show progress frame (created but not packed during build)
        try:
            self._upload_progress_frame.pack(fill="x", pady=(0, 8))
        except Exception:
            pass  # already packed
        for row in self._upload_log_tree.get_children():
            self._upload_log_tree.delete(row)
        self._upload_pb.configure(maximum=total, value=0)
        self._upload_count_var.set(f"0 / {total}")

        def _run():
            conn = get_db_connection()
            try:
                for idx, path in enumerate(pdf_files, 1):
                    fname = os.path.basename(path)
                    self.root.after(0, lambda f=fname, i=idx: (
                        self._upload_file_var.set(f),
                        self._upload_count_var.set(f"{i} / {total}"),
                        self._upload_pb.configure(value=i - 1),
                    ))
                    # Add "running" row
                    iid = fname + str(idx)
                    self.root.after(0, lambda f=fname, i=iid: (
                        self._upload_log_tree.insert(
                            "", "end", iid=i,
                            values=(f, "Verarbeitung …", ""),
                            tags=("running",))
                    ))

                    status, message = process_one_pdf(conn, path)

                    self.root.after(0, lambda i=iid, s=status, m=message: (
                        self._upload_log_tree.item(i, values=(
                            self._upload_log_tree.item(i, "values")[0], s, m),
                            tags=(s,)),
                        self._upload_pb.configure(
                            value=self._upload_pb["value"] + 1),
                    ))
            finally:
                conn.close()

            self.root.after(0, self._on_batch_done)

        import threading
        threading.Thread(target=_run, daemon=True).start()

    def _on_batch_done(self):
        """Called on the main thread after all PDFs are processed."""
        self._upload_file_var.set("Fertig.")
        self._load_history()   # also updates KPIs and no-CBP list
        self._load_db_tables()
        self.status_var.set(
            f"Upload abgeschlossen – Datenbank: {DB_PATH}")
        # Abgleich-DB automatisch neu berechnen wenn ein Re-Read ausgelöst wurde
        if getattr(self, "_abgleich_pending_refresh", False):
            self._abgleich_pending_refresh = False
            self._run_abgleich_db()

    def _reprocess_one_pdf(self, arch_path, fname):
        """Einzelne PDF aus der Übersichtstabelle erneut lesen."""
        if not os.path.isfile(arch_path):
            messagebox.showerror(
                "Archiv-Datei fehlt",
                f"Die archivierte PDF-Datei wurde nicht gefunden:\n{arch_path}")
            return
        # Alle DB-Einträge dieser Datei löschen für frischen Import
        self.conn.execute("DELETE FROM entries WHERE source_file = ?", (fname,))
        self.conn.execute("DELETE FROM processed_files WHERE file_name = ?", (fname,))
        self.conn.commit()
        self.status_var.set(f"Lese erneut: {fname} …")
        self._process_pdf_batch([arch_path])

    def _reprocess_selected_pdfs(self):
        """Ausgewählte Zeilen erneut lesen – nutzt archivierten PDF-Pfad."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo(
                "Keine Auswahl",
                "Bitte mindestens eine Zeile in der Liste markieren,\n"
                "um die PDF-Datei erneut zu lesen."
            )
            return

        # file_name aus den Tree-Werten auslesen
        file_names = [self.tree.item(iid, "values")[0] for iid in selected]

        # archived_path aus der DB holen (neuesten Eintrag pro Dateiname)
        pdf_paths = []
        missing   = []
        for fname in file_names:
            row = self.conn.execute(
                "SELECT archived_path FROM processed_files "
                "WHERE file_name = ? AND archived_path IS NOT NULL "
                "ORDER BY id DESC LIMIT 1",
                (fname,),
            ).fetchone()
            if row and row[0] and os.path.isfile(row[0]):
                pdf_paths.append(row[0])
            else:
                missing.append(fname)

        if missing:
            messagebox.showwarning(
                "Archiv-Datei nicht gefunden",
                "Folgende PDF-Dateien sind nicht im Archiv vorhanden:\n\n"
                + "\n".join(missing)
                + "\n\nNur vorhandene Dateien werden erneut gelesen."
            )

        if not pdf_paths:
            return

        # Bestehende DB-Einträge für diese Dateien löschen damit frischer Import möglich
        for p in pdf_paths:
            fname = os.path.basename(p)
            self.conn.execute("DELETE FROM entries WHERE source_file = ?", (fname,))
            self.conn.execute("DELETE FROM processed_files WHERE file_name = ?", (fname,))
        self.conn.commit()

        self.status_var.set(f"Lese {len(pdf_paths)} PDF(s) erneut …")
        self._process_pdf_batch(pdf_paths)

    def _build_db_tab(self):
        outer = ttk.Frame(self.tab_db, padding=14)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x")
        ttk.Label(top, text="Datenbank-Inhalt - alle Felder tabellarisch",
                  font=("Segoe UI", 15, "bold")).pack(side="left")
        ttk.Button(top, text="Aktualisieren", command=self._load_db_tables).pack(side="right")
        ttk.Button(top, text="Excel-Export (XLSX) erzeugen und oeffnen",
                   command=self._generate_xlsx_export).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Kunden-Claim-Report erzeugen...",
                   command=self._generate_claim_report).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Datenbank leeren...",
                   command=self._clear_database).pack(side="right", padx=(0, 8))

        ttk.Label(
            outer,
            text="Hinweis: zum Lesen waagrecht scrollen (alle Spalten/Felder der Tabellen sind enthalten).",
            foreground="#555",
        ).pack(anchor="w", pady=(2, 10))

        # --- Summen-Leiste ------------------------------------------------
        sum_frame = ttk.LabelFrame(outer, text="Summen aller importierten Entry Summaries", padding=(10, 6))
        sum_frame.pack(fill="x", pady=(0, 10))
        self._db_sum_vars = {}
        _sum_fields = [
            ("mpf_total",   "MPF Summe"),
            ("duty_total",  "37. Duty"),
            ("tax_total",   "38. Tax"),
            ("other_total", "39. Other"),
            ("grand_total", "40. Grand Total"),
        ]
        for col_idx, (field, label) in enumerate(_sum_fields):
            var = tk.StringVar(value="–")
            self._db_sum_vars[field] = var
            cell = ttk.Frame(sum_frame)
            cell.grid(row=0, column=col_idx, padx=16, pady=4, sticky="w")
            ttk.Label(cell, text=label, font=("Segoe UI", 9),
                      foreground="#555").pack(anchor="w")
            ttk.Label(cell, textvariable=var, font=("Segoe UI", 13, "bold"),
                      foreground="#1F4E79").pack(anchor="w")
        # Zollbetrag-Summe aus entry_lines als zusaetzliche Zelle
        var_zoll = tk.StringVar(value="–")
        self._db_sum_vars["duty_amount"] = var_zoll
        cell_zoll = ttk.Frame(sum_frame)
        cell_zoll.grid(row=0, column=len(_sum_fields), padx=16, pady=4, sticky="w")
        ttk.Label(cell_zoll, text="Zollbetrag (Positionen)", font=("Segoe UI", 9),
                  foreground="#555").pack(anchor="w")
        ttk.Label(cell_zoll, textvariable=var_zoll, font=("Segoe UI", 13, "bold"),
                  foreground="#B45309").pack(anchor="w")
        for i in range(len(_sum_fields) + 1):
            sum_frame.columnconfigure(i, weight=1)
        # ------------------------------------------------------------------

        paned = ttk.Panedwindow(outer, orient="vertical")
        paned.pack(fill="both", expand=True)

        entries_frame = ttk.Labelframe(paned, text="Tabelle 'entries' - Kopfdaten je Entry Summary (Felder 1-26, 35-43)")
        lines_frame = ttk.Labelframe(paned, text="Tabelle 'entry_lines' - Warenpositionen (Felder 27-34)")
        paned.add(entries_frame, weight=1)
        paned.add(lines_frame, weight=1)

        self.entries_columns = list(ENTRIES_LABELS.keys())
        self.lines_columns = list(LINES_LABELS.keys())

        # entries: Filter-Leiste oben, Tabelle darunter
        self.entries_table = ScrollableTable(entries_frame, self.entries_columns, ENTRIES_LABELS, col_width=140)
        _ef = ttk.Frame(entries_frame)
        _ef.pack(fill="x", padx=6, pady=(4, 2))
        self._entries_filter_reapply = _setup_filter_bar(
            _ef, self.entries_table, self.entries_columns, ENTRIES_LABELS)
        self.entries_table.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # lines: Filter-Leiste oben, Tabelle darunter
        _lc = self.lines_columns
        _grp_col = _lc.index("entry_id")    if "entry_id"    in _lc else None
        _sum_col = _lc.index("duty_amount") if "duty_amount" in _lc else None
        _lbl_col = _lc.index("description") if "description" in _lc else None
        self.lines_table = ScrollableTable(
            lines_frame, self.lines_columns, LINES_LABELS, col_width=120,
            subtotal_group_col=_grp_col,
            subtotal_sum_col=_sum_col,
            subtotal_label_col=_lbl_col,
        )
        _lf = ttk.Frame(lines_frame)
        _lf.pack(fill="x", padx=6, pady=(4, 2))
        self._lines_filter_reapply = _setup_filter_bar(
            _lf, self.lines_table, self.lines_columns, LINES_LABELS)
        self.lines_table.pack(fill="both", expand=True, padx=6, pady=(0, 6))


    def _load_db_tables(self):
        try:
            cols_e = ", ".join(self.entries_columns)
            rows_e = self.conn.execute(f"SELECT {cols_e} FROM entries ORDER BY id").fetchall()
            # Zahlenfelder → deutsches Format (kein Tausenderkomma, Dezimalpunkt → Komma)
            _de_fields = {"total_entered_value", "mpf_total", "duty_total",
                          "tax_total", "other_total", "grand_total"}
            _de_idx_e = {i for i, c in enumerate(self.entries_columns) if c in _de_fields}
            if _de_idx_e:
                rows_e = [
                    tuple(_to_de(v) if i in _de_idx_e else v
                          for i, v in enumerate(row))
                    for row in rows_e
                ]
            self.entries_table.set_rows(rows_e)

            # filer_code_entry_no is from entries table, not entry_lines – exclude from SQL
            _el_cols = [c for c in self.lines_columns if c != "filer_code_entry_no"]
            cols_l = ", ".join(_el_cols)
            rows_l = self.conn.execute(
                f"SELECT {cols_l} FROM entry_lines ORDER BY entry_id, id"
            ).fetchall()
            # Build entry_id -> filer_code_entry_no + subtotal label mappings
            import re as _re, os as _os
            def _doc_no(fname):
                nums = _re.findall(r'\d+', _os.path.splitext(fname)[0])
                return max(nums, key=len) if nums else fname
            _eno_rows = self.conn.execute(
                "SELECT id, source_file, filer_code_entry_no FROM entries"
            ).fetchall()
            _eid_to_eno = {r[0]: (r[2] or "") for r in _eno_rows}
            _eid_idx = _el_cols.index("entry_id") if "entry_id" in _el_cols else 1
            # Append filer_code_entry_no as last column so the filter dropdown works
            rows_l = [row + (_eid_to_eno.get(row[_eid_idx], ""),) for row in rows_l]
            # Subtotal labels
            _id_lbl = {
                r[0]: f"{r[2] or '–'}  //  {_doc_no(r[1])}"
                for r in _eno_rows
            }
            self.lines_table._subtotal_id_labels = _id_lbl
            # Zollbetrag → deutsches Format
            duty_idx = self.lines_columns.index("duty_amount") if "duty_amount" in self.lines_columns else -1
            if duty_idx >= 0:
                rows_l = [
                    tuple(_to_de(v) if i == duty_idx else v
                          for i, v in enumerate(row))
                    for row in rows_l
                ]
            self.lines_table.set_rows(rows_l)
        except sqlite3.OperationalError:
            pass
        # Summen-Leiste aktualisieren
        if hasattr(self, "_db_sum_vars"):
            try:
                # REPLACE(..., ',', '') entfernt Tausender-Komma vor CAST
                sums = self.conn.execute(
                    """SELECT
                        COALESCE(SUM(CAST(REPLACE(CAST(mpf_total   AS TEXT),',','') AS REAL)),0),
                        COALESCE(SUM(CAST(REPLACE(CAST(duty_total  AS TEXT),',','') AS REAL)),0),
                        COALESCE(SUM(CAST(REPLACE(CAST(tax_total   AS TEXT),',','') AS REAL)),0),
                        COALESCE(SUM(CAST(REPLACE(CAST(other_total AS TEXT),',','') AS REAL)),0),
                        COALESCE(SUM(CAST(REPLACE(CAST(grand_total AS TEXT),',','') AS REAL)),0)
                    FROM entries"""
                ).fetchone()
                _sum_fields = ["mpf_total", "duty_total", "tax_total", "other_total", "grand_total"]
                for field, val in zip(_sum_fields, sums):
                    self._db_sum_vars[field].set(_to_de(val))
                # Zollbetrag-Summe aus entry_lines
                zoll_sum = self.conn.execute(
                    """SELECT COALESCE(
                        SUM(CAST(REPLACE(CAST(duty_amount AS TEXT),',','') AS REAL)),
                        0) FROM entry_lines"""
                ).fetchone()[0]
                self._db_sum_vars["duty_amount"].set(_to_de(zoll_sum))
            except Exception as _e:
                for var in self._db_sum_vars.values():
                    var.set("?")
        # Filter-Zaehler aktualisieren (behaelt aktiven Filtertext bei)
        if hasattr(self, "_entries_filter_reapply"):
            self._entries_filter_reapply()
        if hasattr(self, "_lines_filter_reapply"):
            self._lines_filter_reapply()
        # KPI-Karten aktualisieren
        if hasattr(self, "_kpi_entries_var"):
            try:
                n_e = self.conn.execute("SELECT COUNT(*) FROM entries").fetchone()[0]
                n_l = self.conn.execute("SELECT COUNT(*) FROM entry_lines").fetchone()[0]
                self._kpi_entries_var.set(str(n_e))
                self._kpi_lines_var.set(str(n_l))
            except Exception:
                pass

    def _load_history(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        rows = self.conn.execute(
            "SELECT file_name, processed_at, file_size_bytes, status, message, archived_path "
            "FROM processed_files ORDER BY id DESC"
        ).fetchall()
        for file_name, processed_at, size, status, message, archived_path in rows:
            self._insert_row(file_name, processed_at, size, status, message, archived_path)
        self._load_no_cbp_list()
        self._load_pdf_overview()
        # KPI-Karten aktualisieren
        if hasattr(self, "_kpi_files_var"):
            errors = sum(1 for *_, s, _, _ap in rows
                         if s in ("Fehler", "Uebersprungen", "Keine CBP-Seite gefunden"))
            self._kpi_files_var.set(str(len(rows)))
            self._kpi_errors_var.set(str(errors))

    # ── Status-Chip-Farben ────────────────────────────────────────────────────
    _STATUS_CHIP = {
        "Importiert":             ("#D1FAE5", "#065F46"),   # grün
        "Bereits vorhanden":      ("#E0F2FE", "#0369A1"),   # blau
        "Keine CBP-Seite gefunden": ("#FEF3C7", "#92400E"), # gelb
        "Fehler":                 ("#FEE2E2", "#991B1B"),   # rot
        "Uebersprungen":          ("#F3F4F6", "#4B5563"),   # grau
    }

    def _load_pdf_overview(self):
        """Baut die scrollbare PDF-Übersichtstabelle mit je einem
        'Erneut lesen'-Button pro Zeile neu auf."""
        frame = self._pdf_ov_frame
        # Alle bestehenden Widgets löschen
        for w in frame.winfo_children():
            w.destroy()

        # _pdf_ov_empty wurde durch destroy vernichtet – neu erstellen
        MAIN_BG = "#F4F6FA"
        self._pdf_ov_empty = tk.Label(
            frame,
            text="Noch keine PDF-Dateien importiert.\nDateien per Drag & Drop ablegen.",
            bg=MAIN_BG, fg="#8A9EB5", font=("Segoe UI", 10))

        # Pro PDF den neuesten Eintrag + Positionen aus DB holen
        rows = self.conn.execute("""
            SELECT pf.file_name,
                   pf.processed_at,
                   pf.file_size_bytes,
                   pf.status,
                   pf.message,
                   pf.archived_path,
                   (SELECT COUNT(*) FROM entry_lines el
                    JOIN entries e ON el.entry_id = e.id
                    WHERE e.source_file = pf.file_name) AS line_count
            FROM processed_files pf
            WHERE pf.id = (
                SELECT MAX(id) FROM processed_files
                WHERE file_name = pf.file_name)
            ORDER BY pf.id DESC
        """).fetchall()

        if not rows:
            self._pdf_ov_empty.pack(pady=20)
            return
        self._pdf_ov_empty.pack_forget()

        ROW_EVEN = "#FFFFFF"
        ROW_ODD  = "#F7F9FC"

        for idx, (fname, proc_at, size, status, message, arch_path, line_cnt) in enumerate(rows):
            bg = ROW_EVEN if idx % 2 == 0 else ROW_ODD

            row_f = tk.Frame(frame, bg=bg,
                              highlightthickness=1,
                              highlightbackground="#E4E8EF")
            row_f.pack(fill="x", pady=(0, 1))

            # ── Zeilennummer ─────────────────────────────────────────────────
            tk.Label(row_f, text=str(idx + 1), bg=bg, fg="#8A9EB5",
                     font=("Segoe UI", 9), width=3,
                     anchor="center").pack(side="left", padx=(6, 2), pady=7)

            # ── PDF-Icon + Dateiname ──────────────────────────────────────────
            fname_disp = fname if len(fname) <= 34 else fname[:31] + "…"
            tk.Label(row_f, text=f"📄 {fname_disp}", bg=bg, fg="#1A2742",
                     font=("Segoe UI", 9, "bold"), width=34,
                     anchor="w").pack(side="left", padx=(2, 4), pady=7)

            # ── Größe ─────────────────────────────────────────────────────────
            tk.Label(row_f, text=human_size(size), bg=bg, fg="#5A6A80",
                     font=("Segoe UI", 9), width=7,
                     anchor="e").pack(side="left", padx=(0, 6))

            # ── Datum ─────────────────────────────────────────────────────────
            ts = proc_at or ""
            try:
                ts = datetime.fromisoformat(proc_at).strftime("%d.%m.%Y  %H:%M")
            except (ValueError, TypeError, AttributeError):
                pass
            tk.Label(row_f, text=ts, bg=bg, fg="#5A6A80",
                     font=("Segoe UI", 9), width=16,
                     anchor="w").pack(side="left", padx=(0, 6))

            # ── Status-Chip ───────────────────────────────────────────────────
            chip_bg, chip_fg = self._STATUS_CHIP.get(
                status, ("#F3F4F6", "#4B5563"))
            chip = tk.Label(row_f, text=status or "–", bg=chip_bg, fg=chip_fg,
                            font=("Segoe UI", 8, "bold"),
                            padx=8, pady=2, relief="flat")
            chip.pack(side="left", padx=(0, 8), pady=4)

            # ── Positionen ────────────────────────────────────────────────────
            pos_fg = "#065F46" if (line_cnt or 0) > 0 else "#DC2626"
            tk.Label(row_f, text=str(line_cnt or 0), bg=bg, fg=pos_fg,
                     font=("Segoe UI", 10, "bold"), width=6,
                     anchor="center").pack(side="left", padx=(0, 6))

            # ── Archiv-Indikator ──────────────────────────────────────────────
            arch_ok = arch_path and os.path.isfile(arch_path)
            arch_txt = "✔" if arch_ok else "✖"
            arch_fg  = "#059669" if arch_ok else "#DC2626"
            tk.Label(row_f, text=arch_txt, bg=bg, fg=arch_fg,
                     font=("Segoe UI", 11), width=5,
                     anchor="center").pack(side="left", padx=(0, 6))

            # ── Erneut-lesen-Button ───────────────────────────────────────────
            can_reread = arch_ok
            if can_reread:
                btn = tk.Button(
                    row_f,
                    text="🔄 Erneut lesen",
                    command=lambda p=arch_path, f=fname: self._reprocess_one_pdf(p, f),
                    bg=SIDEBAR_ACT, fg="white", relief="flat",
                    padx=8, pady=2, font=("Segoe UI", 8, "bold"),
                    cursor="hand2",
                )
            else:
                btn = tk.Button(
                    row_f,
                    text="🚫 kein Archiv",
                    state="disabled",
                    bg="#E5E7EB", fg="#9CA3AF", relief="flat",
                    padx=8, pady=2, font=("Segoe UI", 8),
                )
            btn.pack(side="left", padx=(0, 8), pady=4)

            # Tooltip: Detailmeldung bei Hover
            if message:
                _tip_text = message[:120] + ("…" if len(message) > 120 else "")
                def _enter(e, w=row_f, t=_tip_text, b=bg):
                    w.config(highlightbackground=SIDEBAR_ACT)
                def _leave(e, w=row_f):
                    w.config(highlightbackground="#E4E8EF")
                row_f.bind("<Enter>", _enter)
                row_f.bind("<Leave>", _leave)
                for child in row_f.winfo_children():
                    child.bind("<Enter>", _enter)
                    child.bind("<Leave>", _leave)

    def _load_no_cbp_list(self):
        """Zeigt alle hochgeladenen PDF-Dateien, in denen KEINE CBP-Seite
        (Form 7501 / Entry Summary) gefunden wurde - direkt unter der
        Ablage-Zone, damit Problemfaelle sofort sichtbar sind."""
        for row in self.no_cbp_tree.get_children():
            self.no_cbp_tree.delete(row)
        rows = self.conn.execute(
            "SELECT file_name, processed_at, file_size_bytes, message FROM processed_files "
            "WHERE status = 'Keine CBP-Seite gefunden' ORDER BY id DESC"
        ).fetchall()
        for file_name, processed_at, size, message in rows:
            ts = processed_at
            try:
                ts = datetime.fromisoformat(processed_at).strftime("%d.%m.%Y  %H:%M:%S")
            except (ValueError, TypeError):
                pass
            self.no_cbp_tree.insert(
                "", 0,
                values=(file_name, ts, human_size(size), message or ""),
                tags=("Keine CBP-Seite gefunden",),
            )
        if rows:
            self.no_cbp_tree.configure(height=min(6, max(2, len(rows))))
        else:
            self.no_cbp_tree.insert(
                "", "end",
                values=("- keine -", "", "", "Alle hochgeladenen PDFs enthielten eine CBP Form 7501."),
            )

    def _insert_row(self, file_name, processed_at, size, status, message,
                    archived_path=None):
        ts = processed_at
        try:
            ts = datetime.fromisoformat(processed_at).strftime("%d.%m.%Y  %H:%M:%S")
        except (ValueError, TypeError):
            pass
        archived_icon = "✔ vorhanden" if archived_path and os.path.isfile(archived_path) \
                        else ("⚠ fehlt" if archived_path else "–")
        self.tree.insert(
            "", 0,
            values=(file_name, ts, human_size(size), status, message or "", archived_icon),
            tags=(status,),
        )

    def _generate_xlsx_export(self):
        try:
            generate_xlsx_export(self.conn, EXPORT_XLSX_PATH)
        except ImportError as exc:
            messagebox.showerror("Excel-Export nicht moeglich", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("Excel-Export konnte nicht erzeugt werden", str(exc))
            return
        self.status_var.set(f"Excel-Export erzeugt: {EXPORT_XLSX_PATH}")
        try:
            os.startfile(EXPORT_XLSX_PATH)
        except AttributeError:
            try:
                webbrowser.open(f"file:///{EXPORT_XLSX_PATH.replace(os.sep, '/')}")
            except Exception:
                pass
        except Exception:
            pass

    def _clear_database(self):
        """Leert die Datenbank: entfernt alle Entry-Summary-Datensaetze samt
        Warenpositionen (entries / entry_lines, per ON DELETE CASCADE) sowie
        - auf Wunsch - die Liste bereits verarbeiteter Dateien (processed_files),
        damit PDFs (z. B. nach einer Korrektur des Extraktors) erneut komplett
        eingelesen werden koennen. Zur Sicherheit wird zweimal nachgefragt."""
        n_entries = self.conn.execute("SELECT COUNT(*) FROM entries").fetchone()[0]
        n_lines = self.conn.execute("SELECT COUNT(*) FROM entry_lines").fetchone()[0]
        n_files = self.conn.execute("SELECT COUNT(*) FROM processed_files").fetchone()[0]

        if n_entries == 0 and n_lines == 0 and n_files == 0:
            messagebox.showinfo("Datenbank leeren", "Die Datenbank ist bereits leer.")
            return

        also_history = messagebox.askyesnocancel(
            "Datenbank leeren",
            f"Datenbank: {DB_PATH}\n\n"
            f"Aktuell gespeichert:\n"
            f"  - {n_entries} Entry Summary Dokument(e)\n"
            f"  - {n_lines} Warenposition(en)\n"
            f"  - {n_files} Eintraege in der Datei-Verarbeitungsliste\n\n"
            "Sollen auch die Eintraege der Datei-Verarbeitungsliste geloescht "
            "werden (empfohlen, wenn Sie PDFs anschliessend erneut importieren "
            "moechten - z. B. nach einer Korrektur am Lese-Programm)?\n\n"
            "Ja = Datenbank UND Verarbeitungsliste leeren\n"
            "Nein = nur Entry Summaries / Warenpositionen leeren, "
            "Verarbeitungsliste bleibt erhalten\n"
            "Abbrechen = nichts loeschen",
            parent=self.root,
        )
        if also_history is None:
            return

        if not messagebox.askyesno(
            "Wirklich loeschen?",
            "Dieser Schritt kann NICHT rueckgaengig gemacht werden.\n\n"
            "Soll die Datenbank jetzt wirklich geleert werden?",
            icon="warning",
            parent=self.root,
        ):
            return

        try:
            self.conn.execute("DELETE FROM entries")
            self.conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('entries', 'entry_lines')")
            if also_history:
                self.conn.execute("DELETE FROM processed_files")
                self.conn.execute("DELETE FROM sqlite_sequence WHERE name = 'processed_files'")
            self.conn.commit()
        except Exception as exc:
            messagebox.showerror("Datenbank konnte nicht geleert werden", str(exc))
            return

        self._load_db_tables()
        self._load_history()
        self.status_var.set("Datenbank geleert. Bereit fuer neuen Import. Datenbank: " + DB_PATH)
        messagebox.showinfo(
            "Datenbank geleert",
            "Die Datenbank wurde erfolgreich geleert"
            + (" (inkl. Verarbeitungsliste)." if also_history else " (Verarbeitungsliste blieb erhalten).")
            + "\n\nSie koennen jetzt PDF-Dateien per Drag & Drop neu importieren."
        )

    def _generate_claim_report(self):
        period = simpledialog.askstring(
            "Kunden-Claim-Report",
            "Fuer welchen Monat soll der Claim-Report erzeugt werden?\n"
            "Bitte im Format MM/JJJJ eingeben (z. B. 03/2025):",
            initialvalue=datetime.now().strftime("%m/%Y"),
            parent=self.root,
        )
        if not period:
            return
        period = period.strip()
        try:
            mm, yyyy = period.split("/")
            mm = mm.zfill(2)
            assert len(mm) == 2 and len(yyyy) == 4 and mm.isdigit() and yyyy.isdigit()
            assert 1 <= int(mm) <= 12
        except Exception:
            messagebox.showerror("Ungueltige Eingabe", "Bitte das Datum im Format MM/JJJJ eingeben, z. B. 03/2025.")
            return

        out_path = os.path.join(APP_DIR, f"Catensys_Claim-Report_Kunde_{yyyy}-{mm}.xlsx")
        try:
            anzahl, missing = generate_claim_report(self.conn, CLAIM_TEMPLATE_PATH, out_path, mm, yyyy)
        except (ImportError, FileNotFoundError) as exc:
            messagebox.showerror("Claim-Report nicht moeglich", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("Claim-Report konnte nicht erzeugt werden", str(exc))
            return

        if anzahl == 0:
            messagebox.showwarning(
                "Keine Daten gefunden",
                f"Fuer den Zeitraum {mm}/{yyyy} wurden keine Entry Summaries "
                "(Importdatum) in der Datenbank gefunden.\nEs wurde trotzdem "
                "eine leere Vorlage gespeichert."
            )
        else:
            lines = [f"Kunden-Claim-Report fuer {mm}/{yyyy} mit {anzahl} Warenposition(en) erzeugt:",
                     out_path, "", "Folgende Felder konnten NICHT aus der CBP-7501-Datenbank befuellt"
                                   " werden (in der Datei gelb markiert mit Kommentar):"]
            for col, label, _rownums in missing:
                lines.append(f"  - Spalte {col}: {label}")
            lines.append("")
            lines.append("Diese Angaben stammen aus Lieferanten-/Teile-Stammdaten und muessen "
                         "vor Versand an den Kunden ergaenzt werden.")
            messagebox.showinfo("Kunden-Claim-Report erzeugt", "\n".join(lines))

        self.status_var.set(f"Kunden-Claim-Report erzeugt: {out_path}")
        try:
            os.startfile(out_path)
        except AttributeError:
            try:
                webbrowser.open(f"file:///{out_path.replace(os.sep, '/')}")
            except Exception:
                pass
        except Exception:
            pass

    # -----------------------------------------------------------------------
    # SAP-Verbindungstest-Reiter
    # -----------------------------------------------------------------------
    def _build_sap_conn_tab(self):
        outer = ttk.Frame(self.tab_sapconn, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="SAP-Verbindung konfigurieren & testen",
                  font=("Segoe UI", 15, "bold")).pack(anchor="w")
        ttk.Label(
            outer,
            text="Zugangsdaten werden aus der .env-Datei des SAP-Robot-Workers gelesen. "
                 "'Speichern' schreibt die Aenderungen in die .env-Datei zurueck.",
            foreground="#555", wraplength=1100,
        ).pack(anchor="w", pady=(2, 10))

        # ---- zwei Spalten nebeneinander ----
        cols_frame = ttk.Frame(outer)
        cols_frame.pack(fill="x", pady=(0, 8))
        cols_frame.columnconfigure(0, weight=1)
        cols_frame.columnconfigure(1, weight=1)

        # ---- Linke Spalte: ASE Direkt-DB ----
        db_frame = ttk.LabelFrame(cols_frame, text="ASE Direkt-Datenbankverbindung  (fuer MB51-Abruf)", padding=(12, 8))
        db_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        self._sap_db_vars = {}
        db_fields = [
            ("SAP_DB_HOST",     "Host / IP-Adresse:",        False, 22),
            ("SAP_DB_PORT",     "Port:",                      False, 8),
            ("SAP_DB_NAME",     "Datenbank-Name (SID):",     False, 12),
            ("SAP_DB_SCHEMA",   "Schema (z.B. SEP.SAPSR3):", False, 18),
            ("SAP_DB_USER",     "DB-Login-User (ODBC):",     False, 18),
            ("SAP_DB_PASSWORD", "DB-Passwort (ODBC):",       True,  18),
            ("SAP_DB_DRIVER",   "ODBC-Treiber:",              False, 30),
            ("SAP_CLIENT",      "Mandant (MANDT):",           False, 6),
        ]
        for row_i, (key, label, is_pw, width) in enumerate(db_fields):
            ttk.Label(db_frame, text=label, width=22, anchor="e").grid(
                row=row_i, column=0, sticky="e", pady=3, padx=(0, 6))
            var = tk.StringVar()
            self._sap_db_vars[key] = var
            show = "*" if is_pw else ""
            entry = ttk.Entry(db_frame, textvariable=var, width=width, show=show)
            entry.grid(row=row_i, column=1, sticky="w")
            if is_pw:
                # Passwort-Augen-Knopf
                def _toggle(e=entry, s=show):
                    e.config(show="" if e.cget("show") else "*")
                ttk.Button(db_frame, text="👁", width=3, command=_toggle).grid(
                    row=row_i, column=2, padx=(4, 0))

        # Hinweis: DB-User != SAP-User
        ttk.Label(
            db_frame,
            text="⚠  DB-Login-User ist der Sybase-ASE-ODBC-Datenbankbenutzer\n"
                 "   (z.B. readonly_dev) – NICHT der SAP-Anmeldebenutzer (RFC_COFACE o.ae.)!",
            foreground="#8B4513", font=("Segoe UI", 9),
            justify="left",
        ).grid(row=len(db_fields), column=0, columnspan=3, sticky="w", pady=(8, 0))

        db_btn_frame = ttk.Frame(db_frame)
        db_btn_frame.grid(row=len(db_fields) + 1, column=0, columnspan=3, pady=(8, 0), sticky="w")
        ttk.Button(db_btn_frame, text="TCP-Verbindung testen",
                   command=self._test_db_tcp).pack(side="left")
        ttk.Button(db_btn_frame, text="Datenbank-Login testen",
                   command=self._test_db_login).pack(side="left", padx=(8, 0))

        # ---- Rechte Spalte: SAP RFC ----
        rfc_frame = ttk.LabelFrame(cols_frame, text="SAP RFC-Verbindung  (optional, erfordert pyrfc + NW RFC SDK)", padding=(12, 8))
        rfc_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        self._sap_rfc_vars = {}
        rfc_fields = [
            ("SAP_ASHOST",   "Application Server:",  False, 22),
            ("SAP_SYSNR",    "System-Nr.:",           False, 4),
            ("SAP_CLIENT",   "Mandant:",              False, 6),
            ("SAP_USER",     "RFC-Benutzer:",         False, 18),
            ("SAP_PASSWORD", "Passwort:",             True,  18),
            ("SAP_LANG",     "Sprache:",              False, 4),
        ]
        for row_i, (key, label, is_pw, width) in enumerate(rfc_fields):
            ttk.Label(rfc_frame, text=label, width=22, anchor="e").grid(
                row=row_i, column=0, sticky="e", pady=3, padx=(0, 6))
            var = tk.StringVar()
            self._sap_rfc_vars[key] = var
            show = "*" if is_pw else ""
            entry = ttk.Entry(rfc_frame, textvariable=var, width=width, show=show)
            entry.grid(row=row_i, column=1, sticky="w")
            if is_pw:
                def _toggle_rfc(e=entry):
                    e.config(show="" if e.cget("show") else "*")
                ttk.Button(rfc_frame, text="👁", width=3, command=_toggle_rfc).grid(
                    row=row_i, column=2, padx=(4, 0))

        # Info-Label
        ttk.Label(
            rfc_frame,
            text="Hinweis: RFC-Test benoetigt SAP NW RFC SDK und\n"
                 "'pip install pyrfc' im Worker-Virtualenv.",
            foreground="#777", font=("Segoe UI", 9),
        ).grid(row=len(rfc_fields), column=0, columnspan=3, sticky="w", pady=(6, 0))

        rfc_btn_frame = ttk.Frame(rfc_frame)
        rfc_btn_frame.grid(row=len(rfc_fields) + 1, column=0, columnspan=3, pady=(8, 0), sticky="w")
        ttk.Button(rfc_btn_frame, text="TCP-Verbindung testen",
                   command=self._test_rfc_tcp).pack(side="left")
        ttk.Button(rfc_btn_frame, text="RFC-Login testen (RFC_PING)",
                   command=self._test_rfc_login).pack(side="left", padx=(8, 0))

        # ---- Verbindungsprotokoll ----
        log_frame = ttk.LabelFrame(outer, text="Verbindungsprotokoll", padding=(8, 6))
        log_frame.pack(fill="both", expand=True, pady=(4, 0))

        self._conn_log = tk.Text(
            log_frame, height=12, wrap="word",
            font=("Consolas", 9), background="#1e1e1e", foreground="#d4d4d4",
            insertbackground="#d4d4d4", state="disabled",
        )
        self._conn_log.tag_configure("ok",   foreground="#4ec94e")
        self._conn_log.tag_configure("err",  foreground="#f44747")
        self._conn_log.tag_configure("warn", foreground="#ffcc00")
        self._conn_log.tag_configure("info", foreground="#9cdcfe")
        self._conn_log.tag_configure("head", foreground="#ce9178", font=("Consolas", 9, "bold"))

        log_vsb = ttk.Scrollbar(log_frame, orient="vertical", command=self._conn_log.yview)
        self._conn_log.configure(yscrollcommand=log_vsb.set)
        self._conn_log.pack(side="left", fill="both", expand=True)
        log_vsb.pack(side="right", fill="y")

        # ---- untere Knopfleiste ----
        bottom = ttk.Frame(outer)
        bottom.pack(fill="x", pady=(8, 0))
        ttk.Button(bottom, text="Log leeren",
                   command=self._sap_conn_log_clear).pack(side="left")
        ttk.Button(bottom, text="Einstellungen aus .env laden",
                   command=self._load_sap_conn_env).pack(side="left", padx=(8, 0))
        ttk.Button(bottom, text="Einstellungen in .env speichern",
                   command=self._save_sap_conn_env).pack(side="right")

        # Felder beim Start befuellen
        self._load_sap_conn_env()

    # -- Hilfsmethoden Verbindungslog --
    def _sap_conn_log(self, msg: str, tag: str = ""):
        ts = datetime.now().strftime("%H:%M:%S")
        self._conn_log.config(state="normal")
        self._conn_log.insert("end", f"[{ts}]  {msg}\n", tag or "")
        self._conn_log.see("end")
        self._conn_log.config(state="disabled")
        self.root.update_idletasks()

    def _sap_conn_log_clear(self):
        self._conn_log.config(state="normal")
        self._conn_log.delete("1.0", "end")
        self._conn_log.config(state="disabled")

    # -- sap_settings.ini lesen / schreiben --
    def _load_sap_conn_env(self):
        import configparser
        cfg = configparser.ConfigParser()
        if os.path.exists(SAP_CONFIG_PATH):
            cfg.read(SAP_CONFIG_PATH, encoding="utf-8")
            self._sap_conn_log(
                f"Konfiguration geladen: {os.path.basename(SAP_CONFIG_PATH)}", "info")

            # [DB]-Sektion explizit in _sap_db_vars laden
            db_flat = ({k.upper(): v for k, v in cfg.items("DB")}
                       if cfg.has_section("DB") else {})
            for key, var in self._sap_db_vars.items():
                var.set(db_flat.get(key, ""))

            # [RFC]-Sektion explizit in _sap_rfc_vars laden
            rfc_flat = ({k.upper(): v for k, v in cfg.items("RFC")}
                        if cfg.has_section("RFC") else {})
            for key, var in self._sap_rfc_vars.items():
                # SAP_CLIENT aus [DB] nehmen falls in [RFC] nicht eingetragen
                if key == "SAP_CLIENT" and key not in rfc_flat:
                    var.set(db_flat.get(key, ""))
                else:
                    var.set(rfc_flat.get(key, ""))
        else:
            # Fallback: Worker-.env.example (flach lesen, BEIDE User-Felder klar trennen)
            fallback = os.path.join(SAP_WORKER_DIR, ".env.example")
            values = {}
            if os.path.exists(fallback):
                with open(fallback, encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#") and "=" in line:
                            k, _, v = line.partition("=")
                            values[k.strip()] = v.strip()
                self._sap_conn_log(
                    f"sap_settings.ini nicht gefunden – Fallback: {fallback}", "warn")
            else:
                self._sap_conn_log("Keine Einstellungsdatei gefunden.", "warn")
            # DB-Vars und RFC-Vars getrennt befuellen
            for key, var in self._sap_db_vars.items():
                var.set(values.get(key, ""))
            for key, var in self._sap_rfc_vars.items():
                var.set(values.get(key, ""))

    def _save_sap_conn_env(self):
        import configparser
        cfg = configparser.ConfigParser()
        # Vorhandene Datei lesen (Kommentare gehen verloren, Sektionen bleiben)
        if os.path.exists(SAP_CONFIG_PATH):
            cfg.read(SAP_CONFIG_PATH, encoding="utf-8")

        # Sektionen sicherstellen
        for section in ("DB", "RFC", "GUI"):
            if not cfg.has_section(section):
                cfg.add_section(section)

        db_keys  = list(self._sap_db_vars.keys())
        rfc_keys = list(self._sap_rfc_vars.keys())

        for key, var in self._sap_db_vars.items():
            cfg.set("DB", key, var.get())
        for key, var in self._sap_rfc_vars.items():
            cfg.set("RFC", key, var.get())

        try:
            with open(SAP_CONFIG_PATH, "w", encoding="utf-8") as f:
                f.write("# SAP-Verbindungseinstellungen fuer Catensys Claim-Tariff\n")
                f.write("# Wird vom App-Reiter 'SAP-Verbindung' gelesen und gespeichert.\n")
                f.write("# NIEMALS ins Git-Repository committen!\n\n")
                cfg.write(f)
            self._sap_conn_log(f"Gespeichert: {SAP_CONFIG_PATH}", "ok")
        except Exception as exc:
            self._sap_conn_log(f"Speichern fehlgeschlagen: {exc}", "err")

    # -- TCP-Tests --
    def _tcp_test(self, host: str, port: int, label: str):
        """Fuehrt einen TCP-Verbindungstest durch (blockierend, in Thread aufrufen)."""
        import socket
        self._sap_conn_log(f"TCP  →  {host}:{port}  ({label}) …", "info")
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(5)
            s.connect((host, port))
            s.close()
            self._sap_conn_log(f"OK   {host}:{port}  ({label}) erreichbar.", "ok")
            return True
        except Exception as exc:
            self._sap_conn_log(f"FAIL {host}:{port}  ({label}): {exc}", "err")
            return False

    def _test_db_tcp(self):
        host = self._sap_db_vars.get("SAP_DB_HOST", tk.StringVar()).get().strip()
        port_s = self._sap_db_vars.get("SAP_DB_PORT", tk.StringVar()).get().strip()
        if not host:
            self._sap_conn_log("Bitte zuerst SAP_DB_HOST eintragen.", "warn"); return
        try:
            port = int(port_s or "4901")
        except ValueError:
            self._sap_conn_log("Ungueltige Port-Angabe.", "err"); return
        self._sap_conn_log("─── ASE TCP-Test ───────────────────────────────────", "head")
        threading.Thread(
            target=lambda: self._tcp_test(host, port, "Sybase ASE"), daemon=True).start()

    def _test_rfc_tcp(self):
        host  = self._sap_rfc_vars.get("SAP_ASHOST", tk.StringVar()).get().strip()
        sysnr = self._sap_rfc_vars.get("SAP_SYSNR",  tk.StringVar()).get().strip() or "05"
        if not host:
            self._sap_conn_log("Bitte zuerst SAP_ASHOST eintragen.", "warn"); return
        self._sap_conn_log("─── RFC TCP-Test ───────────────────────────────────", "head")
        def _run():
            rfc_port = int(f"33{sysnr.zfill(2)}")
            gw_port  = int(f"32{sysnr.zfill(2)}")
            self._tcp_test(host, rfc_port, "SAP Dispatcher (RFC)")
            self._tcp_test(host, gw_port,  "SAP Gateway")
        threading.Thread(target=_run, daemon=True).start()

    # -- DB-Login-Test --
    def _test_db_login(self):
        self._sap_conn_log("─── ASE Datenbank-Login-Test ───────────────────────", "head")
        python_exe = SAP_WORKER_PYTHON if os.path.exists(SAP_WORKER_PYTHON) else sys.executable
        if not os.path.exists(os.path.join(SAP_WORKER_DIR, "sap_db.py")):
            self._sap_conn_log(
                f"sap_db.py nicht gefunden unter {SAP_WORKER_DIR}", "err"); return

        env_override = {
            **os.environ,
            "PYTHONIOENCODING":  "utf-8",
            "PYTHONUTF8":        "1",
            "SAP_DB_HOST":       self._sap_db_vars["SAP_DB_HOST"].get(),
            "SAP_DB_PORT":       self._sap_db_vars["SAP_DB_PORT"].get(),
            "SAP_DB_NAME":       self._sap_db_vars["SAP_DB_NAME"].get(),
            "SAP_DB_SCHEMA":     self._sap_db_vars["SAP_DB_SCHEMA"].get(),
            "SAP_DB_USER":       self._sap_db_vars["SAP_DB_USER"].get(),
            "SAP_DB_PASSWORD":   self._sap_db_vars["SAP_DB_PASSWORD"].get(),
            "SAP_DB_DRIVER":     self._sap_db_vars["SAP_DB_DRIVER"].get(),
            "SAP_CLIENT":        self._sap_db_vars["SAP_CLIENT"].get(),
        }

        code = (
            "import sys, os; sys.path.insert(0, '.');\n"
            "from sap_db import ping, get_connection;\n"
            "print(ping());\n"
            "conn = get_connection();\n"
            "cur = conn.cursor();\n"
            "mandt = os.environ.get('SAP_CLIENT','600');\n"
            "schema = os.environ.get('SAP_DB_SCHEMA','SAPSR3');\n"
            "pfx = schema + '.' if schema else '';\n"
            "for tbl in ('MKPF', 'MSEG', 'T001', 'MARA'):\n"
            "    try:\n"
            "        cur.execute(f'SELECT count(*) FROM {pfx}{tbl} WHERE MANDT = ?', mandt);\n"
            "        n = cur.fetchone()[0];\n"
            "        print(f'  {tbl:>8}: {n:,} Zeilen (Mandant {mandt})');\n"
            "    except Exception as e:\n"
            "        print(f'  {tbl:>8}: FEHLER – {e}');\n"
            "conn.close();\n"
            "print('Login OK.');\n"
        )

        def _run():
            import subprocess
            try:
                proc = subprocess.Popen(
                    [python_exe, "-c", code],
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, cwd=SAP_WORKER_DIR, env=env_override,
                )
                for line in proc.stdout:
                    line = line.rstrip()
                    tag = "ok" if "OK" in line or "Zeilen" in line else (
                          "err" if "FEHLER" in line or "Error" in line or "error" in line else "")
                    self.root.after(0, self._sap_conn_log, line, tag)
                proc.wait()
                if proc.returncode != 0:
                    self.root.after(0, self._sap_conn_log,
                                    f"Prozess beendet mit Code {proc.returncode}.", "err")
            except Exception as exc:
                self.root.after(0, self._sap_conn_log, f"Fehler: {exc}", "err")

        threading.Thread(target=_run, daemon=True).start()

    # -- RFC-Login-Test --
    def _test_rfc_login(self):
        self._sap_conn_log("─── SAP RFC-Login-Test (RFC_PING) ──────────────────", "head")
        script = os.path.join(SAP_WORKER_DIR, "test_sap_connection.py")
        if not os.path.exists(script):
            self._sap_conn_log(
                f"test_sap_connection.py nicht gefunden unter {SAP_WORKER_DIR}", "err"); return
        python_exe = SAP_WORKER_PYTHON if os.path.exists(SAP_WORKER_PYTHON) else sys.executable

        env_override = {
            **os.environ,
            "PYTHONIOENCODING": "utf-8",
            "PYTHONUTF8":       "1",
            "SAP_ASHOST":       self._sap_rfc_vars["SAP_ASHOST"].get(),
            "SAP_SYSNR":        self._sap_rfc_vars["SAP_SYSNR"].get(),
            "SAP_CLIENT":       self._sap_rfc_vars["SAP_CLIENT"].get(),
            "SAP_USER":         self._sap_rfc_vars["SAP_USER"].get(),
            "SAP_PASSWORD":     self._sap_rfc_vars["SAP_PASSWORD"].get(),
            "SAP_LANG":         self._sap_rfc_vars["SAP_LANG"].get(),
        }

        def _run():
            import subprocess
            try:
                proc = subprocess.Popen(
                    [python_exe, script],
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, cwd=SAP_WORKER_DIR, env=env_override,
                )
                for line in proc.stdout:
                    line = line.rstrip()
                    tag = ("ok"   if "[OK]" in line else
                           "err"  if "[FEHLER]" in line else
                           "warn" if "[WARN]" in line else "")
                    self.root.after(0, self._sap_conn_log, line, tag)
                proc.wait()
                if proc.returncode != 0:
                    self.root.after(0, self._sap_conn_log,
                                    f"Prozess beendet mit Code {proc.returncode}.", "err")
            except Exception as exc:
                self.root.after(0, self._sap_conn_log, f"Fehler: {exc}", "err")

        threading.Thread(target=_run, daemon=True).start()

    # -----------------------------------------------------------------------
    # US.Customs-Reiter: Entry Summary Line Tariff Details
    # -----------------------------------------------------------------------
    def _build_tariff_tab(self):
        outer = ttk.Frame(self.tab_tariff, padding=14)
        outer.pack(fill="both", expand=True)

        # ── KPI-Karten ──
        kpi_row = tk.Frame(outer, bg=MAIN_BG)
        kpi_row.pack(fill="x", pady=(0, 10))
        kpi_row.columnconfigure((0, 1, 2, 3), weight=1, uniform="tkpi")

        self._tc_kpi_rows_var    = tk.StringVar(value="–")
        self._tc_kpi_entries_var = tk.StringVar(value="–")
        self._tc_kpi_value_var   = tk.StringVar(value="–")
        self._tc_kpi_duty_var    = tk.StringVar(value="–")

        for col, (var, label, fgcolor) in enumerate([
            (self._tc_kpi_rows_var,    "Tariff-Zeilen (DB)",     "#7C3AED"),
            (self._tc_kpi_entries_var, "Entry Summaries",        "#D97706"),
            (self._tc_kpi_value_var,   "Warenwert gesamt ($)",   "#059669"),
            (self._tc_kpi_duty_var,    "Zollbetrag gesamt ($)",  "#DC2626"),
        ]):
            card = tk.Frame(kpi_row, bg=CARD_BG,
                            highlightthickness=1, highlightbackground="#E4E8EF")
            card.grid(row=0, column=col, sticky="nsew",
                      padx=(0, 12 if col < 3 else 0))
            tk.Label(card, textvariable=var, bg=CARD_BG, fg=fgcolor,
                     font=("Segoe UI", 24, "bold")).pack(
                         padx=16, pady=(14, 2), anchor="w")
            tk.Label(card, text=label, bg=CARD_BG, fg="#8A9EB5",
                     font=("Segoe UI", 9)).pack(padx=16, pady=(0, 14), anchor="w")

        # ── Drag-&-Drop-Zone ──
        self._tariff_drop_zone = tk.Label(
            outer,
            text="Excel-Datei (Entry Summary Line Tariff Details) hier ablegen",
            font=("Segoe UI", 12), background="#EFF4FA", foreground="#1A2742",
            height=3, highlightthickness=2, highlightbackground="#B8C9DC",
        )
        self._tariff_drop_zone.pack(fill="x", pady=(0, 6))
        self._tariff_drop_zone.drop_target_register(DND_FILES)
        self._tariff_drop_zone.dnd_bind("<<Drop>>",      self._on_tariff_drop)
        self._tariff_drop_zone.dnd_bind("<<DragEnter>>", lambda e: self._tariff_drop_zone.configure(
            background="#D4E8FA", highlightbackground=SIDEBAR_ACT))
        self._tariff_drop_zone.dnd_bind("<<DragLeave>>", lambda e: self._tariff_drop_zone.configure(
            background="#EFF4FA", highlightbackground="#B8C9DC"))

        # ── Schaltflächen & Status ──
        btn_f = ttk.Frame(outer)
        btn_f.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_f, text="Datei auswählen & importieren",
                   command=self._pick_tariff_file).pack(side="left")
        ttk.Button(btn_f, text="Aktualisieren",
                   command=self._load_tariff_data).pack(side="left", padx=(8, 0))
        ttk.Button(btn_f, text="Alle Daten löschen...",
                   command=self._clear_tariff_data).pack(side="left", padx=(8, 0))
        self._tc_status_var = tk.StringVar(value="Noch keine Daten importiert.")
        ttk.Label(btn_f, textvariable=self._tc_status_var,
                  foreground="#1F4E79").pack(side="right")

        # ── Tabelle mit Filter ──
        tbl_frame = ttk.LabelFrame(
            outer, text="Importierte US.Customs-Daten (Entry Summary Line Tariff Details)")
        tbl_frame.pack(fill="both", expand=True)

        self.tariff_table = ScrollableTable(
            tbl_frame, TARIFF_CLAIM_COLUMNS, TARIFF_CLAIM_LABELS, col_width=120)
        _tf = ttk.Frame(tbl_frame)
        _tf.pack(fill="x", padx=6, pady=(4, 2))
        self._tariff_filter_reapply = _setup_filter_bar(
            _tf, self.tariff_table, TARIFF_CLAIM_COLUMNS, TARIFF_CLAIM_LABELS)
        self.tariff_table.pack(fill="both", expand=True, padx=6, pady=(0, 6))

    def _load_tariff_data(self):
        """Laedt alle tariff_claim_lines aus SQLite und zeigt sie in der Tabelle."""
        try:
            cols = ", ".join(TARIFF_CLAIM_COLUMNS)
            rows = self.conn.execute(
                f"SELECT {cols} FROM tariff_claim_lines ORDER BY id").fetchall()
            self.tariff_table.set_rows(rows)
            if hasattr(self, "_tariff_filter_reapply"):
                self._tariff_filter_reapply()
            # KPI-Karten
            if hasattr(self, "_tc_kpi_rows_var"):
                n = len(rows)
                unique_es = self.conn.execute(
                    "SELECT COUNT(DISTINCT entry_summary_number) FROM tariff_claim_lines"
                ).fetchone()[0]
                total_val = self.conn.execute(
                    "SELECT COALESCE(SUM(goods_value),0) FROM tariff_claim_lines"
                ).fetchone()[0]
                total_duty = self.conn.execute(
                    "SELECT COALESCE(SUM(duty_amount),0) FROM tariff_claim_lines"
                ).fetchone()[0]
                self._tc_kpi_rows_var.set(f"{n:,}".replace(",", "."))
                self._tc_kpi_entries_var.set(str(unique_es))
                self._tc_kpi_value_var.set(f"${total_val:,.0f}".replace(",", "."))
                self._tc_kpi_duty_var.set(f"${total_duty:,.0f}".replace(",", "."))
            if rows:
                self._tc_status_var.set(
                    f"{len(rows):,} Zeilen in der Datenbank.".replace(",", "."))
        except Exception as exc:
            if hasattr(self, "_tc_status_var"):
                self._tc_status_var.set(f"Fehler beim Laden: {exc}")

    def _import_tariff_excel(self, path):
        """Liest eine Entry-Summary-Tariff-Excel-Datei und importiert sie in die DB."""
        if openpyxl is None:
            messagebox.showerror("openpyxl fehlt",
                                 "Bitte 'pip install openpyxl' ausfuehren.")
            return
        fname = os.path.basename(path)
        self._tc_status_var.set(f"Importiere {fname} ...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        def _run():
            try:
                from datetime import datetime as _dt
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                next(rows_iter, None)           # Kopfzeile überspringen
                now = _dt.now().isoformat(timespec="seconds")
                batch = []
                for row in rows_iter:
                    def _s(v):
                        if v is None: return None
                        if hasattr(v, "strftime"): return v.strftime("%d.%m.%Y")
                        return v
                    batch.append((
                        fname, now,
                        _s(row[0]),  _s(row[1]),  _s(row[2]),  _s(row[3]),
                        _s(row[4]),  _s(row[5]),  _s(row[6]),
                        row[7],      _s(row[8]),
                        _s(row[9]),  _s(row[10]), _s(row[11]), _s(row[12]),
                        _s(row[13]), _s(row[14]), _s(row[15]), _s(row[16]),
                        row[17],     _s(row[18]),
                        row[19],     _s(row[20]),
                        row[21],     _s(row[22]),
                        row[23],     _s(row[24]),
                        row[25],     row[26],
                    ))
                wb.close()
                conn2 = get_db_connection()
                conn2.executemany(
                    "INSERT INTO tariff_claim_lines "
                    "(source_file, imported_at, "
                    " entry_summary_number, entry_type_code, importer_number, "
                    " port_of_entry_code, entry_date, entry_summary_date, "
                    " initial_create_date, line_number, review_team_number, "
                    " country_of_origin, country_of_export, manufacturer_id, "
                    " foreign_exporter_id, line_spi_code, standard_visa_number, "
                    " textile_category_code, textile_category, tariff_ordinal_number, "
                    " hts_number, quantity_1, uom_1, quantity_2, uom_2, "
                    " quantity_3, uom_3, goods_value, duty_amount) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    batch,
                )
                conn2.commit()
                conn2.close()
                return len(batch), None
            except Exception as exc:
                return 0, str(exc)

        def _done(n, err):
            self.root.config(cursor="")
            if err:
                messagebox.showerror("Import fehlgeschlagen", err)
                self._tc_status_var.set(f"Fehler: {err[:120]}")
            else:
                self._tc_status_var.set(
                    f"{n:,} Zeilen aus '{fname}' importiert.".replace(",", "."))
                self._load_tariff_data()

        threading.Thread(
            target=lambda: self.root.after(0, _done, *_run()),
            daemon=True,
        ).start()

    def _on_tariff_drop(self, event):
        self._tariff_drop_zone.configure(
            background="#EFF4FA", highlightbackground="#B8C9DC")
        paths = self.root.tk.splitlist(event.data)
        xlsx_files = [p for p in paths
                      if os.path.isfile(p) and p.lower().endswith(".xlsx")]
        if not xlsx_files:
            messagebox.showwarning("Kein Excel",
                                   "Bitte eine .xlsx-Datei ablegen.")
            return
        self._import_tariff_excel(xlsx_files[0])

    def _pick_tariff_file(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Entry Summary Tariff Excel auswählen",
            filetypes=[("Excel-Dateien", "*.xlsx *.xlsm"), ("Alle Dateien", "*.*")],
        )
        if path:
            self._import_tariff_excel(path)

    def _clear_tariff_data(self):
        n = self.conn.execute(
            "SELECT COUNT(*) FROM tariff_claim_lines").fetchone()[0]
        if n == 0:
            messagebox.showinfo("Leer", "Keine US.Customs-Daten vorhanden.")
            return
        if not messagebox.askyesno(
            "Wirklich löschen?",
            f"Alle {n:,} US.Customs-Zeilen werden unwiderruflich gelöscht.\n\nFortfahren?".replace(",","."),
            icon="warning", parent=self.root,
        ):
            return
        self.conn.execute("DELETE FROM tariff_claim_lines")
        self.conn.execute(
            "DELETE FROM sqlite_sequence WHERE name='tariff_claim_lines'")
        self.conn.commit()
        self._load_tariff_data()
        self._tc_status_var.set("Daten gelöscht.")

    # -----------------------------------------------------------------------
    # Claim-Kunde Report
    # -----------------------------------------------------------------------
    # Default-Einstellungs-Keys
    _CK_KEYS = [
        "vhm", "commodity_code", "vehicle_line", "production_fcsd",
        "tariff_type", "claim_start", "claim_end",
        "ford_prefix", "ford_suffix", "raw_material",
        "ami_linked", "index_code", "components_per_part",
    ]

    def _build_claim_kunde_tab(self):
        outer = ttk.Frame(self.tab_claim_kunde, padding=14)
        outer.pack(fill="both", expand=True)

        # ── Einstellungsrahmen (globale Felder) — standardmässig eingeklappt ──
        _cfg_hdr = tk.Frame(outer, bg=MAIN_BG)
        _cfg_hdr.pack(fill="x", pady=(0, 2))
        self._ck_cfg_open = tk.BooleanVar(value=False)

        def _toggle_cfg():
            if self._ck_cfg_open.get():
                cfg.pack(fill="x", pady=(0, 8))
                _cfg_btn.config(text="▲  Globale Einstellungen  (To be filled by Supplier)")
            else:
                cfg.pack_forget()
                _cfg_btn.config(text="▶  Globale Einstellungen  (To be filled by Supplier)")

        _cfg_btn = tk.Button(
            _cfg_hdr,
            text="▶  Globale Einstellungen  (To be filled by Supplier)",
            bg="#E8EDF4", fg="#1F4E79",
            font=("Segoe UI", 9, "bold"),
            relief="flat", anchor="w", padx=8, pady=3,
            cursor="hand2",
            command=lambda: [self._ck_cfg_open.set(not self._ck_cfg_open.get()), _toggle_cfg()])
        _cfg_btn.pack(fill="x")

        cfg = ttk.LabelFrame(outer, text="")

        self._ck_vars = {}
        fields = [
            ("vhm",               "VHM",                     0, 0),
            ("commodity_code",    "Commodity Code",           0, 2),
            ("vehicle_line",      "Vehicle Line",             0, 4),
            ("production_fcsd",   "Production / FCSD",        0, 6),
            ("tariff_type",       "Tariff Type",              1, 0),
            ("claim_start",       "Claim Start (mm/dd/yyyy)", 1, 2),
            ("claim_end",         "Claim End (mm/dd/yyyy)",   1, 4),
            ("ford_prefix",       "Ford Part Prefix",         2, 0),
            ("ford_suffix",       "Ford Part Suffix",         2, 2),
            ("raw_material",      "Raw Material",             2, 4),
            ("components_per_part","Components Per Part",     2, 6),
            ("ami_linked",        "AMI Linked (Yes/No)",      3, 0),
            ("index_code",        "Index Code",               3, 2),
        ]
        for key, label, r, c in fields:
            v = tk.StringVar()
            self._ck_vars[key] = v
            ttk.Label(cfg, text=label + ":").grid(
                row=r, column=c, sticky="w", padx=(10 if c == 0 else 4, 2), pady=4)
            ttk.Entry(cfg, textvariable=v, width=18).grid(
                row=r, column=c + 1, sticky="ew", padx=(0, 6), pady=4)
        for c in range(0, 8, 2):
            cfg.columnconfigure(c + 1, weight=1)

        # ── Datenquelle & Verknüpfung — standardmässig eingeklappt ───────────
        _src_hdr = tk.Frame(outer, bg=MAIN_BG)
        _src_hdr.pack(fill="x", pady=(0, 2))
        self._ck_src_open = tk.BooleanVar(value=False)

        def _toggle_src():
            if self._ck_src_open.get():
                src_frame.pack(fill="x", pady=(0, 8))
                _src_btn.config(text="▲  Datenquelle & Verknüpfung")
            else:
                src_frame.pack_forget()
                _src_btn.config(text="▶  Datenquelle & Verknüpfung")

        _src_btn = tk.Button(
            _src_hdr,
            text="▶  Datenquelle & Verknüpfung",
            bg="#E8EDF4", fg="#1F4E79",
            font=("Segoe UI", 9, "bold"),
            relief="flat", anchor="w", padx=8, pady=3,
            cursor="hand2",
            command=lambda: [self._ck_src_open.set(not self._ck_src_open.get()), _toggle_src()])
        _src_btn.pack(fill="x")

        src_frame = ttk.LabelFrame(outer, text="", padding=(10, 6))
        src_frame.columnconfigure(1, weight=2)
        src_frame.columnconfigure(3, weight=3)

        # ── Spalte A: CBP-Einträge ──
        ttk.Label(src_frame, text="CBP-Einträge:",
                  font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, sticky="nw", padx=(0, 4))
        entry_box_frame = ttk.Frame(src_frame)
        entry_box_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=(0, 8))
        self._ck_entry_lb = tk.Listbox(
            entry_box_frame, selectmode="extended", height=5,
            font=("Consolas", 9), exportselection=False,
            bg="#F7F9FC", activestyle="none")
        _lb_sb = ttk.Scrollbar(entry_box_frame, orient="vertical",
                               command=self._ck_entry_lb.yview)
        self._ck_entry_lb.configure(yscrollcommand=_lb_sb.set)
        self._ck_entry_lb.pack(side="left", fill="both", expand=True)
        _lb_sb.pack(side="right", fill="y")

        ebtn = ttk.Frame(src_frame)
        ebtn.grid(row=2, column=0, columnspan=2, sticky="w", pady=(2, 0))
        ttk.Button(ebtn, text="Alle", width=6,
                   command=lambda: self._ck_entry_lb.selection_set(0, "end")).pack(
                       side="left", padx=(0, 4))
        ttk.Button(ebtn, text="Keine", width=6,
                   command=lambda: self._ck_entry_lb.selection_clear(0, "end")).pack(
                       side="left")

        # ── Spalte B: HTS-Code-Auswahl ──
        ttk.Label(src_frame, text="HTS-Codes (zu verrechnen):",
                  font=("Segoe UI", 9, "bold")).grid(
            row=0, column=2, sticky="nw", padx=(0, 4))
        hts_box_frame = ttk.Frame(src_frame)
        hts_box_frame.grid(row=1, column=2, columnspan=2, sticky="nsew", padx=(0, 8))
        self._ck_hts_lb = tk.Listbox(
            hts_box_frame, selectmode="extended", height=5,
            font=("Consolas", 9), exportselection=False,
            bg="#F7F9FC", activestyle="none")
        _hts_sb = ttk.Scrollbar(hts_box_frame, orient="vertical",
                                command=self._ck_hts_lb.yview)
        self._ck_hts_lb.configure(yscrollcommand=_hts_sb.set)
        self._ck_hts_lb.pack(side="left", fill="both", expand=True)
        _hts_sb.pack(side="right", fill="y")

        hbtn = ttk.Frame(src_frame)
        hbtn.grid(row=2, column=2, columnspan=2, sticky="w", pady=(2, 0))
        ttk.Button(hbtn, text="Alle", width=6,
                   command=lambda: self._ck_hts_lb.selection_set(0, "end")).pack(
                       side="left", padx=(0, 4))
        ttk.Button(hbtn, text="Keine", width=6,
                   command=lambda: self._ck_hts_lb.selection_clear(0, "end")).pack(
                       side="left")

        # ── Spalte C: Optionen ──
        opt_frame = ttk.Frame(src_frame)
        opt_frame.grid(row=0, column=4, rowspan=3, sticky="nw", padx=(8, 0))

        ttk.Label(opt_frame, text="SAP-Join-Schlüssel:",
                  font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self._ck_sap_join_var = tk.StringVar(value="Referenzbeleg")
        ttk.Combobox(opt_frame, textvariable=self._ck_sap_join_var,
                     values=["Referenzbeleg", "Belegnummer"],
                     state="readonly", width=16).pack(anchor="w", pady=(2, 12))

        ttk.Button(opt_frame, text="↺  Aktualisieren",
                   command=self._ck_refresh_entry_list).pack(anchor="w", fill="x")

        # Einträge + HTS beim ersten Öffnen laden
        self.root.after(200, self._ck_refresh_entry_list)

        btn_row = ttk.Frame(outer)
        btn_row.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_row, text="Einstellungen speichern",
                   command=self._save_claim_settings).pack(side="left")
        ttk.Button(btn_row, text="↺  Daten aktualisieren",
                   command=self._ck_refresh_and_reload).pack(side="left", padx=(8, 0))
        ttk.Button(btn_row, text="▶  Vorschau generieren",
                   command=self._load_claim_preview).pack(side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Excel-Report exportieren",
                   command=self._export_claim_excel).pack(side="left", padx=(8, 0))
        # Nullwert-Filter + Feldauswahl
        self._ck_hide_zeros = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            btn_row,
            text="Nur Zeilen mit Wert ≠ 0:",
            variable=self._ck_hide_zeros,
            command=self._ck_apply_zero_filter,
        ).pack(side="left", padx=(16, 0))
        # Dropdown: Alle Zahlenfelder ODER ein bestimmtes Feld
        _nz_options = ["— Alle Zahlenfelder —"] + [
            CLAIM_PREVIEW_LABELS[CLAIM_PREVIEW_COLS[i]]
            for i in self._CK_NUMERIC_COLS
        ]
        self._ck_zero_field_var = tk.StringVar(value=_nz_options[0])
        self._ck_zero_field_cb = ttk.Combobox(
            btn_row,
            textvariable=self._ck_zero_field_var,
            values=_nz_options,
            state="readonly",
            width=22,
        )
        self._ck_zero_field_cb.pack(side="left", padx=(4, 0))
        self._ck_zero_field_cb.bind(
            "<<ComboboxSelected>>", lambda _: self._ck_apply_zero_filter())
        self._ck_status_var = tk.StringVar(value="")
        ttk.Label(btn_row, textvariable=self._ck_status_var,
                  foreground="#1F4E79").pack(side="right")

        # ── Ergänzungsfelder (Zeilenbezogen) ─────────────────────────────────
        supp = ttk.LabelFrame(outer, text="Zeilenbezogene Ergänzung (Doppelklick auf Zeile zum Bearbeiten)")
        supp.pack(fill="x", pady=(0, 8))
        ttk.Label(supp, text="Felder Supplier Invoice No, Invoice Date, Ford Part Base,"
                  " Supplier Part Nr., T2 Part Nr. und Weight können"
                  " per Doppelklick auf eine Vorschau-Zeile ergänzt werden.",
                  foreground="#555", wraplength=1000).pack(
                      anchor="w", padx=8, pady=6)

        # ── KPI-Karten ────────────────────────────────────────────────────────
        kpi_row = tk.Frame(outer, bg=MAIN_BG)
        kpi_row.pack(fill="x", pady=(0, 8))
        kpi_row.columnconfigure((0, 1, 2, 3, 4), weight=1, uniform="ckkpi")

        self._ck_kpi_rows_var   = tk.StringVar(value="–")
        self._ck_kpi_entries_var = tk.StringVar(value="–")
        self._ck_kpi_value_var  = tk.StringVar(value="–")
        self._ck_kpi_claim_var  = tk.StringVar(value="–")
        self._ck_kpi_sel_var    = tk.StringVar(value="–")

        for col, (var, label, fgcolor) in enumerate([
            (self._ck_kpi_rows_var,    "Report-Zeilen",         "#7C3AED"),
            (self._ck_kpi_entries_var, "Entry Summaries",       "#D97706"),
            (self._ck_kpi_value_var,   "Invoice Value gesamt",  "#059669"),
            (self._ck_kpi_claim_var,   "Tariff Claim gesamt",   "#DC2626"),
            (self._ck_kpi_sel_var,     "Auswahl Tariff Claim",  "#0F766E"),
        ]):
            card = tk.Frame(kpi_row, bg=CARD_BG,
                            highlightthickness=1, highlightbackground="#E4E8EF")
            card.grid(row=0, column=col, sticky="nsew",
                      padx=(0, 12 if col < 3 else 0))
            # Letztes KPI-Feld zeigt längeren Text → kleinere Schrift
            val_font = ("Segoe UI", 14, "bold") if col == 4 else ("Segoe UI", 22, "bold")
            tk.Label(card, textvariable=var, bg=CARD_BG, fg=fgcolor,
                     font=val_font).pack(
                         padx=16, pady=(12, 2), anchor="w")
            tk.Label(card, text=label, bg=CARD_BG, fg="#8A9EB5",
                     font=("Segoe UI", 9)).pack(padx=16, pady=(0, 12), anchor="w")

        # ── Vorschau-Tabelle ──────────────────────────────────────────────────
        tbl_frame = ttk.LabelFrame(
            outer, text="Vorschau: Claim-Report (alle Template-Felder)")
        tbl_frame.pack(fill="both", expand=True)

        _cf = ttk.Frame(tbl_frame)
        _cf.pack(fill="x", padx=6, pady=(4, 2))
        self.ck_table = ScrollableTable(
            tbl_frame, CLAIM_PREVIEW_COLS, CLAIM_PREVIEW_LABELS, col_width=130)
        self._ck_filter_reapply = _setup_filter_bar(
            _cf, self.ck_table, CLAIM_PREVIEW_COLS, CLAIM_PREVIEW_LABELS)

        # ── Monats-Filter ────────────────────────────────────────────────────────────────────
        _mf = tk.Frame(tbl_frame, bg=MAIN_BG)
        _mf.pack(fill="x", padx=6, pady=(0, 2))
        tk.Label(_mf, text="Import-Monat:", bg=MAIN_BG, fg="#5A6A80",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 4))
        self._ck_month_var = tk.StringVar(value="Alle")
        self._ck_month_cb  = ttk.Combobox(
            _mf, textvariable=self._ck_month_var, state="readonly", width=12)
        self._ck_month_cb["values"] = ["Alle"]
        self._ck_month_cb.pack(side="left")
        self._ck_month_cb.bind("<<ComboboxSelected>>", self._ck_apply_month_filter)

        self._ck_sel_all_btn = tk.Button(
            _mf, text="☑ Alle auswählen",
            bg="#0F766E", fg="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat", padx=8, pady=2,
            cursor="hand2",
            command=self._ck_select_all_filtered)
        self._ck_sel_all_btn.pack(side="left", padx=(10, 0))

        tk.Button(
            _mf, text="☐ Alle abwählen",
            bg="#6B7280", fg="white",
            font=("Segoe UI", 9),
            relief="flat", padx=8, pady=2,
            cursor="hand2",
            command=self._ck_deselect_all_filtered).pack(side="left", padx=(4, 0))

        self.ck_table.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        self.ck_table.tree.configure(selectmode="none")

        # Checkbox-Spalte am Anfang einrichten
        self._ck_setup_checkbox_col()

        # Doppelklick → Ergänzungsdialog; Klick Checkbox → Toggle
        self.ck_table.tree.bind("<Double-1>", self._ck_on_dbl_click)
        self.ck_table.tree.bind("<ButtonRelease-1>", self._ck_toggle_checkbox)

        # Einstellungen laden und Vorschau befüllen
        self._load_claim_settings()

    def _load_claim_settings(self):
        """Laedt gespeicherte Einstellungen aus claim_kunde_settings in die Formularfelder."""
        rows = self.conn.execute(
            "SELECT key, value FROM claim_kunde_settings").fetchall()
        saved = dict(rows)
        for key, var in self._ck_vars.items():
            var.set(saved.get(key, ""))
        self._load_claim_preview()

    def _save_claim_settings(self):
        """Speichert die aktuellen Formularwerte in claim_kunde_settings."""
        for key, var in self._ck_vars.items():
            self.conn.execute(
                "INSERT OR REPLACE INTO claim_kunde_settings (key, value) VALUES (?,?)",
                (key, var.get().strip()),
            )
        self.conn.commit()
        self._ck_status_var.set("Einstellungen gespeichert.")
        self._load_claim_preview()

    def _ck_refresh_entry_list(self):
        """Füllt Entry- und HTS-Listbox mit allen verfügbaren Werten aus der DB."""
        # CBP-Einträge
        if hasattr(self, "_ck_entry_lb"):
            try:
                rows = self.conn.execute(
                    """SELECT filer_code_entry_no, entry_date, exporting_country,
                              total_entered_value
                       FROM entries
                       ORDER BY entry_date DESC, filer_code_entry_no"""
                ).fetchall()
            except Exception:
                rows = []
            self._ck_entry_lb.delete(0, "end")
            for entry_no, edate, country, total in rows:
                label = f"{entry_no}  |  {edate or '?'}  |  {country or '?'}  |  ${total or 0}"
                self._ck_entry_lb.insert("end", label)
            self._ck_entry_lb.selection_set(0, "end")

        # HTS-Codes aus entry_lines
        if hasattr(self, "_ck_hts_lb"):
            try:
                hts_rows = self.conn.execute(
                    """SELECT DISTINCT l.htsus_no, l.htsus_rate,
                              COUNT(*) as cnt,
                              SUM(CASE WHEN l.duty_amount IS NOT NULL
                                       AND l.duty_amount != '' THEN 1 ELSE 0 END) as with_duty
                       FROM entry_lines l
                       WHERE l.htsus_no IS NOT NULL AND l.htsus_no != ''
                             AND l.htsus_no != 'MPF'
                       GROUP BY l.htsus_no, l.htsus_rate
                       ORDER BY l.htsus_no"""
                ).fetchall()
            except Exception:
                hts_rows = []
            self._ck_hts_lb.delete(0, "end")
            for hts_no, hts_rate, cnt, with_duty in hts_rows:
                rate_str = f"  {hts_rate}" if hts_rate else ""
                label = f"{hts_no}{rate_str}  ({cnt} Zeilen, {with_duty} mit Zoll)"
                self._ck_hts_lb.insert("end", label)
            # Alle HTS-Codes standardmäßig auswählen (Benutzer kann abwählen)
            self._ck_hts_lb.selection_set(0, "end")

    def _ck_refresh_and_reload(self):
        """Einträge + HTS-Liste neu laden, dann Vorschau aktualisieren."""
        self._ck_refresh_entry_list()
        self._load_claim_preview()

    def _ck_get_selected_entry_nos(self):
        """Gibt die ausgewählten Entry-Nummern zurück (leer = alle)."""
        if not hasattr(self, "_ck_entry_lb"):
            return []
        sel = self._ck_entry_lb.curselection()
        return [self._ck_entry_lb.get(i).split("|")[0].strip() for i in sel]

    def _ck_get_selected_hts_codes(self):
        """Gibt die ausgewählten HTS-Codes zurück (leer = alle)."""
        if not hasattr(self, "_ck_hts_lb"):
            return []
        sel = self._ck_hts_lb.curselection()
        return [self._ck_hts_lb.get(i).split()[0].strip() for i in sel]

    def _get_claim_settings(self):
        """Gibt die aktuellen Einstellungen als Dict zurueck."""
        return {k: v.get().strip() for k, v in self._ck_vars.items()}

    def _load_claim_preview(self):
        """Erstellt Vorschau aus entries/entry_lines (Datenbank) + SAP MB51 + Supplement."""
        s = self._get_claim_settings()

        def _nz(v, fmt=None):
            if v is None or v == 0 or v == 0.0:
                return ""
            if fmt:
                import re as _re
                m = _re.search(r'\.([0-9]+)([fd])', fmt)
                if m:
                    return _de_num(float(v), int(m.group(1)))
                return fmt.format(v)   # z.B. Prozent-Format unverändert
            return v

        def _parse_rate(rate_str):
            """Wandelt Zollsatz-String ('25%', '0.25', '25') in Float 0-1 um."""
            if not rate_str:
                return 0.0
            s2 = str(rate_str).strip().rstrip("%")
            try:
                v = float(s2)
                return v / 100.0 if v > 1.0 else v
            except ValueError:
                return 0.0

        # SAP-Tabellen prüfen
        has_sap_iv  = self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='sap_iv'"
        ).fetchone() is not None
        has_sap_mb51 = self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='sap_mb51'"
        ).fetchone() is not None

        if has_sap_iv:
            # Primärer Weg: entry_lines.invoice_reference → sap_iv.Ext_Referenz → Material
            # dann sap_mb51[601].Material → Kunde
            _iv_join = (
                'LEFT JOIN (SELECT "Ext_Referenz","Material","Menge","Mengeneinheit","Belegdatum" '
                'FROM sap_iv GROUP BY "Ext_Referenz") sapiv '
                'ON sapiv."Ext_Referenz" = l.invoice_reference '
            )
            _601_join = (
                'LEFT JOIN (SELECT "Material","Kunde","Belegkopftext" '
                'FROM sap_mb51 WHERE "Bewegungsart"=\'601\' GROUP BY "Material") sap601 '
                'ON sap601."Material" = sapiv."Material"'
            ) if has_sap_mb51 else ""
            sap_join     = _iv_join + _601_join
            sap_menge    = 'COALESCE(sapiv."Menge", l.net_quantity, l.manifest_qty, l.invoice_qty, 0)'
            sap_datum    = 'COALESCE(sp.invoice_date, sapiv."Belegdatum", "")'
            sap_mat      = 'COALESCE(sp.ford_part_base, sapiv."Material", "")'
            sap_uom      = 'COALESCE(sapiv."Mengeneinheit", "")'
            sap_material = 'COALESCE(sapiv."Material", "")'
            sap_kunde_nr = 'COALESCE(sap601."Kunde", "")' if has_sap_mb51 else '""'
            sap_kunde_nm = 'COALESCE(sap601."Belegkopftext", "")' if has_sap_mb51 else '""'
        elif has_sap_mb51:
            # Fallback: MB51-101 als Zwischenschritt (alter Weg)
            sap_join  = (
                'LEFT JOIN (SELECT "Referenzbeleg","Material","Menge","Mengeneinheit","Belegdatum" '
                'FROM sap_mb51 WHERE "Bewegungsart"=\'101\' GROUP BY "Referenzbeleg") sap101 '
                'ON sap101."Referenzbeleg" = l.invoice_reference '
                'LEFT JOIN (SELECT "Material","Kunde","Belegkopftext" '
                'FROM sap_mb51 WHERE "Bewegungsart"=\'601\' GROUP BY "Material") sap601 '
                'ON sap601."Material" = sap101."Material"'
            )
            sap_menge    = 'COALESCE(sap101."Menge", l.net_quantity, l.manifest_qty, l.invoice_qty, 0)'
            sap_datum    = 'COALESCE(sp.invoice_date, sap101."Belegdatum", "")'
            sap_mat      = 'COALESCE(sp.ford_part_base, sap101."Material", "")'
            sap_uom      = 'COALESCE(sap101."Mengeneinheit", "")'
            sap_material = 'COALESCE(sap101."Material", "")'
            sap_kunde_nr = 'COALESCE(sap601."Kunde", "")'
            sap_kunde_nm = 'COALESCE(sap601."Belegkopftext", "")'
        else:
            sap_join     = ""
            sap_menge    = "COALESCE(l.net_quantity, l.manifest_qty, l.invoice_qty, 0)"
            sap_datum    = 'COALESCE(sp.invoice_date, "")'
            sap_mat      = 'COALESCE(sp.ford_part_base, "")'
            sap_uom      = '""'
            sap_material = '""'
            sap_kunde_nr = '""'
            sap_kunde_nm = '""'

        # Eintrags-Filter aus Listbox-Selektion
        selected_entries = self._ck_get_selected_entry_nos()
        selected_hts     = self._ck_get_selected_hts_codes()
        where_extra  = ""
        params_extra = []

        if selected_entries:
            ph_in = ",".join("?" * len(selected_entries))
            where_extra  += f" AND e.filer_code_entry_no IN ({ph_in})"
            params_extra += selected_entries

        if selected_hts:
            ph_hts = ",".join("?" * len(selected_hts))
            where_extra  += f" AND l.htsus_no IN ({ph_hts})"
            params_extra += selected_hts

        try:
            rows = self.conn.execute(
                f"""SELECT
                    l.id,
                    e.filer_code_entry_no,
                    e.exporting_country,
                    l.htsus_no,
                    COALESCE(e.import_date, e.entry_date, ''),
                    l.htsus_rate,
                    l.entered_value,
                    l.duty_amount,
                    COALESCE(sp.supplier_invoice_no, l.invoice_no, ''),
                    {sap_datum},
                    {sap_mat},
                    COALESCE(sp.supplier_part_number, ''),
                    COALESCE(sp.t2_part_number, ''),
                    COALESCE(sp.weight, 0),
                    {sap_menge},
                    {sap_uom},
                    {sap_material},
                    {sap_kunde_nr},
                    {sap_kunde_nm}
                FROM entry_lines l
                JOIN entries e ON e.id = l.entry_id
                {sap_join}
                LEFT JOIN claim_kunde_supplement sp ON sp.tcl_id = l.id
                WHERE l.duty_amount IS NOT NULL
                  AND CAST(REPLACE(COALESCE(l.duty_amount,'0'),',','') AS REAL) > 0
                  {where_extra}
                ORDER BY l.id""",
                params_extra
            ).fetchall()
        except Exception as exc:
            if hasattr(self, "_ck_status_var"):
                self._ck_status_var.set(f"Ladefehler: {exc}")
            return

        def _pfloat(v):
            """String-Zahl ('8,151.85' oder '312.55') sicher zu float."""
            if v is None:
                return 0.0
            try:
                return float(str(v).replace(",", ""))
            except (ValueError, TypeError):
                return 0.0

        comp = self._parse_float(s.get("components_per_part", "1") or "1")

        # ── ZuOrd-MAT-KD Lookup-Dict (ford_base → kunden_nr, sap_customer_name) ──
        _zuord_map = {}
        try:
            _has_zuord = self.conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name='stammdaten_zuord_mat_kd'"
            ).fetchone()
            if _has_zuord:
                for _zr in self.conn.execute(
                    "SELECT mat_no_sap, kunden_nr, sap_customer_name "
                    "FROM stammdaten_zuord_mat_kd"
                ).fetchall():
                    _k = str(_zr[0] or "").strip()
                    if _k:
                        _zuord_map[_k] = (
                            _k,
                            str(_zr[1] or "").strip(),
                            str(_zr[2] or "").strip(),
                        )
        except Exception:
            pass

        preview_rows = []
        total_invoice = 0.0
        total_claim   = 0.0
        entry_set     = set()

        for (line_id, entry_no, country, hts, imp_date,
             htsus_rate, entered_val_raw, duty_amt_raw,
             inv_no, inv_date, ford_base, supp_part, t2_part, weight,
             sap_qty, uom, mat_nr, kunde_nr, kunde_nm) in rows:

            duty_amt  = _pfloat(duty_amt_raw)
            goods_val = _pfloat(entered_val_raw)
            qty       = _pfloat(sap_qty)

            # Zollsatz: aus htsus_rate-Feld oder berechnet
            tariff_pct = _parse_rate(htsus_rate) if htsus_rate else 0

            # entered_value meist NULL → rückrechnen aus duty / rate
            if goods_val == 0 and duty_amt > 0 and tariff_pct > 0:
                goods_val = round(duty_amt / tariff_pct, 2)
            elif goods_val == 0 and tariff_pct == 0:
                # Fallback: aus entry gesamt (nur wenn 1 Zeile)
                goods_val = 0

            unit_price = round(goods_val / qty, 4) if qty else 0

            tariff_type   = s.get("tariff_type", "")
            is_steel_alum = "steel" in tariff_type.lower() or "alum" in tariff_type.lower()
            w = float(weight) if weight else 0
            tariff_ppu = round(unit_price * tariff_pct * w, 4) if (is_steel_alum and w) \
                         else round(unit_price * tariff_pct, 4)

            tariff_per_part  = round(tariff_ppu * comp, 4)
            # Tariff Claim = duty_amount aus CBP7501 (direkter Wert, 100% Übereinstimmung)
            # Multiplikator "components_per_part" bleibt erhalten
            tariff_claim_val = round(duty_amt * comp, 2)
            # Stückpreis-Felder: nur informativ (können 0 sein wenn qty fehlt)
            if tariff_claim_val == 0 and duty_amt > 0:
                tariff_claim_val = round(duty_amt, 2)

            total_invoice += goods_val
            total_claim   += tariff_claim_val
            entry_set.add(entry_no)

            preview_rows.append((
                s.get("vhm", ""),
                s.get("commodity_code", ""),
                s.get("vehicle_line", ""),
                s.get("production_fcsd", ""),
                _nz(qty),
                _nz(qty),
                tariff_type,
                s.get("claim_start", ""),
                s.get("claim_end", ""),
                country,
                entry_no,
                hts,
                imp_date,
                _nz(tariff_pct, "{:.0%}"),
                inv_no,
                _nz(qty),
                _nz(unit_price, "{:.4f}"),
                _nz(goods_val, "{:.2f}"),
                inv_date,
                s.get("ford_prefix", ""),
                ford_base,
                s.get("ford_suffix", ""),
                supp_part,
                t2_part,
                s.get("raw_material", ""),
                _nz(w),
                uom,
                s.get("ami_linked", ""),
                s.get("index_code", ""),
                _nz(tariff_ppu, "{:.4f}"),
                _nz(comp),
                _nz(tariff_per_part, "{:.4f}"),
                _nz(tariff_claim_val, "{:.2f}"),
                mat_nr or "",
                kunde_nr or "",
                kunde_nm or "",
            ))
            # ZuOrd-MAT-KD: ford_base als Join-Schlüssel überschreibt SAP-Werte
            _z = _zuord_map.get((ford_base or "").strip())
            if _z:
                preview_rows[-1] = preview_rows[-1][:-3] + (_z[0], _z[1], _z[2])

        # Vollständige Datensätze merken (für Nullwert-Filter)
        self._ck_all_preview_rows = preview_rows
        self._ck_full_tcl_ids      = [r[0] for r in rows]
        self._ck_full_preview_rows = list(preview_rows)
        if hasattr(self, "_ck_month_var"):
            self._ck_update_month_choices()
            self._ck_apply_month_filter()
        else:
            self._ck_all_tcl_ids      = list(self._ck_full_tcl_ids)
            self._ck_all_preview_rows = list(preview_rows)
            self.ck_table._tcl_ids    = self._ck_all_tcl_ids[:]
            self.ck_table.set_rows(preview_rows)
            if hasattr(self, "_ck_hide_zeros") and self._ck_hide_zeros.get():
                self._ck_apply_zero_filter()
            elif hasattr(self, "_ck_filter_reapply"):
                self._ck_filter_reapply()

        n = len(preview_rows)
        self._ck_kpi_rows_var.set(f"{n:,}".replace(",", "."))
        self._ck_kpi_entries_var.set(str(len(entry_set)))
        self._ck_kpi_value_var.set(f"${total_invoice:,.0f}".replace(",", "."))
        self._ck_kpi_claim_var.set(f"$ {_de_num(total_claim)}")
        if hasattr(self, "_ck_status_var"):
            self._ck_status_var.set(
                f"{n} Zeilen geladen." if n
                else "Keine Daten – bitte zuerst CBP 7501 PDFs importieren (entries/entry_lines).")

    @staticmethod
    def _parse_float(s, default=1.0):
        try:
            return float(str(s).replace(",", ".")) or default
        except (ValueError, TypeError):
            return default

    # Spalten-Indizes der Zahlenfelder in CLAIM_PREVIEW_COLS (0-basiert)
    # cmms_volume=4, ship_volume=5, tariff_pct=13, quantity=15,
    # unit_price=16, invoice_value=17, weight=25,
    # tariff_ppu=29, components_per_part=30, tariff_per_part=31, tariff_claim=32
    _CK_NUMERIC_COLS = (4, 5, 13, 15, 16, 17, 25, 29, 30, 31, 32)

    # Nullwert-Check: Werte die als "leer / 0" gelten
    _ZERO_VALS = {"", "0", "0.0", "0.00", "0.0000", "0%", "0.00%",
                  "None", "none", "null"}

    def _ck_apply_zero_filter(self):
        """Filtert Zeilen anhand des gewaehlten Zahlenfeldes (>0 behalten)."""
        if not hasattr(self, "ck_table") or not hasattr(self, "_ck_all_preview_rows"):
            return

        hide_zeros = self._ck_hide_zeros.get()

        # Welches Feld wurde ausgewählt?
        selected_label = (
            self._ck_zero_field_var.get()
            if hasattr(self, "_ck_zero_field_var") else "— Alle Zahlenfelder —"
        )

        if not hide_zeros:
            # Filter deaktiviert → alle Zeilen zeigen
            filtered_rows   = self._ck_all_preview_rows
            filtered_ids    = self._ck_all_tcl_ids
            field_info      = None
        else:
            # Spalten-Index(e) bestimmen
            if selected_label.startswith("—"):
                check_cols = self._CK_NUMERIC_COLS
                field_info = "alle Zahlenfelder"
                # Zeile behalten wenn MINDESTENS EIN Zahlenfeld ≠ 0
                def _keep(row):
                    return any(
                        str(row[i]).strip() not in self._ZERO_VALS
                        for i in check_cols if i < len(row)
                    )
            else:
                # Bestimmtes Feld: Spaltenname → Index über CLAIM_PREVIEW_COLS
                col_idx = next(
                    (i for i in self._CK_NUMERIC_COLS
                     if CLAIM_PREVIEW_LABELS.get(CLAIM_PREVIEW_COLS[i]) == selected_label),
                    None
                )
                field_info = selected_label
                if col_idx is None:
                    filtered_rows = self._ck_all_preview_rows
                    filtered_ids  = self._ck_all_tcl_ids
                    field_info    = None
                else:
                    def _keep(row, _ci=col_idx):
                        return (
                            _ci < len(row)
                            and str(row[_ci]).strip() not in self._ZERO_VALS
                        )

            if field_info is not None:
                pairs = [
                    (row, tid)
                    for row, tid in zip(self._ck_all_preview_rows,
                                        self._ck_all_tcl_ids)
                    if _keep(row)
                ]
                filtered_rows = [p[0] for p in pairs]
                filtered_ids  = [p[1] for p in pairs]

        # Tabelle mit gefilterter Datenmenge neu befüllen
        self.ck_table._tcl_ids = filtered_ids
        self.ck_table.set_rows(filtered_rows)
        if hasattr(self, "_ck_filter_reapply"):
            self._ck_filter_reapply()

        total   = len(self._ck_all_preview_rows)
        shown   = len(filtered_rows)
        hidden  = total - shown
        if hide_zeros and field_info and hidden:
            self._ck_status_var.set(
                f"{shown} angezeigt, {hidden} ausgeblendet "
                f"({field_info} = 0).")
        else:
            self._ck_status_var.set(f"{shown} Zeilen." if shown else "")

    def _ck_update_month_choices(self):
        """Fuellt Monat-Combobox mit einzigartigen Monaten aus import_date."""
        if not hasattr(self, "_ck_month_cb"):
            return
        IDX = CLAIM_PREVIEW_COLS.index("import_date")
        months = set()
        for row in getattr(self, "_ck_full_preview_rows", []):
            d = str(row[IDX] or "").strip()
            if not d:
                continue
            parts = d.split("/")
            if len(parts) >= 3:
                mm   = parts[0].zfill(2)
                yy   = parts[2]
                yyyy = ("20" + yy) if len(yy) == 2 else yy
                months.add(mm + "/" + yyyy)
        sorted_months = sorted(months, key=lambda x: (x[3:], x[:2]))
        values = ["Alle"] + sorted_months
        cur = self._ck_month_var.get()
        self._ck_month_cb["values"] = values
        if cur not in values:
            self._ck_month_var.set("Alle")

    def _ck_apply_month_filter(self, *_):
        """Filtert _full_preview_rows nach gewaehltem Monat und ladet Tabelle."""
        if not hasattr(self, "_ck_full_preview_rows"):
            return
        IDX   = CLAIM_PREVIEW_COLS.index("import_date")
        month = self._ck_month_var.get() if hasattr(self, "_ck_month_var") else "Alle"
        if not month or month == "Alle":
            filtered_rows = self._ck_full_preview_rows
            filtered_ids  = self._ck_full_tcl_ids
        else:
            mm, yyyy = month.split("/")
            yy = yyyy[2:]
            def _match(d, _mm=mm, _yy=yy, _yyyy=yyyy):
                d = str(d or "").strip()
                p = d.split("/")
                if len(p) < 3:
                    return False
                d_mm = p[0].zfill(2)
                d_yy = p[2]
                if len(d_yy) == 2:
                    return d_mm == _mm and d_yy == _yy
                return d_mm == _mm and d_yy == _yyyy
            pairs = [(r, tid) for r, tid in
                     zip(self._ck_full_preview_rows, self._ck_full_tcl_ids)
                     if _match(r[IDX])]
            filtered_rows = [p[0] for p in pairs]
            filtered_ids  = [p[1] for p in pairs]
        self._ck_all_preview_rows = filtered_rows
        self._ck_all_tcl_ids      = filtered_ids
        self.ck_table._tcl_ids    = list(filtered_ids)
        self.ck_table.set_rows(filtered_rows)
        if hasattr(self, "_ck_hide_zeros") and self._ck_hide_zeros.get():
            self._ck_apply_zero_filter()
        elif hasattr(self, "_ck_filter_reapply"):
            self._ck_filter_reapply()

    # ── Checkbox-Spalte: Infrastruktur ──────────────────────────────────────
    def _ck_setup_checkbox_col(self):
        """Fuegt _ck_sel als erste Spalte ein und patched _show/_insert_grand_total."""
        import types as _types
        _SEL = "_ck_sel"
        _all_cols = [_SEL] + list(CLAIM_PREVIEW_COLS)

        # Treeview neu konfigurieren
        self.ck_table.tree.configure(columns=_all_cols)

        # Alle Spaltenköpfe neu setzen (configure() setzt sie auf Default zurück)
        self.ck_table.tree.heading(_SEL, text="✓", anchor="center")
        self.ck_table.tree.column(_SEL, width=32, minwidth=32,
                                  stretch=False, anchor="center")
        for col in CLAIM_PREVIEW_COLS:
            self.ck_table.tree.heading(col, text=CLAIM_PREVIEW_LABELS.get(col, col))
            self.ck_table.tree.column(col, width=130, minwidth=60,
                                      anchor="w", stretch=False)
        self.ck_table.columns = _all_cols

        # Selektion aus DB laden
        self._ck_load_sel_set()

        # _show patchen
        _app = self
        def _patched_show(st, rows):
            for item in st.tree.get_children():
                st.tree.delete(item)
            sel = getattr(_app, "_ck_sel_set", set())
            count = 0
            for row in rows:
                key = _app._ck_row_key(row)
                chk = "☑" if key in sel else "☐"
                vals = (chk,) + tuple("" if v is None else v for v in row)
                tag = "even" if count % 2 == 0 else "odd"
                st.tree.insert("", "end", values=vals, tags=(tag,))
                count += 1
            st._insert_grand_total(rows)
            return len(rows)
        self.ck_table._show = _types.MethodType(_patched_show, self.ck_table)

        # _insert_grand_total patchen (weiss, dass Spalte 0 = _ck_sel)
        def _patched_igt(st, rows):
            if not rows:
                return
            ncols = len(st.columns)          # 37 (mit _ck_sel)
            sums = {}
            for ci, col in enumerate(st.columns):
                if ci == 0 or not _is_sum_col(col):
                    continue
                total = 0.0
                found = False
                for row in rows:
                    v = row[ci - 1] if (ci - 1) < len(row) else None
                    parsed = _parse_de(v)
                    if parsed is not None:
                        total += parsed
                        found = True
                if found:
                    sums[ci] = total
            if not sums:
                return
            footer = [""] * ncols
            footer[0] = ""
            footer[1] = "∑ GESAMT"
            for ci, total in sums.items():
                footer[ci] = _de_num(total)
            st.tree.insert("", "end", values=tuple(footer), tags=("grand_total",))
        self.ck_table._insert_grand_total = _types.MethodType(_patched_igt, self.ck_table)

    def _ck_row_key(self, row):
        """Erstellt einen eindeutigen Schluessel fuer eine Claim-Vorschau-Zeile.
        None-Werte werden als '' normalisiert, damit Treeview-Werte (None→'')
        und DB-Rohwerte (None) immer denselben Schluessel ergeben."""
        # entry_no=10, hts_code=11, import_date=12, supplier_invoice_no=14, tariff_claim=32
        def _s(v):
            return "" if (v is None or v == "None") else str(v).strip()
        return "|".join(_s(row[i]) if i < len(row) else ""
                        for i in (10, 11, 12, 14, 32))

    def _ck_load_sel_set(self):
        """Laedt gespeicherte Zeilenauswahl aus der DB."""
        self._ck_sel_set = set()
        try:
            rows = self.conn.execute(
                "SELECT row_key FROM ck_row_selection WHERE selected=1").fetchall()
            for (key,) in rows:
                self._ck_sel_set.add(key)
        except Exception:
            pass

    def _ck_save_sel_row(self, key, selected):
        """Speichert Auswahlzustand einer Zeile sofort in der DB."""
        try:
            self.conn.execute(
                "INSERT OR REPLACE INTO ck_row_selection (row_key, selected) VALUES (?, ?)",
                (key, int(selected)))
            self.conn.commit()
        except Exception:
            pass

    def _ck_toggle_checkbox(self, event):
        """Behandelt Klick auf Checkbox-Spalte: Toggle + DB-Speicherung + KPI."""
        tv = self.ck_table.tree
        region = tv.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tv.identify_column(event.x)
        if col != "#1":          # Nur die erste Spalte (_ck_sel)
            return
        iid = tv.identify_row(event.y)
        if not iid:
            return
        vals = tv.item(iid, "values")
        if not vals:
            return
        # Footer-Zeile ignorieren
        if str(vals[0]) == "" and len(vals) > 1 and str(vals[1]).startswith("∑"):
            return
        # Originaldaten: vals[1:] (ohne Checkbox-Zeichen)
        orig_row = list(vals[1:])
        key = self._ck_row_key(orig_row)
        if key in self._ck_sel_set:
            self._ck_sel_set.discard(key)
            tv.set(iid, "_ck_sel", "☐")
            self._ck_save_sel_row(key, 0)
        else:
            self._ck_sel_set.add(key)
            tv.set(iid, "_ck_sel", "☑")
            self._ck_save_sel_row(key, 1)
        self._ck_update_selection_sum()

    def _ck_update_selection_sum(self):
        """Aktualisiert KPI 'Auswahl Tariff Claim' basierend auf _ck_sel_set."""
        if not hasattr(self, "_ck_kpi_sel_var"):
            return
        IDX_CLAIM = CLAIM_PREVIEW_COLS.index("tariff_claim")
        total = 0.0
        count = 0
        for row in getattr(self, "_ck_full_preview_rows", []):
            key = self._ck_row_key(row)
            if key in self._ck_sel_set:
                parsed = _parse_de(row[IDX_CLAIM] if IDX_CLAIM < len(row) else None)
                if parsed is not None:
                    total += parsed
                    count += 1
        if count:
            self._ck_kpi_sel_var.set("$ " + _de_num(total) + " (" + str(count) + " Zeilen)")
        else:
            self._ck_kpi_sel_var.set("–")

    def _ck_select_all_filtered(self):
        """Markiert alle aktuell sichtbaren (gefilterten) Zeilen als ausgewaehlt."""
        rows = getattr(self, "_ck_all_preview_rows", [])
        for row in rows:
            key = self._ck_row_key(row)
            if key not in self._ck_sel_set:
                self._ck_sel_set.add(key)
                self._ck_save_sel_row(key, 1)
        self._ck_refresh_checkboxes()
        self._ck_update_selection_sum()

    def _ck_deselect_all_filtered(self):
        """Hebt die Auswahl aller aktuell sichtbaren (gefilterten) Zeilen auf."""
        rows = getattr(self, "_ck_all_preview_rows", [])
        for row in rows:
            key = self._ck_row_key(row)
            if key in self._ck_sel_set:
                self._ck_sel_set.discard(key)
                self._ck_save_sel_row(key, 0)
        self._ck_refresh_checkboxes()
        self._ck_update_selection_sum()

    def _ck_refresh_checkboxes(self):
        """Aktualisiert die Checkbox-Symbole aller sichtbaren Treeview-Zeilen."""
        tv   = self.ck_table.tree
        sel  = self._ck_sel_set
        for iid in tv.get_children():
            vals = tv.item(iid, "values")
            if not vals:
                continue
            # Footer-Zeile überspringen
            if str(vals[0]) == "" and len(vals) > 1 and str(vals[1]).startswith("∑"):
                continue
            orig_row = list(vals[1:])
            key = self._ck_row_key(orig_row)
            tv.set(iid, "_ck_sel", "☑" if key in sel else "☐")

    def _ck_on_dbl_click(self, event):
        tree = self.ck_table.tree
        iid = tree.identify_row(event.y)
        if not iid:
            return
        idx = tree.index(iid)
        if not hasattr(self.ck_table, "_tcl_ids") or idx >= len(self.ck_table._tcl_ids):
            return
        self._ck_edit_supplement(self.ck_table._tcl_ids[idx])

    def _ck_edit_supplement(self, tcl_id):
        existing = self.conn.execute(
            "SELECT supplier_invoice_no, invoice_date, ford_part_base,"
            " supplier_part_number, t2_part_number, weight"
            " FROM claim_kunde_supplement WHERE tcl_id=?",
            (tcl_id,)
        ).fetchone() or ("", "", "", "", "", "")

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Zeilenergaenzung - ID {tcl_id}")
        dlg.resizable(False, False)
        dlg.grab_set()

        labels = ["Supplier Invoice No", "Invoice Date",
                  "Ford Part Base Nr.", "Supplier Part Nr.",
                  "T2 Part Nr.", "Weight"]
        vars_ = []
        for r, (lbl, val) in enumerate(zip(labels, existing)):
            ttk.Label(dlg, text=lbl + ":").grid(
                row=r, column=0, sticky="w", padx=(12, 4), pady=4)
            v = tk.StringVar(value=str(val) if val else "")
            ttk.Entry(dlg, textvariable=v, width=30).grid(
                row=r, column=1, sticky="ew", padx=(0, 12), pady=4)
            vars_.append(v)
        dlg.columnconfigure(1, weight=1)

        def _save():
            inv_no, inv_date, ford_base, supp_part, t2_part, weight_s = (
                v.get().strip() for v in vars_)
            wval = self._parse_float(weight_s, None) if weight_s else None
            self.conn.execute(
                "INSERT OR REPLACE INTO claim_kunde_supplement"
                " (tcl_id, supplier_invoice_no, invoice_date, ford_part_base,"
                "  supplier_part_number, t2_part_number, weight)"
                " VALUES (?,?,?,?,?,?,?)",
                (tcl_id, inv_no, inv_date, ford_base, supp_part, t2_part, wval),
            )
            self.conn.commit()
            dlg.destroy()
            self._load_claim_preview()

        bf = ttk.Frame(dlg)
        bf.grid(row=len(labels), column=0, columnspan=2,
                pady=10, padx=12, sticky="e")
        ttk.Button(bf, text="Abbrechen", command=dlg.destroy).pack(
            side="right", padx=(0, 8))
        ttk.Button(bf, text="Speichern", command=_save).pack(side="right")

    def _export_claim_excel(self):
        if openpyxl is None:
            messagebox.showerror("openpyxl fehlt", "pip install openpyxl")
            return
        from tkinter import filedialog
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        out_path = filedialog.asksaveasfilename(
            title="Claim-Report speichern",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"Claim-Report_Kunde_{datetime.now():%Y-%m-%d}.xlsx",
        )
        if not out_path:
            return

        s = self._get_claim_settings()
        has_sap_iv2  = self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='sap_iv'"
        ).fetchone() is not None
        has_sap_mb512 = self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='sap_mb51'"
        ).fetchone() is not None

        if has_sap_iv2:
            _iv_join2 = (
                'LEFT JOIN (SELECT "Ext_Referenz","Material","Menge","Mengeneinheit","Belegdatum" '
                'FROM sap_iv GROUP BY "Ext_Referenz") sapiv '
                'ON sapiv."Ext_Referenz" = l.invoice_reference '
            )
            _601_join2 = (
                'LEFT JOIN (SELECT "Material","Kunde","Belegkopftext" '
                'FROM sap_mb51 WHERE "Bewegungsart"=\'601\' GROUP BY "Material") sap601 '
                'ON sap601."Material" = sapiv."Material"'
            ) if has_sap_mb512 else ""
            sap_join2     = _iv_join2 + _601_join2
            sap_menge2    = 'COALESCE(sapiv."Menge", l.net_quantity, l.manifest_qty, l.invoice_qty, 0)'
            sap_datum2    = "COALESCE(sp.invoice_date, sapiv.\"Belegdatum\", '')"
            sap_mat2      = "COALESCE(sp.ford_part_base, sapiv.\"Material\", '')"
            sap_uom2      = 'COALESCE(sapiv."Mengeneinheit", \'\')'
            sap_material2 = 'COALESCE(sapiv."Material", \'\')'
            sap_kunde_nr2 = 'COALESCE(sap601."Kunde", \'\')' if has_sap_mb512 else "''"
            sap_kunde_nm2 = 'COALESCE(sap601."Belegkopftext", \'\')' if has_sap_mb512 else "''"
        elif has_sap_mb512:
            sap_join2  = (
                'LEFT JOIN (SELECT "Referenzbeleg","Material","Menge","Mengeneinheit","Belegdatum" '
                'FROM sap_mb51 WHERE "Bewegungsart"=\'101\' GROUP BY "Referenzbeleg") sap101 '
                'ON sap101."Referenzbeleg" = l.invoice_reference '
                'LEFT JOIN (SELECT "Material","Kunde","Belegkopftext" '
                'FROM sap_mb51 WHERE "Bewegungsart"=\'601\' GROUP BY "Material") sap601 '
                'ON sap601."Material" = sap101."Material"'
            )
            sap_menge2    = 'COALESCE(sap101."Menge", l.net_quantity, l.manifest_qty, l.invoice_qty, 0)'
            sap_datum2    = "COALESCE(sp.invoice_date, sap101.\"Belegdatum\", '')"
            sap_mat2      = "COALESCE(sp.ford_part_base, sap101.\"Material\", '')"
            sap_uom2      = 'COALESCE(sap101."Mengeneinheit", \'\')'
            sap_material2 = 'COALESCE(sap101."Material", \'\')'
            sap_kunde_nr2 = 'COALESCE(sap601."Kunde", \'\')'
            sap_kunde_nm2 = 'COALESCE(sap601."Belegkopftext", \'\')'
        else:
            sap_join2     = ""
            sap_menge2    = "COALESCE(l.net_quantity, l.manifest_qty, l.invoice_qty, 0)"
            sap_datum2    = "COALESCE(sp.invoice_date, '')"
            sap_mat2      = "COALESCE(sp.ford_part_base, '')"
            sap_uom2      = "''"
            sap_material2 = "''"
            sap_kunde_nr2 = "''"
            sap_kunde_nm2 = "''"
        try:
            rows = self.conn.execute(
                f"""SELECT
                    l.id,
                    e.filer_code_entry_no,
                    e.exporting_country,
                    l.htsus_no,
                    COALESCE(e.import_date, e.entry_date, ''),
                    l.htsus_rate,
                    l.entered_value,
                    l.duty_amount,
                    COALESCE(sp.supplier_invoice_no, l.invoice_no, ''),
                    {sap_datum2},
                    {sap_mat2},
                    COALESCE(sp.supplier_part_number, ''),
                    COALESCE(sp.t2_part_number, ''),
                    COALESCE(sp.weight, 0),
                    {sap_menge2},
                    {sap_uom2},
                    {sap_material2},
                    {sap_kunde_nr2},
                    {sap_kunde_nm2}
                FROM entry_lines l
                JOIN entries e ON e.id = l.entry_id
                {sap_join2}
                LEFT JOIN claim_kunde_supplement sp ON sp.tcl_id = l.id
                WHERE l.duty_amount IS NOT NULL
                  AND CAST(REPLACE(COALESCE(l.duty_amount,'0'),',','') AS REAL) > 0
                ORDER BY l.id"""
            ).fetchall()
        except Exception as exc:
            messagebox.showerror("Datenbankfehler", str(exc))
            return

        if not rows:
            messagebox.showwarning("Keine Daten",
                                   "Keine Daten in entries/entry_lines – bitte zuerst CBP 7501 PDFs importieren.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Claim Report"

        _db = "FF1A2742"
        _gd = "FFD4A843"
        _lb = "FFD6E4F0"
        _wh = "FFFFFFFF"
        _thin = Side(style="thin", color="FFB8C9DC")
        _brd  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        _ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _lft  = Alignment(horizontal="left",   vertical="center")

        def hf(color=_wh, sz=9, bold=True):
            return Font(name="Segoe UI", bold=bold, color=color, size=sz)

        def fl(hex_):
            return PatternFill("solid", fgColor=hex_)

        ws.merge_cells("A1:AG1")
        c = ws["A1"]
        c.value = "To be filled by Supplier"
        c.font  = hf(sz=11)
        c.fill  = fl(_db)
        c.alignment = _ctr

        for rng, txt, fhex, fc in [
            ("A2:D2",   "General Info",          _db, _wh),
            ("E2:F2",   "Volume",                _gd, "FF000000"),
            ("G2:I2",   "Tariff",                _db, _wh),
            ("J2:R2",   "7501 Form",             _lb, "FF1A2742"),
            ("S2:AC2",  "Supplier Part Details", _db, _wh),
            ("AD2:AG2", "Tariff Impact",         _gd, "FF000000"),
        ]:
            ws.merge_cells(rng)
            c = ws[rng.split(":")[0]]
            c.value = txt
            c.font  = hf(color=fc)
            c.fill  = fl(fhex)
            c.alignment = _ctr

        col_hdrs = [
            "VHM", "Commodity Code", "Vehicle Line", "Production/FCSD parts",
            "CMMS Actual\nVolume", "Supplier Shipment\nVolume",
            "Tariff Type\n(Steel/Aluminum/China Tariff/Non USMCA/Others)",
            "Claim Start Period\n(mm/dd/yyyy)", "Claim End Period\n(mm/dd/yyyy)",
            "Exporting Country\n(14)", "Filer Code/\nEntry No (1)",
            "HTS Code", "Import Date\n(11)", "Tariff %\n(33)",
            "Supplier Invoice No", "Quantity\n(30)", "Unit Price", "Invoice Value",
            "Invoice Date", "Ford Part\nPrefix", "Ford\nPart Base Number",
            "Ford Part\nSuffix", "Supplier Part\nnumber",
            "T2 part number,\nwherever applicable",
            "Raw Material", "Weight\n(Raw material content\nin Ford part)", "UoM",
            "AMI Linked\n(Yes/No)", "If Linked,\nIndex Code",
            "Tariff Price Per Unit\n(Invoice Price x Tariff rate)",
            "Components\nPer Part", "Tariff $\nper part", "Tariff Claim\n($)",
        ]
        for ci, hdr in enumerate(col_hdrs, 1):
            c = ws.cell(row=3, column=ci, value=hdr)
            c.font      = hf(sz=8)
            c.fill      = fl(_db)
            c.alignment = _ctr
            c.border    = _brd

        comp_v = self._parse_float(s.get("components_per_part", "1") or "1")
        ttype  = s.get("tariff_type", "")
        is_sa  = "steel" in ttype.lower() or "alum" in ttype.lower()

        def _parse_rate_x(rate_str):
            if not rate_str:
                return 0.0
            s2 = str(rate_str).strip().rstrip("%")
            try:
                v = float(s2)
                return v / 100.0 if v > 1.0 else v
            except ValueError:
                return 0.0

        def _pfloat2(v):
            if v is None:
                return 0.0
            try:
                return float(str(v).replace(",", ""))
            except (ValueError, TypeError):
                return 0.0

        def _xnz(v):
            """None statt 0 fuer Excel – leere Zelle statt 0."""
            return None if (v == 0 or v == 0.0) else v

        for dr, (_, entry_no, country, hts, imp_date,
                  htsus_rate, gv_raw, duty_raw,
                  inv_no, inv_date, ford_base, supp_part, t2_part,
                  weight, sap_qty, uom,
                  mat_nr, kunde_nr, kunde_nm) in enumerate(rows, 4):
            duty = _pfloat2(duty_raw)
            gv   = _pfloat2(gv_raw)
            qty  = _pfloat2(sap_qty)
            tp   = _parse_rate_x(htsus_rate)
            # entered_value meist NULL → aus duty/rate berechnen
            if gv == 0 and duty > 0 and tp > 0:
                gv = round(duty / tp, 2)
            up   = round(gv / qty, 4) if qty else 0
            wv   = float(weight) if weight else 0
            ppu  = up * tp * wv if (is_sa and wv) else up * tp
            ppp  = ppu * comp_v
            # Tariff Claim = duty_amount aus CBP7501 (direkter Wert, 100% Übereinstimmung)
            tc   = round(duty * comp_v, 2)

            vals = [
                s.get("vhm",""), s.get("commodity_code",""),
                s.get("vehicle_line",""), s.get("production_fcsd",""),
                _xnz(qty), _xnz(qty), ttype,
                s.get("claim_start",""), s.get("claim_end",""),
                country, entry_no, hts, imp_date,
                _xnz(tp),
                inv_no, _xnz(qty), _xnz(up), _xnz(gv), inv_date,
                s.get("ford_prefix",""), ford_base, s.get("ford_suffix",""),
                supp_part, t2_part, s.get("raw_material",""),
                wv if wv else None, uom,
                s.get("ami_linked",""), s.get("index_code",""),
                _xnz(ppu), _xnz(comp_v), _xnz(ppp), _xnz(tc),
                mat_nr or "", kunde_nr or "", kunde_nm or "",
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=dr, column=ci, value=val)
                c.font      = Font(name="Segoe UI", size=9)
                c.border    = _brd
                c.alignment = _lft
                if ci in (5, 6, 16):
                    c.number_format = "#,##0"
                elif ci == 14:
                    c.number_format = "0.00%"
                elif ci in (17, 30, 31, 32):
                    c.number_format = "#,##0.0000"
                elif ci in (18, 33):
                    c.number_format = "#,##0.00"

        for ci, w in enumerate(
            [8,14,14,14, 10,10, 24,14,14, 10,18,14,12,8,
             16,10,10,12,12, 10,14,10,16,18,12,8,6,
             10,14, 18,10,12,14], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[3].height = 48
        ws.freeze_panes = "A4"

        try:
            wb.save(out_path)
            self._ck_status_var.set(
                f"Gespeichert: {os.path.basename(out_path)}")
            try:
                os.startfile(out_path)
            except Exception:
                pass
        except Exception as exc:
            messagebox.showerror("Export fehlgeschlagen", str(exc))

    # -----------------------------------------------------------------------
    # SAP-Reiter: MB51 Warenbewegungen
    # -----------------------------------------------------------------------
    def _build_sap_tab(self):
        outer = ttk.Frame(self.tab_sap, padding=14)
        outer.pack(fill="both", expand=True)

        # Titel
        top = ttk.Frame(outer)
        top.pack(fill="x", pady=(0, 8))
        ttk.Label(top, text="SAP-Daten (MB51) – Warenbewegungen",
                  font=("Segoe UI", 15, "bold")).pack(side="left")

        # ── Abruf-Parameter ─────────────────────────────────────────────────
        pf = ttk.LabelFrame(outer, text="MB51 Abruf-Parameter", padding=(12, 8))
        pf.pack(fill="x", pady=(0, 8))

        from datetime import timedelta as _td
        today_str = datetime.now().strftime("%Y%m%d")
        jan1_str  = (datetime.now() - _td(days=730)).strftime("%Y%m%d")  # 2 Jahre zurück

        # Zeile 0: Datum + Werk + Material
        base_defs = [
            ("Datum von (YYYYMMDD)",  "sap_date_from", jan1_str,  12),
            ("Datum bis (YYYYMMDD)",  "sap_date_to",   today_str, 12),
            ("Werk (leer = alle)",    "sap_werks",     "",        10),
            ("Material (Teilstring)", "sap_matnr",     "",        18),
        ]
        self._sap_param_vars = {}
        for col_idx, (label, key, default, width) in enumerate(base_defs):
            ttk.Label(pf, text=label).grid(row=0, column=col_idx*2,
                                           sticky="w", padx=(0 if col_idx==0 else 12, 2))
            var = tk.StringVar(value=default)
            ttk.Entry(pf, textvariable=var, width=width).grid(row=0, column=col_idx*2+1,
                                                               sticky="w")
            self._sap_param_vars[key] = var

        # Zeile 1: Bewegungsart – Checkboxen 101 / 601 + Freitext
        bwart_frame = ttk.Frame(pf)
        bwart_frame.grid(row=1, column=0, columnspan=len(base_defs)*2+2,
                         sticky="w", pady=(8, 0))
        ttk.Label(bwart_frame, text="Bewegungsart:").pack(side="left")
        self._sap_bwart_101 = tk.BooleanVar(value=True)
        self._sap_bwart_601 = tk.BooleanVar(value=True)
        ttk.Checkbutton(bwart_frame, text="101 (Wareneingang)",
                        variable=self._sap_bwart_101).pack(side="left", padx=(8, 0))
        ttk.Checkbutton(bwart_frame, text="601 (Lieferung/Ausgang)",
                        variable=self._sap_bwart_601).pack(side="left", padx=(8, 0))
        ttk.Label(bwart_frame, text="  Weitere:").pack(side="left", padx=(16, 2))
        self._sap_param_vars["sap_bwart_extra"] = tk.StringVar(value="")
        ttk.Entry(bwart_frame, textvariable=self._sap_param_vars["sap_bwart_extra"],
                  width=14).pack(side="left")
        ttk.Label(bwart_frame,
                  text="(kommagetrennt, z.B. 201,261)",
                  foreground="#777").pack(side="left", padx=(4, 0))

        # Buttons
        btn_frame = ttk.Frame(pf)
        btn_frame.grid(row=1, column=len(base_defs)*2+1, padx=(20, 0), pady=(8, 0))
        ttk.Button(btn_frame, text="▶  MB51 abrufen",
                   command=self._run_mb51_fetch).pack(side="left")
        ttk.Button(btn_frame, text="Tabelle aktualisieren",
                   command=self._load_sap_data).pack(side="left", padx=(8, 0))

        # Status-Zeile
        self._sap_status_var = tk.StringVar(value="")
        ttk.Label(pf, textvariable=self._sap_status_var,
                  foreground="#555").grid(row=2, column=0, columnspan=len(base_defs)*2+2,
                                          sticky="w", pady=(4, 0))

        # Log-Bereich
        log_frame = ttk.LabelFrame(outer, text="Worker-Ausgabe", padding=4)
        log_frame.pack(fill="x", pady=(0, 8))
        self._sap_log = tk.Text(log_frame, height=5, font=("Consolas", 9),
                                state="disabled", bg="#0B1929", fg="#A8D5FF",
                                insertbackground="white", relief="flat")
        self._sap_log.pack(fill="x")
        sb = ttk.Scrollbar(log_frame, orient="vertical", command=self._sap_log.yview)
        self._sap_log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._sap_log.pack(fill="x")

        sap_frame = ttk.LabelFrame(outer, text="MB51 Warenbewegungen")
        sap_frame.pack(fill="both", expand=True)

        self.sap_table = ScrollableTable(
            sap_frame, SAP_MB51_COLUMNS,
            {c: c for c in SAP_MB51_COLUMNS}, col_width=120)
        _sf = ttk.Frame(sap_frame)
        _sf.pack(fill="x", padx=6, pady=(4, 2))
        self._sap_filter_reapply = _setup_filter_bar(
            _sf, self.sap_table, SAP_MB51_COLUMNS,
            {c: c for c in SAP_MB51_COLUMNS})
        self.sap_table.pack(fill="both", expand=True, padx=6, pady=(0, 6))

    def _load_sap_data(self):
        try:
            cols = ", ".join(f'"{c}"' for c in SAP_MB51_COLUMNS)
            rows = self.conn.execute(
                f'SELECT {cols} FROM sap_mb51 ORDER BY rowid'
            ).fetchall()
            self.sap_table.set_rows(rows)
            if hasattr(self, "_sap_status_var"):
                self._sap_status_var.set(f"{len(rows)} Zeilen in der lokalen DB.")
        except sqlite3.OperationalError:
            self.sap_table.set_rows([])
            if hasattr(self, "_sap_status_var"):
                self._sap_status_var.set("Noch keine SAP-Daten importiert.")
        if hasattr(self, "_sap_filter_reapply"):
            self._sap_filter_reapply()

    def _sap_log_write(self, text: str):
        def _append():
            if not hasattr(self, "_sap_log"):
                return
            self._sap_log.configure(state="normal")
            self._sap_log.insert("end", text + "\n")
            self._sap_log.see("end")
            self._sap_log.configure(state="disabled")
        self.root.after(0, _append)

    def _run_mb51_fetch(self):
        import subprocess, threading
        if not os.path.exists(SAP_MB51_SCRIPT):
            messagebox.showerror("Skript nicht gefunden",
                                 f"mb51_extract.py nicht gefunden:\n{SAP_MB51_SCRIPT}")
            return

        # Bewegungsarten sammeln
        bwart_list = []
        if getattr(self, "_sap_bwart_101", None) and self._sap_bwart_101.get():
            bwart_list.append("101")
        if getattr(self, "_sap_bwart_601", None) and self._sap_bwart_601.get():
            bwart_list.append("601")
        extra = self._sap_param_vars.get("sap_bwart_extra", tk.StringVar()).get().strip()
        for b in extra.split(","):
            b = b.strip()
            if b and b not in bwart_list:
                bwart_list.append(b)
        if not bwart_list:
            messagebox.showwarning("Keine Bewegungsart",
                                   "Bitte mindestens eine Bewegungsart auswaehlen.")
            return

        python_exe = SAP_WORKER_PYTHON if os.path.exists(SAP_WORKER_PYTHON) else sys.executable
        v = self._sap_param_vars

        env_override = os.environ.copy()
        if os.path.exists(SAP_CONFIG_PATH):
            import configparser
            cfg = configparser.ConfigParser()
            cfg.read(SAP_CONFIG_PATH, encoding="utf-8")
            for section in cfg.sections():
                for k, val in cfg.items(section):
                    env_override[k.upper()] = val

        self._sap_log_write(f"Starte MB51-Abruf fuer Bewegungsarten: {', '.join(bwart_list)} ...")
        if hasattr(self, "_sap_status_var"):
            self._sap_status_var.set("Abruf laeuft ...")

        def _import_one(bwart, xl_path):
            """Importiert eine Excel-Datei fuer eine bestimmte Bewegungsart."""
            import openpyxl as _opx
            import sqlite3 as _sq3
            wb = _opx.load_workbook(xl_path, read_only=True, data_only=True)
            xl_rows = list(wb.active.iter_rows(min_row=2, values_only=True))
            wb.close()
            tconn = _sq3.connect(DB_PATH)
            col_defs = ", ".join(f'"{c}" TEXT' for c in SAP_MB51_COLUMNS)
            tconn.execute(f"CREATE TABLE IF NOT EXISTS sap_mb51 ({col_defs})")
            # Nur Zeilen dieser Bewegungsart loeschen (andere bleiben erhalten)
            bwart_col_idx = SAP_MB51_COLUMNS.index("Bewegungsart") if "Bewegungsart" in SAP_MB51_COLUMNS else -1
            if bwart_col_idx >= 0:
                tconn.execute('DELETE FROM sap_mb51 WHERE "Bewegungsart" = ?', (bwart,))
            else:
                tconn.execute("DELETE FROM sap_mb51")
            ph = ", ".join("?" * len(SAP_MB51_COLUMNS))
            for row in xl_rows:
                padded = list(row) + [None] * max(0, len(SAP_MB51_COLUMNS) - len(row))
                tconn.execute(f"INSERT INTO sap_mb51 VALUES ({ph})",
                              padded[:len(SAP_MB51_COLUMNS)])
            tconn.commit()
            tconn.close()
            return len(xl_rows)

        def _worker():
            total_rows = 0
            for bwart in bwart_list:
                ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_xl = os.path.join(APP_DIR, f"MB51_{bwart}_{ts}.xlsx")
                args_cli = [python_exe, SAP_MB51_SCRIPT, "--out", out_xl,
                            "--bwart", bwart]
                for flag, key in [("--date-from", "sap_date_from"),
                                  ("--date-to",   "sap_date_to"),
                                  ("--werks",     "sap_werks"),
                                  ("--matnr",     "sap_matnr")]:
                    val = v[key].get().strip()
                    if val:
                        args_cli += [flag, val]
                self._sap_log_write(f"── Bewegungsart {bwart} ──")
                try:
                    proc = subprocess.Popen(
                        args_cli, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, encoding="utf-8", errors="replace",
                        cwd=SAP_WORKER_DIR, env=env_override)
                    for line in proc.stdout:
                        self._sap_log_write(line.rstrip())
                    proc.wait()
                except Exception as exc:
                    self._sap_log_write(f"FEHLER {bwart}: {exc}")
                    continue
                if not os.path.exists(out_xl):
                    self._sap_log_write(f"Keine Ausgabedatei fuer {bwart} – Verbindung pruefen.")
                    continue
                self._sap_log_write(f"Importiere {os.path.basename(out_xl)} ...")
                try:
                    n = _import_one(bwart, out_xl)
                    total_rows += n
                    if n == 0:
                        self._sap_log_write(f"HINWEIS {bwart}: 0 Zeilen – Filter pruefen.")
                    else:
                        self._sap_log_write(f"OK {bwart} – {n} Zeilen gespeichert.")
                    try:
                        os.remove(out_xl)
                    except Exception:
                        pass
                except Exception as exc:
                    self._sap_log_write(f"Import-Fehler {bwart}: {exc}")
            # Abschluss
            self.root.after(0, self._load_sap_data)
            self.root.after(0, lambda r=total_rows: (
                self._sap_status_var.set(
                    f"{r} Zeilen gesamt importiert ({', '.join(bwart_list)}).")
                if hasattr(self, "_sap_status_var") else None))

        threading.Thread(target=_worker, daemon=True).start()

    # ── Eingangsrechnungen (IV) ──────────────────────────────────────────────

    def _build_sap_iv_tab(self):
        outer = ttk.Frame(self.tab_sap_iv, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="Eingangsrechnungen (IV) – Lieferantenrechnungen",
                  font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 10))

        # ── Parameter-Frame ──────────────────────────────────────────────────
        pf = ttk.LabelFrame(outer, text="Abruf-Parameter", padding=(12, 8))
        pf.pack(fill="x", pady=(0, 8))

        from datetime import timedelta as _td
        today_str = datetime.now().strftime("%Y%m%d")
        jan1_str  = (datetime.now() - _td(days=730)).strftime("%Y%m%d")

        base_iv = [
            ("Datum von (YYYYMMDD)", "iv_date_from", jan1_str,  12),
            ("Datum bis (YYYYMMDD)", "iv_date_to",   today_str, 12),
            ("Lieferant-Nr.",        "iv_lifnr",     "",        12),
            ("Material (Teilstring)","iv_matnr",     "",        18),
            ("Werk (leer = alle)",   "iv_werks",     "",        10),
        ]
        self._iv_param_vars = {}
        for ci, (label, key, default, width) in enumerate(base_iv):
            ttk.Label(pf, text=label).grid(row=0, column=ci*2,
                                            sticky="w", padx=(0 if ci==0 else 12, 2))
            var = tk.StringVar(value=default)
            ttk.Entry(pf, textvariable=var, width=width).grid(row=0, column=ci*2+1, sticky="w")
            self._iv_param_vars[key] = var

        btn_iv = ttk.Frame(pf)
        btn_iv.grid(row=0, column=len(base_iv)*2, padx=(20, 0))
        ttk.Button(btn_iv, text="▶  IV abrufen",
                   command=self._run_iv_fetch).pack(side="left")
        ttk.Button(btn_iv, text="Tabelle aktualisieren",
                   command=self._load_sap_iv_data).pack(side="left", padx=(8, 0))

        self._iv_status_var = tk.StringVar(value="")
        ttk.Label(pf, textvariable=self._iv_status_var,
                  foreground="#555").grid(row=1, column=0,
                                          columnspan=len(base_iv)*2+2,
                                          sticky="w", pady=(4, 0))

        # ── Log ─────────────────────────────────────────────────────────────
        log_frame = ttk.LabelFrame(outer, text="Worker-Ausgabe", padding=4)
        log_frame.pack(fill="x", pady=(0, 8))
        self._iv_log = tk.Text(log_frame, height=4, font=("Consolas", 9),
                               state="disabled", bg="#0B1929", fg="#A8D5FF",
                               insertbackground="white", relief="flat")
        self._iv_log.pack(fill="x")
        sb = ttk.Scrollbar(log_frame, orient="vertical", command=self._iv_log.yview)
        self._iv_log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        # ── Tabelle ──────────────────────────────────────────────────────────
        tbl_frame = ttk.LabelFrame(outer, text="Eingangsrechnungen")
        tbl_frame.pack(fill="both", expand=True)

        self.iv_table = ScrollableTable(
            tbl_frame, SAP_IV_COLUMNS,
            {c: c for c in SAP_IV_COLUMNS}, col_width=120)
        _sf = ttk.Frame(tbl_frame)
        _sf.pack(fill="x", padx=6, pady=(4, 2))
        self._iv_filter_reapply = _setup_filter_bar(
            _sf, self.iv_table, SAP_IV_COLUMNS,
            {c: c for c in SAP_IV_COLUMNS})
        self.iv_table.pack(fill="both", expand=True, padx=6, pady=(0, 6))

    def _iv_log_write(self, text: str):
        def _append():
            if not hasattr(self, "_iv_log"):
                return
            self._iv_log.configure(state="normal")
            self._iv_log.insert("end", text + "\n")
            self._iv_log.see("end")
            self._iv_log.configure(state="disabled")
        self.root.after(0, _append)

    def _load_sap_iv_data(self):
        try:
            cols = ", ".join(f'"{c}"' for c in SAP_IV_COLUMNS)
            rows = self.conn.execute(
                f'SELECT {cols} FROM sap_iv ORDER BY rowid'
            ).fetchall()
            self.iv_table.set_rows(rows)
            if hasattr(self, "_iv_status_var"):
                self._iv_status_var.set(f"{len(rows)} Zeilen in der lokalen DB.")
        except sqlite3.OperationalError:
            self.iv_table.set_rows([])
            if hasattr(self, "_iv_status_var"):
                self._iv_status_var.set("Noch keine IV-Daten importiert.")
        if hasattr(self, "_iv_filter_reapply"):
            self._iv_filter_reapply()

    def _run_iv_fetch(self):
        import subprocess, threading
        if not os.path.exists(SAP_IV_SCRIPT):
            messagebox.showerror("Skript nicht gefunden",
                                 f"iv_extract.py nicht gefunden:\n{SAP_IV_SCRIPT}")
            return
        python_exe = SAP_WORKER_PYTHON if os.path.exists(SAP_WORKER_PYTHON) else sys.executable
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_xl = os.path.join(APP_DIR, f"IV_{ts}.xlsx")
        v      = self._iv_param_vars
        args_cli = [python_exe, SAP_IV_SCRIPT, "--out", out_xl]
        for flag, key in [("--date-from", "iv_date_from"), ("--date-to", "iv_date_to"),
                          ("--lifnr",     "iv_lifnr"),     ("--matnr", "iv_matnr"),
                          ("--werks",     "iv_werks")]:
            val = v[key].get().strip()
            if val:
                args_cli += [flag, val]

        env_override = os.environ.copy()
        if os.path.exists(SAP_CONFIG_PATH):
            import configparser
            cfg = configparser.ConfigParser()
            cfg.read(SAP_CONFIG_PATH, encoding="utf-8")
            for section in cfg.sections():
                for k, val in cfg.items(section):
                    env_override[k.upper()] = val

        self._iv_log_write("Starte IV-Abruf (Eingangsrechnungen) ...")
        if hasattr(self, "_iv_status_var"):
            self._iv_status_var.set("Abruf laeuft ...")

        def _worker():
            try:
                proc = subprocess.Popen(
                    args_cli, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    cwd=SAP_WORKER_DIR, env=env_override)
                for line in proc.stdout:
                    self._iv_log_write(line.rstrip())
                proc.wait()
            except Exception as exc:
                self._iv_log_write(f"FEHLER: {exc}")
                self.root.after(0, lambda e=exc: self._iv_status_var.set(
                    f"Fehler: {e}") if hasattr(self, "_iv_status_var") else None)
                return
            if not os.path.exists(out_xl):
                self._iv_log_write("Keine Ausgabedatei – Verbindung pruefen.")
                self.root.after(0, lambda: self._iv_status_var.set(
                    "Keine Ausgabedatei – Verbindung pruefen.")
                    if hasattr(self, "_iv_status_var") else None)
                return
            self._iv_log_write(f"Importiere {os.path.basename(out_xl)} ...")
            try:
                import openpyxl as _opx, sqlite3 as _sq3
                wb = _opx.load_workbook(out_xl, read_only=True, data_only=True)
                xl_rows = list(wb.active.iter_rows(min_row=2, values_only=True))
                wb.close()
                tconn = _sq3.connect(DB_PATH)
                col_defs = ", ".join(f'"{c}" TEXT' for c in SAP_IV_COLUMNS)
                tconn.execute(f"CREATE TABLE IF NOT EXISTS sap_iv ({col_defs})")
                # Unique-Index für Upsert anlegen (Rechnungsnummer + Jahr + Position)
                has_uq = tconn.execute(
                    "SELECT name FROM sqlite_master WHERE type='index' "
                    "AND tbl_name='sap_iv' AND name='sap_iv_uq'"
                ).fetchone() is not None
                if not has_uq:
                    try:
                        tconn.execute(
                            'CREATE UNIQUE INDEX sap_iv_uq '
                            'ON sap_iv("Rechnungsnummer","Jahr","Position")'
                        )
                    except Exception:
                        pass  # Bestehende Duplikate – Index wird später angelegt
                # Upsert: vorhandene Zeilen aktualisieren, neue einfügen
                ph = ", ".join("?" * len(SAP_IV_COLUMNS))
                inserted = updated = 0
                for row in xl_rows:
                    padded = list(row) + [None] * max(0, len(SAP_IV_COLUMNS) - len(row))
                    tconn.execute(f"INSERT OR REPLACE INTO sap_iv VALUES ({ph})",
                                  padded[:len(SAP_IV_COLUMNS)])
                    if tconn.execute("SELECT changes()").fetchone()[0] == 1:
                        inserted += 1
                    else:
                        updated += 1
                tconn.commit(); tconn.close()
                n = len(xl_rows)
                self._iv_log_write(
                    f"OK \u2013 {n} Zeilen verarbeitet "
                    f"({inserted} neu, {n - inserted} aktualisiert)."
                )
                self.root.after(0, lambda r=n: (
                    self._iv_status_var.set(f"{r} Zeilen importiert.")
                    if hasattr(self, "_iv_status_var") else None))
                self.root.after(0, self._load_sap_iv_data)
                try:
                    os.remove(out_xl)
                except Exception:
                    pass
            except Exception as exc:
                self._iv_log_write(f"Import-Fehler: {exc}")
                self.root.after(0, lambda e=exc: (
                    self._iv_status_var.set(f"Import-Fehler: {e}")
                ) if hasattr(self, "_iv_status_var") else None)

        threading.Thread(target=_run, daemon=True).start()

    # -----------------------------------------------------------------------
    # CAPE Entry Summary Tab
    # -----------------------------------------------------------------------
    def _build_cape_tab(self):
        """CBP-Report > CAPE Entry Summary – Excel Drag & Drop Import + Tabelle."""
        parent = self.tab_cape
        self._cape_status_var = tk.StringVar(value="Bitte Excel-Datei hier ablegen …")

        # ── Top-Toolbar ──────────────────────────────────────────────────────
        toolbar = tk.Frame(parent, bg=CARD_BG, height=54)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        btn_reload = tk.Button(
            toolbar, text="↻  Aktualisieren", command=self._load_cape_data,
            bg=SIDEBAR_ACT, fg="white", relief="flat", padx=14, pady=6,
            font=("Segoe UI", 10, "bold"), cursor="hand2",
        )
        btn_reload.pack(side="left", padx=16, pady=10)

        btn_clear = tk.Button(
            toolbar, text="🗑  Tabelle leeren", command=self._cape_clear_table,
            bg="#D32F2F", fg="white", relief="flat", padx=14, pady=6,
            font=("Segoe UI", 10), cursor="hand2",
        )
        btn_clear.pack(side="left", padx=(0, 16), pady=10)

        tk.Label(toolbar, textvariable=self._cape_status_var,
                 bg=CARD_BG, fg="#5A6A80", font=("Segoe UI", 10)).pack(
            side="left", padx=8)

        tk.Frame(parent, bg="#E4E8EF", height=1).pack(fill="x")

        # ── Drag & Drop Zone ─────────────────────────────────────────────────
        drop_frame = tk.Frame(parent, bg="#EAF3FB", height=90)
        drop_frame.pack(fill="x", padx=20, pady=12)
        drop_frame.pack_propagate(False)

        drop_lbl = tk.Label(
            drop_frame,
            text="📊   Excel-Datei hier ablegen  (CAPE Entry Summary Report – Main Report)",
            bg="#EAF3FB", fg="#1A5B8A",
            font=("Segoe UI", 12, "bold"),
        )
        drop_lbl.place(relx=0.5, rely=0.5, anchor="center")

        # TkinterDnD Drag-&-Drop registrieren
        try:
            from tkinterdnd2 import DND_FILES
            for w in (drop_frame, drop_lbl):
                w.drop_target_register(DND_FILES)
                w.dnd_bind("<<Drop>>", self._cape_on_drop)
        except Exception:
            pass  # DnD nicht verfügbar – kein Absturz

        # ── Suchleiste + Spaltenfilter ────────────────────────────────────────
        sf_outer = tk.Frame(parent, bg=MAIN_BG)
        sf_outer.pack(fill="x", padx=20, pady=(0, 4))

        sf = tk.Frame(sf_outer, bg=MAIN_BG)
        sf.pack(fill="x")
        tk.Label(sf, text="🔍", bg=MAIN_BG, fg="#8A9EB5",
                 font=("Segoe UI", 12)).pack(side="left", padx=(0, 6))
        self._cape_search_var = tk.StringVar()
        self._cape_search_var.trace_add("write", lambda *_: self._cape_filter())
        tk.Entry(sf, textvariable=self._cape_search_var, relief="flat",
                 font=("Segoe UI", 11), bg="white", fg="#1A2742",
                 insertbackground="#1A2742").pack(side="left", fill="x",
                                                  expand=True, ipady=5)
        self._cape_toggle_lbl = tk.StringVar(value="▶ Spaltenfilter")
        tk.Button(sf, textvariable=self._cape_toggle_lbl,
                  command=self._cape_toggle_panel,
                  bg="#E4E8EF", fg="#1A2742", relief="flat", padx=8, pady=3,
                  font=("Segoe UI", 9), cursor="hand2").pack(side="left", padx=(6, 4))
        tk.Button(sf, text="✕", command=self._cape_reset_filter,
                  bg="#E4E8EF", fg="#1A2742", relief="flat", padx=8, pady=3,
                  font=("Segoe UI", 10), cursor="hand2").pack(side="left")

        self._cape_panel_open = False
        self._cape_panel, self._cape_get_col_filters, self._cape_reset_col = (
            _build_col_filter_panel(sf_outer, CAPE_HEADERS, self._cape_filter, cols_per_row=5))

        # ── Tabelle (Treeview) ───────────────────────────────────────────────
        tv_frame = tk.Frame(parent, bg=MAIN_BG)
        tv_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        vsb = ttk.Scrollbar(tv_frame, orient="vertical")
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal")
        self._cape_tv = ttk.Treeview(
            tv_frame,
            columns=CAPE_COLUMNS,
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._cape_tv.yview)
        hsb.config(command=self._cape_tv.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._cape_tv.pack(side="left", fill="both", expand=True)

        # Spaltenbreiten
        col_widths = {
            "claim_number": 110, "claim_status": 90, "filer_code": 70,
            "importer_number": 110, "cf4811_number": 100, "cf4811_name": 180,
            "entry_summary_number": 130, "entry_date": 90,
            "liquidation_date": 100, "control_team_number": 80,
            "center_id_code": 80, "center_id_name": 120,
            "refund_amount": 90, "interest_amount": 90,
            "refund_number": 100, "refund_date": 90, "refund_status": 90,
        }
        for col, hdr in zip(CAPE_COLUMNS, CAPE_HEADERS):
            self._cape_tv.heading(col, text=hdr,
                                  command=lambda c=col: self._cape_sort(c))
            self._cape_tv.column(col, width=col_widths.get(col, 100),
                                 minwidth=60, stretch=False, anchor="w")

        # Zebra-Stripes
        self._cape_tv.tag_configure("odd",  background="#F7F9FC")
        self._cape_tv.tag_configure("even", background="white")
        self._cape_tv.tag_configure("grand_total", background="#1F4E79",
                                    foreground="white", font=("Segoe UI", 9, "bold"))

        # Alternating-row Farben bei Selektion
        self._cape_tv.tag_configure("selected", background=SIDEBAR_ACT, foreground="white")

        # DB-Tabelle sicherstellen & Daten laden
        self._cape_ensure_db()
        self._load_cape_data()
        self._cape_sort_col = None
        self._cape_sort_rev = False

    # -----------------------------------------------------------------------
    def _cape_ensure_db(self):
        """Erstellt die cape_entry_summary-Tabelle falls nicht vorhanden."""
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cols_def = ", ".join(f"{c} TEXT" for c in CAPE_COLUMNS)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS cape_entry_summary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_date TEXT DEFAULT (date('now')),
                {cols_def}
            )
        """)
        conn.commit()
        conn.close()

    def _cape_on_drop(self, event):
        """Wird aufgerufen wenn eine Datei in die Drop-Zone gezogen wird."""
        raw = event.data.strip()
        # TkinterDnD liefert geschweifte Klammern bei Leerzeichen im Pfad
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw
        if not os.path.isfile(path):
            self._cape_status_var.set(f"Datei nicht gefunden: {path}")
            return
        if not path.lower().endswith((".xlsx", ".xls")):
            self._cape_status_var.set("Nur Excel-Dateien (.xlsx / .xls) werden unterstützt.")
            return
        self._cape_status_var.set(f"Importiere: {os.path.basename(path)} …")
        threading.Thread(target=self._cape_import_excel, args=(path,),
                         daemon=True).start()

    def _cape_import_excel(self, path):
        """Excel-Datei in DB importieren (läuft im Hintergrund-Thread)."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            # Sheet "Main Report" suchen (case-insensitive)
            sheet_name = None
            for n in wb.sheetnames:
                if n.strip().lower() == "main report":
                    sheet_name = n
                    break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]  # Fallback: erstes Sheet
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                self.root.after(0, lambda: self._cape_status_var.set("Excel-Datei ist leer."))
                return

            # Erste Zeile = Header, rest = Daten
            data_rows = rows[1:]  # Zeile 1 = Spaltenköpfe überspringen

            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            inserted = 0
            source = os.path.basename(path)
            for row in data_rows:
                # Auf 17 Spalten auffüllen / kürzen
                vals = list(row[:17])
                while len(vals) < 17:
                    vals.append(None)
                # Leere Zeilen überspringen
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                # Werte normalisieren
                norm = []
                for v in vals:
                    if v is None:
                        norm.append(None)
                    else:
                        s = str(v).strip()
                        norm.append(s if s else None)
                cols_sql = ", ".join(CAPE_COLUMNS)
                placeholders = ", ".join(["?"] * len(CAPE_COLUMNS))
                cur.execute(
                    f"INSERT INTO cape_entry_summary ({cols_sql}) VALUES ({placeholders})",
                    norm,
                )
                inserted += 1
            conn.commit()
            conn.close()
            self.root.after(0, lambda n=inserted, s=source: (
                self._cape_status_var.set(f"✓ {n} Zeilen importiert aus '{s}'."),
                self._load_cape_data(),
            ))
        except ImportError:
            self.root.after(0, lambda: self._cape_status_var.set(
                "openpyxl nicht installiert – bitte: pip install openpyxl"))
        except Exception as exc:
            self.root.after(0, lambda e=exc: self._cape_status_var.set(
                f"Import-Fehler: {e}"))

    def _load_cape_data(self, *_):
        """Daten aus DB laden und Treeview füllen."""
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute(
                f"SELECT {', '.join(CAPE_COLUMNS)} FROM cape_entry_summary ORDER BY id DESC"
            )
            rows = cur.fetchall()
            conn.close()
        except Exception:
            rows = []
        self._cape_all_rows = rows
        self._cape_filter()

    def _cape_toggle_panel(self):
        if self._cape_panel_open:
            self._cape_panel.pack_forget()
            self._cape_panel_open = False
            self._cape_toggle_lbl.set("▶ Spaltenfilter")
        else:
            self._cape_panel.pack(fill="x", pady=(2, 0))
            self._cape_panel_open = True
            self._cape_toggle_lbl.set("▼ Spaltenfilter")

    def _cape_reset_filter(self):
        self._cape_search_var.set("")
        self._cape_reset_col()
        self._cape_filter()
        if self._cape_panel_open:
            self._cape_toggle_panel()

    def _cape_filter(self, *_):
        """Treeview mit AND-Mehrfach-Filter befuellen."""
        term = self._cape_search_var.get().lower().strip()
        col_filters = self._cape_get_col_filters()
        tv = self._cape_tv
        for item in tv.get_children():
            tv.delete(item)
        count = 0
        _visible = []
        for row in self._cape_all_rows:
            if not _row_passes(row, col_filters, term):
                continue
            _visible.append(row)
            tag = "even" if count % 2 == 0 else "odd"
            tv.insert("", "end", values=row, tags=(tag,))
            count += 1
        self._cape_insert_grand_total(_visible)
        total = len(self._cape_all_rows)
        active = len(col_filters)
        badge = f"  [{active} Spaltenfilter]" if active else ""
        if term or active:
            self._cape_status_var.set(f"{count} von {total} Eintraegen{badge} angezeigt.")
        else:
            self._cape_status_var.set(f"{total} Eintraege in der Datenbank.")

    def _cape_insert_grand_total(self, rows):
        """Gesamtsummen-Zeile: Refund Amt. + Interest Amt."""
        if not rows:
            return
        _SUM_IDX = [
            (CAPE_COLUMNS.index("refund_amount"),   "Refund Amt."),
            (CAPE_COLUMNS.index("interest_amount"), "Interest Amt."),
        ]
        sums = {ci: 0.0 for ci, _ in _SUM_IDX}
        for row in rows:
            for ci, _ in _SUM_IDX:
                parsed = _parse_de(row[ci] if ci < len(row) else None)
                if parsed is not None:
                    sums[ci] += parsed
        footer = [""] * len(CAPE_COLUMNS)
        footer[0] = "\u2211 GESAMT"
        for ci, _ in _SUM_IDX:
            footer[ci] = _de_num(sums[ci])
        self._cape_tv.insert("", "end", values=footer, tags=("grand_total",))

    def _cape_sort(self, col):
        """Treeview nach Spalte sortieren (Toggle asc/desc)."""
        if self._cape_sort_col == col:
            self._cape_sort_rev = not self._cape_sort_rev
        else:
            self._cape_sort_col = col
            self._cape_sort_rev = False
        idx = CAPE_COLUMNS.index(col)
        self._cape_all_rows.sort(
            key=lambda r: (r[idx] is None, str(r[idx] or "").lower()),
            reverse=self._cape_sort_rev,
        )
        self._cape_filter()

    def _cape_clear_table(self):
        """Alle Einträge aus der DB-Tabelle löschen (nach Bestätigung)."""
        import tkinter.messagebox as mb
        if not mb.askyesno(
            "Tabelle leeren",
            "Alle CAPE-Entry-Summary-Daten aus der Datenbank löschen?\nDieser Vorgang kann nicht rückgängig gemacht werden.",
        ):
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM cape_entry_summary")
            conn.commit()
            conn.close()
        except Exception as exc:
            self._cape_status_var.set(f"Fehler: {exc}")
            return
        self._load_cape_data()

    # -----------------------------------------------------------------------
    # Stammdaten – ZuOrd-MAT-KD Tab
    # -----------------------------------------------------------------------
    def _build_zuord_mat_kd_tab(self):
        """Stammdaten > ZuOrd-MAT-KD – Excel Drag & Drop Import + Tabelle."""
        parent = self.tab_zuord_mat_kd
        outer  = tk.Frame(parent, bg=MAIN_BG)
        outer.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = tk.Frame(outer, bg=MAIN_BG)
        toolbar.pack(fill="x", pady=(0, 6))
        self._zuord_status_var = tk.StringVar(value="Bitte Excel-Datei hier ablegen …")

        tk.Button(toolbar, text="🗑  Tabelle leeren",
                  command=self._zuord_clear_table,
                  bg="#DC2626", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Label(toolbar, textvariable=self._zuord_status_var,
                 bg=MAIN_BG, fg="#1F4E79", font=("Segoe UI", 9)).pack(side="left")

        # ── Drop-Zone ────────────────────────────────────────────────────────
        drop_frame = tk.LabelFrame(
            outer,
            text="Excel-Import  (Sheet: StammdatenKundenMAT, Header Zeile 1)",
            bg=MAIN_BG, fg="#5A6A80", font=("Segoe UI", 9, "bold"),
            padx=10, pady=8)
        drop_frame.pack(fill="x", pady=(0, 8))

        drop_lbl = tk.Label(
            drop_frame,
            text="📂  Excel-Datei hier ablegen (.xlsx / .xls)  |  Spalten: Mat-No. SAP | Material CAT-SAP | Kunden-Nr | SAP Customer Name",
            bg="#EFF6FF", fg="#1D4ED8",
            font=("Segoe UI", 10), relief="groove",
            padx=20, pady=18, cursor="hand2")
        drop_lbl.pack(fill="x")
        try:
            drop_lbl.drop_target_register("DND_Files")
            drop_lbl.dnd_bind("<<Drop>>", self._zuord_on_drop)
        except Exception:
            pass

        # ── Suchleiste + Spaltenfilter ────────────────────────────────────────
        sf_outer = tk.Frame(outer, bg=MAIN_BG)
        sf_outer.pack(fill="x", pady=(0, 4))
        sf = tk.Frame(sf_outer, bg=MAIN_BG)
        sf.pack(fill="x")

        self._zuord_search_var = tk.StringVar()
        self._zuord_search_var.trace_add("write", lambda *_: self._zuord_filter())
        tk.Label(sf, text="Suche:", bg=MAIN_BG, fg="#5A6A80",
                 font=("Segoe UI", 9)).pack(side="left")
        ttk.Entry(sf, textvariable=self._zuord_search_var, width=26).pack(
            side="left", padx=(4, 10))

        self._zuord_toggle_lbl = tk.StringVar(value="▶ Spaltenfilter")
        tk.Button(sf, textvariable=self._zuord_toggle_lbl,
                  command=self._zuord_toggle_panel,
                  bg="#E8EDF4", fg="#1F4E79", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left")
        tk.Button(sf, text="✕", command=self._zuord_reset_filter,
                  bg="#E8EDF4", fg="#6B7280", font=("Segoe UI", 9),
                  relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left", padx=(2, 0))

        self._zuord_panel_open = False
        self._zuord_panel, self._zuord_get_col_filters, self._zuord_reset_col = (
            _build_col_filter_panel(sf_outer, ZUORD_HEADERS, self._zuord_filter, cols_per_row=4))

        # ── Treeview ─────────────────────────────────────────────────────────
        tv_frame = tk.Frame(outer, bg=MAIN_BG)
        tv_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(tv_frame, orient="vertical")
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal")
        self._zuord_tv = ttk.Treeview(
            tv_frame,
            columns=ZUORD_COLUMNS,
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._zuord_tv.yview)
        hsb.config(command=self._zuord_tv.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._zuord_tv.pack(side="left", fill="both", expand=True)

        col_widths = {
            "mat_no_sap": 180, "material_cat_sap": 280,
            "kunden_nr": 100, "sap_customer_name": 280,
        }
        for col, hdr in zip(ZUORD_COLUMNS, ZUORD_HEADERS):
            self._zuord_tv.heading(col, text=hdr,
                                   command=lambda c=col: self._zuord_sort(c))
            self._zuord_tv.column(col, width=col_widths.get(col, 150),
                                  minwidth=60, stretch=False, anchor="w")

        self._zuord_tv.tag_configure("odd",  background="#F7F9FC")
        self._zuord_tv.tag_configure("even", background="white")
        self._zuord_tv.tag_configure("grand_total", background="#1F4E79",
                                     foreground="white", font=("Segoe UI", 9, "bold"))

        self._zuord_ensure_db()
        self._load_zuord_data()
        self._zuord_sort_col = None
        self._zuord_sort_rev = False

    def _zuord_ensure_db(self):
        conn = sqlite3.connect(DB_PATH)
        conn.executescript(ZUORD_SCHEMA)
        conn.commit()
        conn.close()

    def _zuord_on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw
        if not os.path.isfile(path):
            self._zuord_status_var.set(f"Datei nicht gefunden: {path}")
            return
        if not path.lower().endswith((".xlsx", ".xls")):
            self._zuord_status_var.set("Nur Excel-Dateien (.xlsx / .xls) werden unterstützt.")
            return
        self._zuord_status_var.set(f"Importiere: {os.path.basename(path)} …")
        threading.Thread(target=self._zuord_import_excel, args=(path,),
                         daemon=True).start()

    def _zuord_import_excel(self, path):
        try:
            import openpyxl
            from datetime import datetime as _dt
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            # Sheet suchen
            sheet_name = None
            for n in wb.sheetnames:
                if "stamm" in n.lower() or "zuord" in n.lower():
                    sheet_name = n
                    break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Header Zeile 1 (index 0), Daten ab Zeile 2 (index 1)
            data_rows = rows[1:]
            conn = sqlite3.connect(DB_PATH)
            cur  = conn.cursor()
            cur.execute("DELETE FROM stammdaten_zuord_mat_kd")
            inserted = 0
            source = os.path.basename(path)
            for row in data_rows:
                vals = list(row[:4])
                while len(vals) < 4:
                    vals.append(None)
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                norm = []
                for v in vals:
                    if v is None:
                        norm.append(None)
                    elif isinstance(v, float) and v == int(v):
                        norm.append(str(int(v)))
                    else:
                        s = str(v).strip()
                        norm.append(s if s else None)
                cur.execute(
                    "INSERT INTO stammdaten_zuord_mat_kd "
                    "(mat_no_sap, material_cat_sap, kunden_nr, sap_customer_name) "
                    "VALUES (?, ?, ?, ?)",
                    norm,
                )
                inserted += 1
            conn.commit()
            conn.close()
            self.root.after(0, lambda n=inserted, s=source: (
                self._zuord_status_var.set(f"✓ {n} Zeilen importiert aus '{s}' (Tabelle ersetzt)."),
                self._load_zuord_data(),
            ))
        except ImportError:
            self.root.after(0, lambda: self._zuord_status_var.set(
                "openpyxl nicht installiert – bitte: pip install openpyxl"))
        except Exception as exc:
            self.root.after(0, lambda e=exc: self._zuord_status_var.set(
                f"Import-Fehler: {e}"))

    def _load_zuord_data(self, *_):
        try:
            conn = sqlite3.connect(DB_PATH)
            cur  = conn.cursor()
            cur.execute("SELECT " + ", ".join(ZUORD_COLUMNS) +
                        " FROM stammdaten_zuord_mat_kd ORDER BY id DESC")
            rows = cur.fetchall()
            conn.close()
        except Exception:
            rows = []
        self._zuord_all_rows = rows
        self._zuord_filter()

    def _zuord_toggle_panel(self):
        if self._zuord_panel_open:
            self._zuord_panel.pack_forget()
            self._zuord_panel_open = False
            self._zuord_toggle_lbl.set("▶ Spaltenfilter")
        else:
            self._zuord_panel.pack(fill="x", pady=(2, 0))
            self._zuord_panel_open = True
            self._zuord_toggle_lbl.set("▼ Spaltenfilter")

    def _zuord_reset_filter(self):
        self._zuord_search_var.set("")
        self._zuord_reset_col()
        self._zuord_filter()
        if self._zuord_panel_open:
            self._zuord_toggle_panel()

    def _zuord_filter(self, *_):
        term = self._zuord_search_var.get().lower().strip()
        col_filters = self._zuord_get_col_filters()
        tv = self._zuord_tv
        for item in tv.get_children():
            tv.delete(item)
        count = 0
        for row in getattr(self, "_zuord_all_rows", []):
            if not _row_passes(row, col_filters, term):
                continue
            tag = "even" if count % 2 == 0 else "odd"
            tv.insert("", "end", values=row, tags=(tag,))
            count += 1
        total = len(getattr(self, "_zuord_all_rows", []))
        active = len(col_filters)
        badge = f"  [{active} Spaltenfilter]" if active else ""
        if term or active:
            self._zuord_status_var.set(f"{count} von {total} Einträgen{badge} angezeigt.")
        else:
            self._zuord_status_var.set(f"{total} Einträge in der Datenbank.")

    def _zuord_sort(self, col):
        if self._zuord_sort_col == col:
            self._zuord_sort_rev = not self._zuord_sort_rev
        else:
            self._zuord_sort_col = col
            self._zuord_sort_rev = False
        idx = ZUORD_COLUMNS.index(col)
        self._zuord_all_rows.sort(
            key=lambda r: (r[idx] is None, str(r[idx] or "").lower()),
            reverse=self._zuord_sort_rev,
        )
        self._zuord_filter()

    def _zuord_clear_table(self):
        import tkinter.messagebox as mb
        if not mb.askyesno(
            "Tabelle leeren",
            "Alle ZuOrd-MAT-KD-Daten aus der Datenbank löschen?\nDieser Vorgang kann nicht rückgängig gemacht werden.",
        ):
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM stammdaten_zuord_mat_kd")
            conn.commit()
            conn.close()
        except Exception as exc:
            self._zuord_status_var.set(f"Fehler: {exc}")
            return
        self._load_zuord_data()


    # -----------------------------------------------------------------------
    # db_Tariff Tab
    # -----------------------------------------------------------------------
    def _build_db_tariff_tab(self):
        """Stammdaten > db_Tariff – Excel Drag & Drop Import + Tabelle."""
        parent = self.tab_db_tariff
        outer  = tk.Frame(parent, bg=MAIN_BG)
        outer.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = tk.Frame(outer, bg=MAIN_BG)
        toolbar.pack(fill="x", pady=(0, 6))
        self._dbt_status_var = tk.StringVar(value="Bitte Excel-Datei hier ablegen …")

        tk.Button(toolbar, text="📊  Excel Report",
                  command=self._dbt_export_excel,
                  bg="#166534", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="🗑  Tabelle leeren",
                  command=self._dbt_clear_table,
                  bg="#DC2626", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="🔍  Abgleich Claim-Report",
                  command=self._dbt_run_abgleich,
                  bg="#1D4ED8", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(toolbar, text="✕ Abgleich zurücksetzen",
                  command=self._dbt_reset_abgleich,
                  bg="#6B7280", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=4, cursor="hand2").pack(side="left", padx=(0, 12))
        tk.Label(toolbar, textvariable=self._dbt_status_var,
                 bg=MAIN_BG, fg="#1F4E79", font=("Segoe UI", 9)).pack(side="left")

        # ── Drop-Zone ────────────────────────────────────────────────────────
        drop_frame = tk.LabelFrame(
            outer,
            text="Excel-Import  (Sheet: Data, Header Zeile 1, Daten ab Zeile 2)",
            bg=MAIN_BG, fg="#5A6A80", font=("Segoe UI", 9, "bold"),
            padx=10, pady=8)
        drop_frame.pack(fill="x", pady=(0, 8))

        drop_lbl = tk.Label(
            drop_frame,
            text="📂  Excel-Datei hier ablegen (.xlsx / .xls)  |  64 Spalten: Year/Month, Entry-No, HTS-Codes, Vendor, Mat-No. SAP, Claim Values …",
            bg="#EFF6FF", fg="#1D4ED8",
            font=("Segoe UI", 10), relief="groove",
            padx=20, pady=18, cursor="hand2")
        drop_lbl.pack(fill="x")
        try:
            drop_lbl.drop_target_register("DND_Files")
            drop_lbl.dnd_bind("<<Drop>>", self._dbt_on_drop)
        except Exception:
            pass

        # ── Suchleiste + Spaltenfilter ────────────────────────────────────────
        sf_outer = tk.Frame(outer, bg=MAIN_BG)
        sf_outer.pack(fill="x", pady=(0, 4))
        sf = tk.Frame(sf_outer, bg=MAIN_BG)
        sf.pack(fill="x")

        self._dbt_search_var = tk.StringVar()
        self._dbt_search_var.trace_add("write", lambda *_: self._dbt_filter())
        tk.Label(sf, text="Suche:", bg=MAIN_BG, fg="#5A6A80",
                 font=("Segoe UI", 9)).pack(side="left")
        ttk.Entry(sf, textvariable=self._dbt_search_var, width=26).pack(
            side="left", padx=(4, 10))

        self._dbt_toggle_lbl = tk.StringVar(value="▶ Spaltenfilter")
        tk.Button(sf, textvariable=self._dbt_toggle_lbl,
                  command=self._dbt_toggle_panel,
                  bg="#E8EDF4", fg="#1F4E79", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left")
        tk.Button(sf, text="✕", command=self._dbt_reset_filter,
                  bg="#E8EDF4", fg="#6B7280", font=("Segoe UI", 9),
                  relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left", padx=(2, 0))

        self._dbt_panel_open = False
        self._dbt_panel, self._dbt_get_col_filters, self._dbt_reset_col = (
            _build_col_filter_panel(sf_outer, DB_TARIFF_HEADERS, self._dbt_filter, cols_per_row=6))

        # Referenzen für späteren Abgleich-Panel-Toggle
        self._dbt_outer = outer

        # ── Detail-Panel (Abgleich-Bericht) – unten eingeblendet bei Abgleich ──
        self._dbt_detail_outer = tk.Frame(outer, bg=MAIN_BG)
        # wird initial NICHT gepackt
        _dlf = tk.LabelFrame(
            self._dbt_detail_outer,
            text="📊  Abgleich-Bericht  (Gründe für Unterschiede)",
            bg=MAIN_BG, fg="#1F4E79", font=("Segoe UI", 9, "bold"),
            padx=6, pady=4)
        _dlf.pack(fill="both", expand=True)
        _d_vsb = ttk.Scrollbar(_dlf, orient="vertical")
        _d_hsb = ttk.Scrollbar(_dlf, orient="horizontal")
        self._dbt_detail_text = tk.Text(
            _dlf, height=11,
            bg="#F8FAFC", fg="#1A2742",
            font=("Consolas", 9), relief="flat",
            state="disabled", wrap="none",
            yscrollcommand=_d_vsb.set,
            xscrollcommand=_d_hsb.set)
        _d_vsb.config(command=self._dbt_detail_text.yview)
        _d_hsb.config(command=self._dbt_detail_text.xview)
        _d_vsb.pack(side="right", fill="y")
        _d_hsb.pack(side="bottom", fill="x")
        self._dbt_detail_text.pack(fill="both", expand=True)
        self._dbt_detail_text.tag_configure("h1",     font=("Consolas", 9, "bold"), foreground="#1A2742")
        self._dbt_detail_text.tag_configure("red",    foreground="#991B1B")
        self._dbt_detail_text.tag_configure("orange", foreground="#92400E")
        self._dbt_detail_text.tag_configure("green",  foreground="#065F46")
        self._dbt_detail_text.tag_configure("grey",   foreground="#6B7280")
        self._dbt_detail_text.tag_configure("bold",   font=("Consolas", 9, "bold"))

        # ── Treeview ─────────────────────────────────────────────────────────
        self._dbt_tv_frame = tk.Frame(outer, bg=MAIN_BG)
        tv_frame = self._dbt_tv_frame
        tv_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(tv_frame, orient="vertical")
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal")
        self._dbt_tv = ttk.Treeview(
            tv_frame,
            columns=DB_TARIFF_COLUMNS,
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._dbt_tv.yview)
        hsb.config(command=self._dbt_tv.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._dbt_tv.pack(side="left", fill="both", expand=True)

        # Column widths – wider for descriptive cols, narrow for HTS amounts
        _wide = {"year_month": 90, "dachser_duty_doc_no": 130, "entry_no_new": 120,
                 "line_item": 70, "vendor_name": 200, "sap_customer_name": 200,
                 "material_cat_sap": 250, "mat_no_sap": 140, "invoice_no_delivery_note": 160,
                 "po_order_no": 130, "customer_part_no": 140, "claim_status": 100,
                 "dachser_dokument": 120, "entry_document": 120}
        for col, hdr in zip(DB_TARIFF_COLUMNS, DB_TARIFF_HEADERS):
            self._dbt_tv.heading(col, text=hdr,
                                 command=lambda c=col: self._dbt_sort(c))
            w = _wide.get(col, 90)
            self._dbt_tv.column(col, width=w, minwidth=50, stretch=False, anchor="w")

        self._dbt_tv.tag_configure("odd",  background="#F7F9FC")
        self._dbt_tv.tag_configure("even", background="white")
        self._dbt_tv.tag_configure("grand_total", background="#1F4E79",
                                   foreground="white", font=("Segoe UI", 9, "bold"))
        # Abgleich-Tags
        self._dbt_tv.tag_configure("abgleich_fehlt",
                                   background="#FEE2E2", foreground="#991B1B")
        self._dbt_tv.tag_configure("abgleich_ok",
                                   background="#D1FAE5", foreground="#065F46")
        self._dbt_tv.tag_configure("abgleich_nur_db",
                                   background="#FEF9C3", foreground="#78350F")

        # Abgleich-Zustand
        self._dbt_abgleich_active = False
        self._dbt_abgleich_map    = {}   # norm_entry → "ok" | "fehlt" | "nur_db"

        self._dbt_ensure_db()
        self._load_dbt_data()
        self._dbt_sort_col = None
        self._dbt_sort_rev = False

    # ═══════════════════════════════════════════════════════════════════════
    # HTSUS CLAIM CONFIG – Konfiguration Erstattungsfähig?
    # ═══════════════════════════════════════════════════════════════════════
    def _build_htsus_claim_cfg_tab(self):
        """STAMMDATEN > HTSUS Claim – Konfiguration welche HTSUS-Nr. für Claim Value 1 gilt."""
        parent = self.tab_htsus_claim_cfg
        outer  = tk.Frame(parent, bg=MAIN_BG)
        outer.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Header ───────────────────────────────────────────────────────────
        hdr = tk.Frame(outer, bg=MAIN_BG)
        hdr.pack(fill="x", pady=(0, 8))
        tk.Label(hdr, text="⚙️  HTSUS Claim Konfiguration",
                 bg=MAIN_BG, fg="#1F4E79",
                 font=("Segoe UI", 13, "bold")).pack(side="left")

        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = tk.Frame(outer, bg=MAIN_BG)
        toolbar.pack(fill="x", pady=(0, 6))

        tk.Button(toolbar, text="📂  Excel importieren",
                  command=self._hcc_import_excel,
                  bg="#1D4ED8", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2"
                  ).pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="➕  Zeile hinzufügen",
                  command=self._hcc_add_row,
                  bg="#166534", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2"
                  ).pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="🗑  Zeile löschen",
                  command=self._hcc_delete_row,
                  bg="#DC2626", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2"
                  ).pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="🔄  Aktualisieren",
                  command=self._hcc_load,
                  bg="#6B7280", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=4, cursor="hand2"
                  ).pack(side="left", padx=(0, 16))

        self._hcc_status_var = tk.StringVar(value="")
        tk.Label(toolbar, textvariable=self._hcc_status_var,
                 bg=MAIN_BG, fg="#1F4E79", font=("Segoe UI", 9)).pack(side="left")

        # ── Info-Banner ───────────────────────────────────────────────────────
        info = tk.Label(outer,
            text='ℹ️  Zeilen mit "Erstattungsfähig? = ja" werden in der Berechnung Claim Value 1 (Auswertung) berücksichtigt. '
                 'Doppelklick auf eine Zelle zum Bearbeiten.',
            bg="#EFF6FF", fg="#1D4ED8",
            font=("Segoe UI", 9), relief="groove",
            padx=10, pady=6, anchor="w", justify="left")
        info.pack(fill="x", pady=(0, 8))

        # ── Suchleiste ────────────────────────────────────────────────────────
        sf = tk.Frame(outer, bg=MAIN_BG)
        sf.pack(fill="x", pady=(0, 4))
        self._hcc_search_var = tk.StringVar()
        self._hcc_search_var.trace_add("write", lambda *_: self._hcc_apply_filter())
        tk.Label(sf, text="Suche:", bg=MAIN_BG, fg="#5A6A80",
                 font=("Segoe UI", 9)).pack(side="left")
        ttk.Entry(sf, textvariable=self._hcc_search_var, width=26
                  ).pack(side="left", padx=(4, 8))
        tk.Button(sf, text="✕", command=lambda: self._hcc_search_var.set(""),
                  bg="#E8EDF4", fg="#6B7280", font=("Segoe UI", 9),
                  relief="flat", padx=6, pady=2, cursor="hand2"
                  ).pack(side="left")

        # ── Tabelle ───────────────────────────────────────────────────────────
        tbl_frame = tk.Frame(outer, bg=MAIN_BG)
        tbl_frame.pack(fill="both", expand=True)

        _HCC_COLS = ("id", "htsus_no", "htsus_rate", "erstattungsfaehig",
                     "comment_1", "comment_2", "comment_3")
        _HCC_HDRS = ("ID", "HTSUS-Nr.", "HTSUS-Satz", "Erstattungsfähig?",
                     "Comment 1", "Comment 2", "Comment 3")
        _HCC_WIDTHS = (40, 110, 80, 120, 280, 220, 180)

        style = ttk.Style()
        style.configure("HCC.Treeview",
                        rowheight=22, font=("Segoe UI", 9),
                        background="#FFFFFF", fieldbackground="#FFFFFF")
        style.configure("HCC.Treeview.Heading",
                        font=("Segoe UI", 9, "bold"), background="#1F4E79",
                        foreground="white")
        style.map("HCC.Treeview", background=[
            ("selected", "#DBEAFE"),
            ("!selected", ""),
        ])

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical")
        hsb = ttk.Scrollbar(tbl_frame, orient="horizontal")
        self._hcc_tv = ttk.Treeview(
            tbl_frame, columns=_HCC_COLS, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set,
            style="HCC.Treeview", selectmode="browse")
        vsb.config(command=self._hcc_tv.yview)
        hsb.config(command=self._hcc_tv.xview)

        for col, hdr, w in zip(_HCC_COLS, _HCC_HDRS, _HCC_WIDTHS):
            self._hcc_tv.heading(col, text=hdr,
                command=lambda c=col: self._hcc_sort(c))
            self._hcc_tv.column(col, width=w, minwidth=40,
                                stretch=(col in ("comment_1", "comment_2")))

        # Farbmarkierung: ja=grün, nein=rot, None=grau
        self._hcc_tv.tag_configure("ja",   background="#D1FAE5", foreground="#065F46")
        self._hcc_tv.tag_configure("nein", background="#FEE2E2", foreground="#991B1B")
        self._hcc_tv.tag_configure("none", background="#F3F4F6", foreground="#6B7280")

        self._hcc_tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tbl_frame.rowconfigure(0, weight=1)
        tbl_frame.columnconfigure(0, weight=1)

        # Doppelklick zum Bearbeiten
        self._hcc_tv.bind("<Double-1>", self._hcc_on_double_click)

        # Interne Daten
        self._hcc_all_rows = []
        self._hcc_sort_col = None
        self._hcc_sort_rev = False

        self._hcc_load()

    def _hcc_load(self):
        """Daten aus htsus_claim_config laden."""
        with sqlite3.connect(DB_PATH) as con:
            rows = con.execute(
                "SELECT id, htsus_no, htsus_rate, erstattungsfaehig, "
                "       comment_1, comment_2, comment_3 "
                "FROM htsus_claim_config ORDER BY htsus_no"
            ).fetchall()
        self._hcc_all_rows = rows
        self._hcc_apply_filter()
        ja_count = sum(1 for r in rows if r[3] == "ja")
        self._hcc_status_var.set(
            f"{len(rows)} Zeilen  |  {ja_count} Erstattungsfähig")

    def _hcc_apply_filter(self):
        """Suchfilter anwenden."""
        q = self._hcc_search_var.get().lower().strip()
        filtered = [r for r in self._hcc_all_rows
                    if not q or any(q in str(v).lower() for v in r)]
        self._hcc_tv.delete(*self._hcc_tv.get_children())
        for row in filtered:
            erst = row[3] or ""
            tag  = "ja" if erst == "ja" else ("nein" if erst == "nein" else "none")
            self._hcc_tv.insert("", "end",
                values=[v if v is not None else "" for v in row],
                tags=(tag,))

    def _hcc_sort(self, col):
        """Spalte sortieren."""
        cols = ("id", "htsus_no", "htsus_rate", "erstattungsfaehig",
                "comment_1", "comment_2", "comment_3")
        idx = cols.index(col)
        rev = (self._hcc_sort_col == col and not self._hcc_sort_rev)
        self._hcc_sort_col = col
        self._hcc_sort_rev = rev
        self._hcc_all_rows.sort(
            key=lambda r: (r[idx] is None, str(r[idx] or "").lower()),
            reverse=rev)
        self._hcc_apply_filter()

    def _hcc_on_double_click(self, event):
        """Doppelklick → Zeile bearbeiten (Erstattungsfähig toggle oder Text-Editor)."""
        item = self._hcc_tv.identify_row(event.y)
        col  = self._hcc_tv.identify_column(event.x)
        if not item or not col:
            return
        col_idx = int(col.replace("#", "")) - 1  # 0-basiert
        col_name = ("id", "htsus_no", "htsus_rate", "erstattungsfaehig",
                    "comment_1", "comment_2", "comment_3")[col_idx]
        vals = list(self._hcc_tv.item(item, "values"))
        row_id = int(vals[0])

        if col_name == "erstattungsfaehig":
            # Toggle ja ↔ nein
            current = vals[3]
            new_val = "nein" if current == "ja" else "ja"
            self._hcc_save_cell(row_id, "erstattungsfaehig", new_val)
            vals[3] = new_val
            tag = "ja" if new_val == "ja" else "nein"
            self._hcc_tv.item(item, values=vals, tags=(tag,))
            # Auch in _hcc_all_rows aktualisieren
            for i, r in enumerate(self._hcc_all_rows):
                if r[0] == row_id:
                    lst = list(r); lst[3] = new_val
                    self._hcc_all_rows[i] = tuple(lst)
                    break
            ja_count = sum(1 for r in self._hcc_all_rows if r[3] == "ja")
            self._hcc_status_var.set(
                f"{len(self._hcc_all_rows)} Zeilen  |  {ja_count} Erstattungsfähig")
            # Auswertung automatisch neu berechnen
            if hasattr(self, "_aus_load"):
                self._aus_load()
        elif col_name != "id":
            # Text-Editor als Popup
            self._hcc_edit_cell(item, col, col_idx, col_name, row_id, vals)

    def _hcc_edit_cell(self, item, col, col_idx, col_name, row_id, vals):
        """Inline-Texteditor für eine Zelle."""
        bbox = self._hcc_tv.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox
        var = tk.StringVar(value=vals[col_idx])
        entry = ttk.Entry(self._hcc_tv, textvariable=var, font=("Segoe UI", 9))
        entry.place(x=x, y=y, width=max(w, 100), height=h)
        entry.focus_set()
        entry.select_range(0, "end")

        def _commit(event=None):
            new_val = var.get().strip() or None
            if col_name == "erstattungsfaehig" and new_val not in ("ja", "nein", None):
                new_val = None
            self._hcc_save_cell(row_id, col_name, new_val)
            vals[col_idx] = new_val or ""
            erst = vals[3]
            tag  = "ja" if erst == "ja" else ("nein" if erst == "nein" else "none")
            self._hcc_tv.item(item, values=vals, tags=(tag,))
            for i, r in enumerate(self._hcc_all_rows):
                if r[0] == row_id:
                    lst = list(r); lst[col_idx] = new_val
                    self._hcc_all_rows[i] = tuple(lst)
                    break
            entry.destroy()
            # Bei Änderung des Erstattungsstatus Auswertung neu laden
            if col_name == "erstattungsfaehig" and hasattr(self, "_aus_load"):
                self._aus_load()

        entry.bind("<Return>",  _commit)
        entry.bind("<FocusOut>", _commit)
        entry.bind("<Escape>",  lambda e: entry.destroy())

    def _hcc_save_cell(self, row_id, col_name, value):
        """Einzelne Zelle in DB speichern."""
        allowed = {"htsus_no", "htsus_rate", "erstattungsfaehig",
                   "comment_1", "comment_2", "comment_3"}
        if col_name not in allowed:
            return
        with sqlite3.connect(DB_PATH) as con:
            con.execute(
                f"UPDATE htsus_claim_config SET {col_name}=? WHERE id=?",
                (value, row_id))

    def _hcc_add_row(self):
        """Neue leere Zeile hinzufügen."""
        dlg = tk.Toplevel(self)
        dlg.title("Neue HTSUS-Nr. hinzufügen")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg=MAIN_BG)

        fields = [("HTSUS-Nr. *", "htsus_no"),
                  ("HTSUS-Satz",  "htsus_rate"),
                  ("Erstattungsfähig? (ja/nein)", "erstattungsfaehig"),
                  ("Comment 1", "comment_1"),
                  ("Comment 2", "comment_2"),
                  ("Comment 3", "comment_3")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg=MAIN_BG, fg="#1F4E79",
                     font=("Segoe UI", 9)).grid(row=i, column=0, sticky="w",
                                                 padx=10, pady=4)
            v = tk.StringVar()
            ttk.Entry(dlg, textvariable=v, width=36).grid(
                row=i, column=1, padx=10, pady=4)
            vars_[key] = v

        status_lbl = tk.Label(dlg, text="", bg=MAIN_BG, fg="#DC2626",
                              font=("Segoe UI", 9))
        status_lbl.grid(row=len(fields), column=0, columnspan=2, padx=10)

        def _save():
            htsus_no = vars_["htsus_no"].get().strip()
            if not htsus_no:
                status_lbl.config(text="HTSUS-Nr. ist Pflichtfeld!")
                return
            erst = vars_["erstattungsfaehig"].get().strip().lower() or None
            if erst and erst not in ("ja", "nein"):
                status_lbl.config(text="Erstattungsfähig? nur 'ja' oder 'nein'")
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.execute(
                        "INSERT INTO htsus_claim_config "
                        "(htsus_no, htsus_rate, erstattungsfaehig, comment_1, comment_2, comment_3) "
                        "VALUES (?,?,?,?,?,?)",
                        (htsus_no,
                         vars_["htsus_rate"].get().strip() or None,
                         erst,
                         vars_["comment_1"].get().strip() or None,
                         vars_["comment_2"].get().strip() or None,
                         vars_["comment_3"].get().strip() or None))
                dlg.destroy()
                self._hcc_load()
                if hasattr(self, "_aus_load"):
                    self._aus_load()
            except sqlite3.IntegrityError:
                status_lbl.config(text=f"HTSUS-Nr. '{htsus_no}' existiert bereits!")

        btn_f = tk.Frame(dlg, bg=MAIN_BG)
        btn_f.grid(row=len(fields)+1, column=0, columnspan=2, pady=10)
        tk.Button(btn_f, text="Speichern", command=_save,
                  bg="#166534", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=12, pady=4, cursor="hand2"
                  ).pack(side="left", padx=6)
        tk.Button(btn_f, text="Abbrechen", command=dlg.destroy,
                  bg="#6B7280", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=12, pady=4, cursor="hand2"
                  ).pack(side="left", padx=6)

    def _hcc_delete_row(self):
        """Ausgewählte Zeile löschen."""
        sel = self._hcc_tv.selection()
        if not sel:
            messagebox.showwarning("Keine Auswahl", "Bitte erst eine Zeile auswählen.",
                                   parent=self)
            return
        vals   = self._hcc_tv.item(sel[0], "values")
        row_id = int(vals[0])
        htsus  = vals[1]
        if not messagebox.askyesno("Löschen bestätigen",
                                   f"HTSUS-Nr. '{htsus}' wirklich löschen?",
                                   parent=self):
            return
        with sqlite3.connect(DB_PATH) as con:
            con.execute("DELETE FROM htsus_claim_config WHERE id=?", (row_id,))
        self._hcc_load()
        if hasattr(self, "_aus_load"):
            self._aus_load()

    def _hcc_import_excel(self):
        """Excel-Datei (Format wie db_Tariffs.xlsx) importieren."""
        import re
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            title="db_Tariffs.xlsx öffnen",
            filetypes=[("Excel-Dateien", "*.xlsx *.xls"), ("Alle Dateien", "*.*")],
            parent=self)
        if not path:
            return
        try:
            import openpyxl
            wb  = openpyxl.load_workbook(path, data_only=True)
            # Versuche zuerst 'Tabelle2', sonst erstes Sheet
            ws  = wb["Tabelle2"] if "Tabelle2" in wb.sheetnames else wb.active
            pat = re.compile(r"^\d{3,4}\.\d{2}\.\d{2,4}$")
            ins = upd = 0
            with sqlite3.connect(DB_PATH) as con:
                for row in ws.iter_rows(values_only=True):
                    htsus_no = row[1]
                    if not htsus_no or not isinstance(htsus_no, str):
                        continue
                    htsus_no = htsus_no.strip()
                    if not pat.match(htsus_no):
                        continue
                    htsus_rate = str(row[2]).strip() if row[2] else None
                    erst = str(row[3]).strip().lower() if row[3] else None
                    if erst not in ("ja", "nein"):
                        erst = None
                    c1 = str(row[4]).strip() if row[4] else None
                    c2 = str(row[5]).strip() if row[5] else None
                    c3 = str(row[6]).strip() if row[6] else None
                    cur = con.execute(
                        "SELECT id FROM htsus_claim_config WHERE htsus_no=?",
                        (htsus_no,))
                    existing = cur.fetchone()
                    if existing:
                        con.execute(
                            "UPDATE htsus_claim_config SET htsus_rate=?, erstattungsfaehig=?, "
                            "comment_1=?, comment_2=?, comment_3=? WHERE id=?",
                            (htsus_rate, erst, c1, c2, c3, existing[0]))
                        upd += 1
                    else:
                        con.execute(
                            "INSERT INTO htsus_claim_config "
                            "(htsus_no, htsus_rate, erstattungsfaehig, comment_1, comment_2, comment_3) "
                            "VALUES (?,?,?,?,?,?)",
                            (htsus_no, htsus_rate, erst, c1, c2, c3))
                        ins += 1
            self._hcc_load()
            self._hcc_status_var.set(
                f"Import OK – {ins} neu, {upd} aktualisiert")
            if hasattr(self, "_aus_load"):
                self._aus_load()
        except Exception as exc:
            messagebox.showerror("Import-Fehler", str(exc), parent=self)

    # ═══════════════════════════════════════════════════════════════════════
    # AUSWERTUNG – entry_lines JOIN entries (alle Felder)
    # ═══════════════════════════════════════════════════════════════════════
    def _build_auswertung_tab(self):
        parent = self.tab_auswertung
        outer  = tk.Frame(parent, bg=MAIN_BG)
        outer.pack(fill="both", expand=True, padx=14, pady=14)

        # ── Toolbar ─────────────────────────────────────────────────────────
        toolbar = tk.Frame(outer, bg=MAIN_BG)
        toolbar.pack(fill="x", pady=(0, 6))

        tk.Button(toolbar, text="🔄  Aktualisieren",
                  command=self._aus_load,
                  bg="#1D4ED8", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="📊  Excel Export",
                  command=self._aus_export_excel,
                  bg="#166534", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="📄  PDF Vendor-Daten einlesen",
                  command=self._aus_pdf_vendor_import,
                  bg="#7C3AED", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=(0, 12))
        self._aus_status_var = tk.StringVar(value="")
        tk.Label(toolbar, textvariable=self._aus_status_var,
                 bg=MAIN_BG, fg="#1F4E79", font=("Segoe UI", 9)).pack(side="left")

        # ── KPI-Leiste ──────────────────────────────────────────────────────
        kpi_frame = tk.Frame(outer, bg=CARD_BG, relief="flat",
                             highlightbackground="#DBEAFE", highlightthickness=1)
        kpi_frame.pack(fill="x", pady=(0, 8))

        self._aus_kpi_vars = {}
        _kpis = [
            ("entries",       "Entries",            "#1F4E79"),
            ("positionen",    "Positionen",          "#1F4E79"),
            ("sum_ev",        "Eingetr. Wert (USD)", "#1F4E79"),
            ("sum_duty",      "Zollbetrag (USD)",    "#B45309"),
            ("sum_mpf",       "MPF (USD)",           "#1F4E79"),
            ("sum_hmf",       "HMF (USD)",           "#1F4E79"),
            ("sum_total",     "Gesamt (USD)",        "#065F46"),
            ("sum_claim_v1",  "Claim Value 1 (USD)", "#7C3AED"),
        ]
        for ci, (key, lbl, fg) in enumerate(_kpis):
            var = tk.StringVar(value="–")
            self._aus_kpi_vars[key] = var
            cell = tk.Frame(kpi_frame, bg=CARD_BG)
            cell.grid(row=0, column=ci, padx=18, pady=8, sticky="w")
            tk.Label(cell, text=lbl, bg=CARD_BG, fg="#8A9EB5",
                     font=("Segoe UI", 8)).pack(anchor="w")
            tk.Label(cell, textvariable=var, bg=CARD_BG, fg=fg,
                     font=("Segoe UI", 12, "bold")).pack(anchor="w")
            kpi_frame.columnconfigure(ci, weight=1)

        # ── Suchleiste + Spaltenfilter ────────────────────────────────────
        sf_outer = tk.Frame(outer, bg=MAIN_BG)
        sf_outer.pack(fill="x", pady=(0, 4))
        sf = tk.Frame(sf_outer, bg=MAIN_BG)
        sf.pack(fill="x")

        self._aus_search_var = tk.StringVar()
        self._aus_search_var.trace_add("write", lambda *_: self._aus_filter())
        tk.Label(sf, text="Suche:", bg=MAIN_BG, fg="#5A6A80",
                 font=("Segoe UI", 9)).pack(side="left")
        ttk.Entry(sf, textvariable=self._aus_search_var, width=28).pack(
            side="left", padx=(4, 10))

        self._aus_toggle_lbl = tk.StringVar(value="▶ Spaltenfilter")
        tk.Button(sf, textvariable=self._aus_toggle_lbl,
                  command=self._aus_toggle_panel,
                  bg="#E8EDF4", fg="#1F4E79", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left")
        tk.Button(sf, text="✕", command=self._aus_reset_filter,
                  bg="#E8EDF4", fg="#6B7280", font=("Segoe UI", 9),
                  relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left", padx=(2, 0))

        self._aus_panel_open = False
        self._aus_panel, self._aus_get_col_filters, self._aus_reset_col = (
            _build_col_filter_panel(sf_outer, AUS_HEADERS, self._aus_filter, cols_per_row=6))

        # ── Treeview ─────────────────────────────────────────────────────
        tv_frame = tk.Frame(outer, bg=MAIN_BG)
        tv_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(tv_frame, orient="vertical")
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal")
        self._aus_tv = ttk.Treeview(
            tv_frame,
            columns=AUS_COLUMNS,
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._aus_tv.yview)
        hsb.config(command=self._aus_tv.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._aus_tv.pack(side="left", fill="both", expand=True)

        for col, hdr in zip(AUS_COLUMNS, AUS_HEADERS):
            self._aus_tv.heading(col, text=hdr,
                                 command=lambda c=col: self._aus_sort(c))
            w = _AUS_COL_WIDTHS.get(col, 90)
            self._aus_tv.column(col, width=w, minwidth=50, stretch=False, anchor="w")

        self._aus_tv.tag_configure("odd",  background="#F7F9FC")
        self._aus_tv.tag_configure("even", background="white")
        self._aus_tv.tag_configure("grand_total",
                                   background="#1F4E79", foreground="white",
                                   font=("Segoe UI", 9, "bold"))

        self._aus_all_rows  = []   # alle geladenen Zeilen
        self._aus_sort_col  = None
        self._aus_sort_rev  = False

        self._aus_load()

    # ── Daten laden ─────────────────────────────────────────────────────────
    def _aus_load(self):
        try:
            rows = self.conn.execute('''
                SELECT
                    e.filer_code_entry_no, e.entry_type, e.import_date, e.entry_date,
                    e.country_of_origin, e.exporting_country, e.importing_carrier,
                    e.mode_of_transport, e.us_port_of_unlading, e.source_file,
                    e.total_entered_value, e.duty_total, e.mpf_total, e.grand_total,
                    el.line_no, el.htsus_no, el.description, el.country_of_origin_line,
                    el.program_code, el.gross_weight, el.net_quantity, el.manifest_qty,
                    el.entered_value, el.htsus_rate, el.duty_amount,
                    el.mpf_rate, el.mpf_amount, el.hmf_rate, el.hmf_amount,
                    el.invoice_no, el.invoice_qty, el.invoice_value_amount,
                    el.invoice_value_rate, el.invoice_value_currency,
                    el.relationship, el.visa_no, el.source_block,
                    -- Lieferanten-Daten: db_Tariff (Primär) mit PDF-Extraktion als Fallback
                    COALESCE(dbt.dachser_dokument,  pve.dhl_doc_no)         AS dachser_dokument,
                    COALESCE(dbt.entry_document,    pve.entry_no)           AS entry_document,
                    dbt.vendor,
                    COALESCE(dbt.vendor_name,       pve.vendor_name)        AS vendor_name,
                    dbt.vendor_origin_o,
                    dbt.vendor_origin_e,
                    COALESCE(dbt.invoice_no_delivery_note, pve.vendor_invoice_no)
                                                                             AS invoice_no_delivery_note,
                    COALESCE(dbt.po_order_no,       pve.vendor_po_no)       AS po_order_no,
                    COALESCE(dbt.mat_no_sap,        pve.vendor_catalogue_no) AS mat_no_sap,
                    COALESCE(dbt.material_cat_sap,  pve.vendor_tariff_code) AS material_cat_sap,
                    COALESCE(dbt.quantity,           pve.vendor_quantity)    AS quantity,
                    dbt.rate,
                    COALESCE(dbt.invoice_doc_value,  pve.vendor_value)      AS invoice_doc_value,
                    COALESCE(dbt.doc_currency,       pve.vendor_currency)   AS doc_currency,
                    dbt.local_value_sap,
                    dbt.local_currency,
                    -- ZuOrd-Mat-KD: Kundenzuordnung (dbt direkt, Fallback zmk)
                    COALESCE(dbt.sap_customer_no,   zmk.kunden_nr)          AS kd_sap_customer_no,
                    COALESCE(dbt.sap_customer_name, zmk.sap_customer_name)  AS kd_sap_customer_name,
                    dbt.customer_part_no                                     AS kd_customer_part_no,
                    -- CBP-Reports: Claim-Felder
                    dbt.claim_status                                         AS dbt_claim_status,
                    -- Claim Value 1: aus CBP-Reports, nur wenn HTSUS in htsus_claim_config mit erstattungsfaehig='ja'
                    CASE WHEN EXISTS (
                             SELECT 1 FROM htsus_claim_config hcc
                             WHERE hcc.htsus_no = el.htsus_no
                               AND hcc.erstattungsfaehig = 'ja'
                         ) THEN dbt.claim_value_1
                         ELSE NULL
                    END                                                      AS dbt_claim_value_1,
                    dbt.claim_value_2                                        AS dbt_claim_value_2
                FROM entry_lines el
                JOIN entries e ON e.id = el.entry_id
                -- JOIN 1: db_Tariff Excel (Primärquelle)
                LEFT JOIN (
                    SELECT
                        LOWER(REPLACE(REPLACE(REPLACE(
                            COALESCE(entry_document, entry_no_new, ''),
                            '-',''), ' ',''), '/','')) AS norm_e,
                        MAX(dachser_dokument)          AS dachser_dokument,
                        MAX(entry_document)            AS entry_document,
                        MAX(vendor)                    AS vendor,
                        MAX(vendor_name)               AS vendor_name,
                        MAX(vendor_origin_o)           AS vendor_origin_o,
                        MAX(vendor_origin_e)           AS vendor_origin_e,
                        GROUP_CONCAT(DISTINCT NULLIF(invoice_no_delivery_note,''))
                                                       AS invoice_no_delivery_note,
                        GROUP_CONCAT(DISTINCT NULLIF(po_order_no,''))
                                                       AS po_order_no,
                        GROUP_CONCAT(DISTINCT NULLIF(mat_no_sap,''))
                                                       AS mat_no_sap,
                        GROUP_CONCAT(DISTINCT NULLIF(material_cat_sap,''))
                                                       AS material_cat_sap,
                        MAX(quantity)                  AS quantity,
                        MAX(rate)                      AS rate,
                        MAX(invoice_doc_value)         AS invoice_doc_value,
                        MAX(doc_currency)              AS doc_currency,
                        MAX(local_value_sap)           AS local_value_sap,
                        MAX(local_currency)            AS local_currency,
                        MAX(sap_customer_no)           AS sap_customer_no,
                        MAX(sap_customer_name)         AS sap_customer_name,
                        MAX(customer_part_no)          AS customer_part_no,
                        MAX(claim_status)              AS claim_status,
                        MAX(claim_value_1)             AS claim_value_1,
                        MAX(claim_value_2)             AS claim_value_2,
                        -- first mat_no_sap für zmk-Fallback-JOIN
                        MIN(NULLIF(mat_no_sap,''))     AS first_mat_no_sap
                    FROM stammdaten_db_tariff
                    WHERE COALESCE(entry_document, entry_no_new, '') != ''
                    GROUP BY norm_e
                ) dbt ON LOWER(REPLACE(REPLACE(REPLACE(
                    COALESCE(e.filer_code_entry_no,''), '-',''), ' ',''), '/',''))
                    = dbt.norm_e
                -- JOIN 2: PDF-Extraktion (Fallback für Entries ohne db_Tariff-Daten)
                LEFT JOIN (
                    SELECT
                        LOWER(REPLACE(REPLACE(REPLACE(
                            COALESCE(entry_no, ''), '-',''), ' ',''), '/','')) AS norm_e,
                        MAX(dhl_doc_no)        AS dhl_doc_no,
                        MAX(entry_no)          AS entry_no,
                        MAX(vendor_name)       AS vendor_name,
                        GROUP_CONCAT(DISTINCT NULLIF(vendor_invoice_no,''))
                                               AS vendor_invoice_no,
                        GROUP_CONCAT(DISTINCT NULLIF(vendor_po_no,''))
                                               AS vendor_po_no,
                        GROUP_CONCAT(DISTINCT NULLIF(vendor_catalogue_no,''))
                                               AS vendor_catalogue_no,
                        MAX(vendor_tariff_code) AS vendor_tariff_code,
                        MAX(vendor_quantity)    AS vendor_quantity,
                        MAX(vendor_value)       AS vendor_value,
                        MAX(vendor_currency)    AS vendor_currency
                    FROM pdf_vendor_extract
                    WHERE COALESCE(entry_no, '') != ''
                    GROUP BY norm_e
                ) pve ON LOWER(REPLACE(REPLACE(REPLACE(
                    COALESCE(e.filer_code_entry_no,''), '-',''), ' ',''), '/',''))
                    = pve.norm_e
                -- JOIN 3: ZuOrd-Mat-KD – Kundenzuordnung über Mat-Nr. (Fallback wenn dbt leer)
                LEFT JOIN stammdaten_zuord_mat_kd zmk
                    ON dbt.sap_customer_no IS NULL
                    AND zmk.mat_no_sap = COALESCE(
                        dbt.first_mat_no_sap,
                        SUBSTR(
                            COALESCE(pve.vendor_catalogue_no,''),
                            1,
                            CASE WHEN INSTR(COALESCE(pve.vendor_catalogue_no,''),',') > 0
                                 THEN INSTR(pve.vendor_catalogue_no,',') - 1
                                 ELSE LENGTH(COALESCE(pve.vendor_catalogue_no,''))
                            END
                        )
                    )
                ORDER BY e.import_date DESC, e.filer_code_entry_no,
                         CAST(el.line_no AS INTEGER)
            ''').fetchall()
            self._aus_all_rows = [tuple(str(v) if v is not None else "" for v in r)
                                   for r in rows]
            # KPIs
            kpi = self.conn.execute('''
                SELECT
                    COUNT(DISTINCT e.id),
                    COUNT(el.id),
                    SUM(CAST(REPLACE(COALESCE(el.entered_value,"0"),",","") AS REAL)),
                    SUM(CAST(REPLACE(COALESCE(el.duty_amount,  "0"),",","") AS REAL)),
                    SUM(CAST(REPLACE(COALESCE(el.mpf_amount,   "0"),",","") AS REAL)),
                    SUM(CAST(REPLACE(COALESCE(el.hmf_amount,   "0"),",","") AS REAL)),
                    -- Claim Value 1 KPI: nur HTSUS mit erstattungsfaehig='ja' in htsus_claim_config
                    SUM(CASE WHEN EXISTS (
                                     SELECT 1 FROM htsus_claim_config hcc
                                     WHERE hcc.htsus_no = el.htsus_no
                                       AND hcc.erstattungsfaehig = 'ja'
                                 )
                             THEN CAST(REPLACE(COALESCE(dbt.claim_value_1,'0'),',','') AS REAL)
                             ELSE 0 END)
                FROM entry_lines el
                JOIN entries e ON e.id = el.entry_id
                LEFT JOIN (
                    SELECT
                        LOWER(REPLACE(REPLACE(REPLACE(
                            COALESCE(entry_document, entry_no_new, ''),
                            '-',''), ' ',''), '/','')) AS norm_e,
                        MAX(claim_value_1) AS claim_value_1
                    FROM stammdaten_db_tariff
                    WHERE COALESCE(entry_document, entry_no_new, '') != ''
                    GROUP BY norm_e
                ) dbt ON LOWER(REPLACE(REPLACE(REPLACE(
                    COALESCE(e.filer_code_entry_no,''), '-',''), ' ',''), '/',''))
                    = dbt.norm_e
            ''').fetchone()
            def _fmt(v): return f"${v:,.2f}" if v else "–"
            self._aus_kpi_vars["entries"].set(f"{kpi[0]:,}")
            self._aus_kpi_vars["positionen"].set(f"{kpi[1]:,}")
            self._aus_kpi_vars["sum_ev"].set(_fmt(kpi[2]))
            self._aus_kpi_vars["sum_duty"].set(_fmt(kpi[3]))
            self._aus_kpi_vars["sum_mpf"].set(_fmt(kpi[4]))
            self._aus_kpi_vars["sum_hmf"].set(_fmt(kpi[5]))
            total = (kpi[3] or 0) + (kpi[4] or 0) + (kpi[5] or 0)
            self._aus_kpi_vars["sum_total"].set(_fmt(total))
            self._aus_kpi_vars["sum_claim_v1"].set(_fmt(kpi[6]))
            self._aus_status_var.set(
                f"{kpi[1]:,} Positionen  |  {kpi[0]:,} Entries  |  "
                f"Zoll: {_fmt(kpi[3])}  |  Claim V1: {_fmt(kpi[6])}  |  Gesamt: {_fmt(total)}")
        except Exception as exc:
            self._aus_status_var.set(f"Fehler: {exc}")
            return
        self._aus_filter()
    def _aus_filter(self):
        q           = self._aus_search_var.get().strip().lower()
        col_filters = self._aus_get_col_filters() if hasattr(self, "_aus_get_col_filters") else []
        rows        = self._aus_all_rows

        if q or col_filters:
            rows = [r for r in rows if _row_passes(r, col_filters, q)]

        # Sort
        if self._aus_sort_col is not None:
            sc = self._aus_sort_col
            def _key(r):
                v = r[sc]
                try: return (0, float(v.replace(",","")))
                except: return (1, v.lower())
            rows = sorted(rows, key=_key, reverse=self._aus_sort_rev)

        self._aus_tv.delete(*self._aus_tv.get_children())
        for i, row in enumerate(rows):
            tag = "odd" if i % 2 == 0 else "even"
            self._aus_tv.insert("", "end", values=row, tags=(tag,))

        # Summenzeile (sichtbare Zeilen)
        def _s(col_name):
            ci = AUS_COLUMNS.index(col_name)
            t = 0.0
            for r in rows:
                try: t += float(r[ci].replace(",",""))
                except: pass
            return t

        sum_vals = ["∑  SUMME (gefiltert)"] + [""] * (len(AUS_COLUMNS) - 1)
        for col in ("entered_value","duty_amount","mpf_amount","hmf_amount","invoice_value_amount",
                    "dbt_claim_value_1","dbt_claim_value_2"):
            if col in AUS_COLUMNS:
                ci = AUS_COLUMNS.index(col)
                sum_vals[ci] = f"{_s(col):,.2f}"
        self._aus_tv.insert("", "end", values=tuple(sum_vals),
                            tags=("grand_total",))

    def _aus_sort(self, col):
        ci = AUS_COLUMNS.index(col)
        if self._aus_sort_col == ci:
            self._aus_sort_rev = not self._aus_sort_rev
        else:
            self._aus_sort_col = ci
            self._aus_sort_rev = False
        self._aus_filter()

    def _aus_toggle_panel(self):
        if self._aus_panel_open:
            self._aus_panel.pack_forget()
            self._aus_toggle_lbl.set("▶ Spaltenfilter")
        else:
            self._aus_panel.pack(fill="x", pady=(2, 0))
            self._aus_toggle_lbl.set("▼ Spaltenfilter")
        self._aus_panel_open = not self._aus_panel_open

    def _aus_reset_filter(self):
        self._aus_search_var.set("")
        self._aus_reset_col()

    # ── Excel-Export ────────────────────────────────────────────────────────
    def _aus_export_excel(self):
        from tkinter import filedialog as _fd
        import threading as _thr
        path = _fd.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Auswertung_entry_lines.xlsx",
            title="Auswertung speichern als …",
        )
        if not path:
            return
        self._aus_status_var.set("Excel wird erstellt …")
        _thr.Thread(target=self._aus_export_worker, args=(path,), daemon=True).start()

    def _aus_export_worker(self, path):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime as _dt

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Auswertung"

            thin = Side(style="thin", color="B0C4DE")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _hcell(r, c, val, bg="1F4E79", fg="FFFFFF", bold=True):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font      = Font(bold=bold, color=fg, name="Arial", size=9)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = brd

            # Titel
            ts = _dt.now().strftime("%d.%m.%Y %H:%M")
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1,  end_column=len(AUS_COLUMNS))
            tc = ws.cell(row=1, column=1,
                         value=f"Auswertung – entry_lines  |  Erstellt: {ts}")
            tc.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            tc.fill      = PatternFill("solid", fgColor="0F2D54")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # Header
            for ci, hdr in enumerate(AUS_HEADERS, 1):
                _hcell(2, ci, hdr, bg="BDD7EE", fg="1F4E79")
            ws.row_dimensions[2].height = 28

            # Column widths
            for ci, col in enumerate(AUS_COLUMNS, 1):
                ws.column_dimensions[get_column_letter(ci)].width = (
                    _AUS_COL_WIDTHS.get(col, 9))

            # Data
            rows = self._aus_all_rows
            num_idx = {AUS_COLUMNS.index(c) for c in _AUS_NUM_COLS
                       if c in AUS_COLUMNS}
            for ri, row in enumerate(rows):
                er  = ri + 3
                bg  = "EFF6FF" if ri % 2 == 0 else "FFFFFF"
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=er, column=ci)
                    if (ci - 1) in num_idx:
                        try:
                            c.value = float(val.replace(",","")) if val else None
                            c.number_format = "#,##0.00"
                        except:
                            c.value = val
                    else:
                        c.value = val
                    c.font      = Font(name="Arial", size=9)
                    c.fill      = PatternFill("solid", fgColor=bg)
                    c.alignment = Alignment(vertical="center")
                    c.border    = brd

            # Summenzeile
            sr = len(rows) + 3
            ws.cell(row=sr, column=1, value="∑  GESAMT")
            ws.cell(row=sr, column=1).font = Font(bold=True, color="FFFFFF",
                                                   name="Arial", size=9)
            ws.cell(row=sr, column=1).fill = PatternFill("solid", fgColor="1E3A5F")
            ws.cell(row=sr, column=1).alignment = Alignment(horizontal="center",
                                                             vertical="center")
            for col in _AUS_NUM_COLS:
                if col in AUS_COLUMNS:
                    ci = AUS_COLUMNS.index(col) + 1
                    cl = get_column_letter(ci)
                    c = ws.cell(row=sr, column=ci,
                                value=f"=SUM({cl}3:{cl}{sr-1})")
                    c.font         = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                    c.fill         = PatternFill("solid", fgColor="1E3A5F")
                    c.number_format = "#,##0.00"
                    c.alignment    = Alignment(horizontal="right", vertical="center")
                    c.border       = brd

            ws.freeze_panes = "A3"
            ws.auto_filter.ref = f"A2:{get_column_letter(len(AUS_COLUMNS))}2"

            wb.save(path)
            self.after(0, lambda: self._aus_status_var.set(
                f"✓  Excel gespeichert: {path}"))
        except Exception as exc:
            self.after(0, lambda: self._aus_status_var.set(f"Fehler: {exc}"))

    def _aus_pdf_vendor_import(self):
        """Startet PDF Vendor-Extraktion.
        M-*.pdf: Textextraktion (kein API, kostenlos, schnell).
        Alle anderen: Claude Vision (API-Key nötig).
        """
        pdf_dir = os.path.join(APP_DIR, "pdf_archive")
        if not os.path.isdir(pdf_dir):
            messagebox.showerror("PDF-Archiv nicht gefunden",
                                 f"Ordner nicht gefunden:\n{pdf_dir}")
            return

        all_pdfs  = [f for f in os.listdir(pdf_dir) if f.lower().endswith(".pdf")]
        m_pdfs    = [f for f in all_pdfs if f.startswith("M")]
        img_pdfs  = [f for f in all_pdfs if not f.startswith("M")]
        already   = self.conn.execute(
            "SELECT COUNT(*) FROM pdf_vendor_extract").fetchone()[0]

        # API-Key nur wenn Bild-PDFs vorhanden und noch nicht alle verarbeitet
        api_key = ""
        img_pending = [f for f in img_pdfs
                       if not self.conn.execute(
                           "SELECT 1 FROM pdf_vendor_extract WHERE pdf_filename=?",
                           (f,)).fetchone()]
        if img_pending:
            import configparser as _cp
            cfg = _cp.ConfigParser()
            cfg.read(SAP_CONFIG_PATH, encoding="utf-8")
            api_key = cfg.get("CLAUDE", "ANTHROPIC_API_KEY", fallback="").strip()
            if not api_key:
                api_key = simpledialog.askstring(
                    "Anthropic API-Key",
                    "Für Bild-PDFs (DACHSER) wird Claude Vision benötigt.\n"
                    "API-Key eingeben (oder leer lassen für nur M-PDFs):",
                    show="*") or ""
                api_key = api_key.strip()
                if api_key:
                    if not cfg.has_section("CLAUDE"):
                        cfg.add_section("CLAUDE")
                    cfg.set("CLAUDE", "ANTHROPIC_API_KEY", api_key)
                    with open(SAP_CONFIG_PATH, "w", encoding="utf-8") as fh:
                        cfg.write(fh)

        info = (
            f"{len(all_pdfs)} PDF-Dateien im Archiv  "
            f"(M-PDFs: {len(m_pdfs)}, Sonstige: {len(img_pdfs)})\n"
            f"Bereits verarbeitet: {already}\n\n"
            f"M-*.pdf → Textextraktion (kostenlos, schnell)\n"
        )
        if img_pending and api_key:
            info += (f"{len(img_pending)} Bild-PDFs → Claude Vision "
                     f"(~${len(img_pending)*0.002:.2f} Kosten)\n")
        elif img_pending:
            info += f"{len(img_pending)} Bild-PDFs werden übersprungen (kein API-Key)\n"

        if not messagebox.askyesno("PDF Vendor-Daten einlesen",
                                   info + "\nFortschritt im Status-Label. Weiter?"):
            return

        self._aus_status_var.set(f"PDF-Extraktion läuft … ({len(all_pdfs)} Dateien)")
        self._aus_pdf_stop = threading.Event()

        def _worker():
            try:
                from pdf_vendor_extractor import process_m_pdfs, process_all

                def _log(txt):
                    short = txt.strip()[-100:] if txt.strip() else ""
                    self.after(0, lambda: self._aus_status_var.set(short))

                # Schritt 1: M-PDFs (Textextraktion, kostenlos)
                r1 = process_m_pdfs(
                    db_path=str(DB_PATH),
                    pdf_dir=pdf_dir,
                    force=False,
                    log_fn=_log,
                )

                # Schritt 2: Bild-PDFs via Claude Vision (wenn API-Key vorhanden)
                r2 = {"processed": 0, "skipped": 0, "errors": 0}
                if api_key and img_pending:
                    r2 = process_all(
                        db_path=str(DB_PATH),
                        pdf_dir=pdf_dir,
                        api_key=api_key,
                        force=False,
                        log_fn=_log,
                        stop_flag=self._aus_pdf_stop,
                    )

                total_proc = r1["processed"] + r2["processed"]
                total_skip = r1["skipped"]  + r2["skipped"]
                total_err  = r1["errors"]   + r2["errors"]
                msg = (f"✓ PDF-Extraktion fertig: {total_proc} verarbeitet, "
                       f"{total_skip} übersprungen, {total_err} Fehler")
                self.after(0, lambda: self._aus_status_var.set(msg))
                self.after(0, self._aus_load)
            except ImportError:
                self.after(0, lambda: messagebox.showerror(
                    "Fehlende Pakete",
                    "Bitte installieren:\n  pip install pdfplumber\n\n"
                    "Für Bild-PDFs zusätzlich:\n  pip install pymupdf anthropic\n\n"
                    "Dann App neu starten."))
            except Exception as exc:
                self.after(0, lambda: messagebox.showerror(
                    "PDF-Extraktion Fehler", str(exc)[:500]))
    def _dbt_ensure_db(self):
        conn = sqlite3.connect(DB_PATH)
        conn.executescript(DB_TARIFF_SCHEMA)
        conn.commit()
        conn.close()

    def _dbt_on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw
        if not os.path.isfile(path):
            self._dbt_status_var.set(f"Datei nicht gefunden: {path}")
            return
        if not path.lower().endswith((".xlsx", ".xls")):
            self._dbt_status_var.set("Nur Excel-Dateien (.xlsx / .xls) werden unterstützt.")
            return
        self._dbt_status_var.set(f"Importiere: {os.path.basename(path)} …")
        threading.Thread(target=self._dbt_import_excel, args=(path,),
                         daemon=True).start()

    def _dbt_import_excel(self, path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            # Prefer sheet named "Data", else first sheet
            sheet_name = "Data" if "Data" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Header row 1 (index 0), data from row 2 (index 1)
            data_rows = rows[1:]
            n_cols = len(DB_TARIFF_COLUMNS)
            conn = sqlite3.connect(DB_PATH)
            cur  = conn.cursor()
            cur.execute("DELETE FROM stammdaten_db_tariff")
            inserted = 0
            source = os.path.basename(path)
            placeholders = ", ".join(["?"] * n_cols)
            col_list = ", ".join(DB_TARIFF_COLUMNS)
            for row in data_rows:
                vals = list(row[:n_cols])
                while len(vals) < n_cols:
                    vals.append(None)
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                norm = []
                for v in vals:
                    if v is None:
                        norm.append(None)
                    elif isinstance(v, float) and v == int(v):
                        norm.append(str(int(v)))
                    else:
                        s = str(v).strip().replace("\n", " ")
                        norm.append(s if s else None)
                cur.execute(
                    f"INSERT INTO stammdaten_db_tariff ({col_list}) VALUES ({placeholders})",
                    norm,
                )
                inserted += 1
            conn.commit()
            conn.close()
            self.root.after(0, lambda n=inserted, s=source: (
                self._dbt_status_var.set(
                    f"✓ {n} Zeilen importiert aus '{s}' (Tabelle ersetzt)."),
                self._load_dbt_data(),
            ))
        except ImportError:
            self.root.after(0, lambda: self._dbt_status_var.set(
                "openpyxl nicht installiert – bitte: pip install openpyxl"))
        except Exception as exc:
            self.root.after(0, lambda e=exc: self._dbt_status_var.set(
                f"Import-Fehler: {e}"))

    def _load_dbt_data(self, *_):
        try:
            conn = sqlite3.connect(DB_PATH)
            cur  = conn.cursor()
            cur.execute("SELECT " + ", ".join(DB_TARIFF_COLUMNS) +
                        " FROM stammdaten_db_tariff ORDER BY id DESC")
            rows = cur.fetchall()
            conn.close()
        except Exception:
            rows = []
        self._dbt_all_rows = rows
        self._dbt_filter()

    def _dbt_toggle_panel(self):
        if self._dbt_panel_open:
            self._dbt_panel.pack_forget()
            self._dbt_panel_open = False
            self._dbt_toggle_lbl.set("▶ Spaltenfilter")
        else:
            self._dbt_panel.pack(fill="x", pady=(2, 0))
            self._dbt_panel_open = True
            self._dbt_toggle_lbl.set("▼ Spaltenfilter")

    def _dbt_reset_filter(self):
        self._dbt_search_var.set("")
        self._dbt_reset_col()
        self._dbt_filter()
        if self._dbt_panel_open:
            self._dbt_toggle_panel()

    def _dbt_filter(self, *_):
        term = self._dbt_search_var.get().lower().strip()
        col_filters = self._dbt_get_col_filters()
        tv = self._dbt_tv
        for item in tv.get_children():
            tv.delete(item)
        count = 0
        for row in getattr(self, "_dbt_all_rows", []):
            if not _row_passes(row, col_filters, term):
                continue
            tag = "even" if count % 2 == 0 else "odd"
            tv.insert("", "end", values=row, tags=(tag,))
            count += 1
        total = len(getattr(self, "_dbt_all_rows", []))
        active = len(col_filters)
        badge = f"  [{active} Spaltenfilter]" if active else ""
        if term or active:
            self._dbt_status_var.set(f"{count} von {total} Einträgen{badge} angezeigt.")
        else:
            self._dbt_status_var.set(f"{total} Einträge in der Datenbank.")

    def _dbt_sort(self, col):
        if self._dbt_sort_col == col:
            self._dbt_sort_rev = not self._dbt_sort_rev
        else:
            self._dbt_sort_col = col
            self._dbt_sort_rev = False
        idx = DB_TARIFF_COLUMNS.index(col)
        self._dbt_all_rows.sort(
            key=lambda r: (r[idx] is None, str(r[idx] or "").lower()),
            reverse=self._dbt_sort_rev,
        )
        self._dbt_filter()

    def _dbt_clear_table(self):
        import tkinter.messagebox as mb
        if not mb.askyesno(
            "Tabelle leeren",
            "Alle db_Tariff-Daten aus der Datenbank löschen?\nDieser Vorgang kann nicht rückgängig gemacht werden.",
        ):
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM stammdaten_db_tariff")
            conn.commit()
            conn.close()
        except Exception as exc:
            self._dbt_status_var.set(f"Fehler: {exc}")
            return
        self._load_dbt_data()

    def _dbt_run_abgleich(self):
        """Vergleicht db_Tariff mit Claim-Report.
        Berechnet Werte-Differenzen und erklärt Ursachen im Detail-Panel."""
        import re as _re
        from datetime import datetime as _dt

        def _norm(s):
            return _re.sub(r'\D', '', str(s or ""))

        def _v(s):
            """String-Zahl sicher zu float."""
            if s is None:
                return None
            try:
                return float(str(s).replace(",", "").strip())
            except (ValueError, TypeError):
                return None

        TOLERANCE = 0.05   # USD Rundungstoleranz

        # ── 1. Claim-Report: duty_amount-Summe pro Entry aus DB ───────────
        try:
            conn = sqlite3.connect(DB_PATH)
            claim_raw = conn.execute(
                """SELECT e.filer_code_entry_no,
                          SUM(CAST(REPLACE(COALESCE(l.duty_amount,'0'),',','') AS REAL))
                   FROM entry_lines l
                   JOIN entries e ON e.id = l.entry_id
                   WHERE l.duty_amount IS NOT NULL
                   GROUP BY e.filer_code_entry_no"""
            ).fetchall()
            conn.close()
        except Exception as exc:
            self._dbt_status_var.set(f"Abgleich-Fehler (DB): {exc}")
            return

        if not claim_raw:
            self._dbt_status_var.set("⚠ Keine Claim-Daten in entry_lines gefunden.")
            return

        # Claim: norm_key → (original_entry_no, duty_sum)
        claim_map = {_norm(r[0]): (r[0], r[1] or 0.0) for r in claim_raw if r[0]}

        # ── 2. db_Tariff: summe_tax + HTS-Felder pro Entry ────────────────
        IDX_ED  = DB_TARIFF_COLUMNS.index("entry_document")
        IDX_EN  = DB_TARIFF_COLUMNS.index("entry_no_new")
        IDX_DDD = DB_TARIFF_COLUMNS.index("dachser_duty_doc_no")
        IDX_ST  = DB_TARIFF_COLUMNS.index("summe_tax")
        IDX_TVT = DB_TARIFF_COLUMNS.index("tax_vol_total_doc")
        # HTS-Betrags-Spalten (indices 10-40)
        HTS_INDICES = list(range(
            DB_TARIFF_COLUMNS.index("hts_3926_90_9989"),
            DB_TARIFF_COLUMNS.index("ascertained_other") + 1
        ))

        # Baue pro norm_key: {summe_tax, hts_vals, dachser_doc, orig_entry_doc}
        db_entry_data = {}  # norm_key → dict
        for row in getattr(self, "_dbt_all_rows", []):
            _ed  = str(row[IDX_ED]  or "").strip()
            _en  = str(row[IDX_EN]  or "").strip()
            _key = _norm(_ed) if _ed else _norm(_en)
            if not _key:
                continue
            if _key not in db_entry_data:
                db_entry_data[_key] = {
                    "orig_doc":   _ed or _en,
                    "dachser":    str(row[IDX_DDD] or "").strip(),
                    "summe_tax":  None,
                    "tax_vol":    None,
                    "hts_vals":   {},    # header → value
                }
            d = db_entry_data[_key]
            _st = _v(row[IDX_ST])
            if _st is not None and _st != 0.0:
                if d["summe_tax"] is None:
                    d["summe_tax"] = _st
                else:
                    d["summe_tax"] += _st
            _tv = _v(row[IDX_TVT])
            if _tv is not None and _tv != 0.0 and d["tax_vol"] is None:
                d["tax_vol"] = _tv
            for hi in HTS_INDICES:
                hv = _v(row[hi] if hi < len(row) else None)
                if hv is not None and hv != 0.0:
                    hdr = DB_TARIFF_HEADERS[hi]
                    d["hts_vals"][hdr] = d["hts_vals"].get(hdr, 0.0) + hv

        # ── 3. Mapping + Gründe aufbauen ──────────────────────────────────
        amap      = {}
        ok_cnt    = 0
        fehlt_cnt = 0
        diff_cnt  = 0
        nur_db    = 0

        # Alle Entry-Keys aus beiden Quellen
        all_keys = set(db_entry_data.keys()) | set(claim_map.keys())

        # Bericht-Zeilen (list of (tag, text))
        report_lines = []

        def _rline(tag, text):
            report_lines.append((tag, text))

        nicht_in_db_list = []
        nur_db_list      = []
        ok_list          = []
        diff_list        = []
        fehlt_list       = []

        for key in sorted(all_keys):
            in_claim = key in claim_map
            in_db    = key in db_entry_data

            if in_db:
                d = db_entry_data[key]
                orig = d["orig_doc"] or key
                dach = d["dachser"]
            else:
                orig = claim_map[key][0]
                dach = ""

            claim_val = claim_map[key][1] if in_claim else None
            db_val    = db_entry_data[key]["summe_tax"] if in_db else None

            if not in_db:
                # Claim-Eintrag existiert nicht in db_Tariff
                amap[key] = "fehlt"
                nicht_in_db_list.append((orig, claim_val))
                continue

            if not in_claim:
                # db_Tariff Eintrag ohne Claim-Report-Match
                amap[key] = "nur_db"
                nur_db_list.append((orig, dach, db_val))
                nur_db += 1
                continue

            # Beide vorhanden – Werte vergleichen
            if db_val is None:
                # summe_tax leer
                amap[key] = "fehlt"
                fehlt_cnt += 1
                d2 = db_entry_data[key]
                fehlt_list.append({
                    "orig": orig, "dach": dach, "key": key,
                    "claim": claim_val,
                    "tax_vol": d2["tax_vol"],
                    "hts_vals": d2["hts_vals"],
                })
            else:
                diff = claim_val - db_val
                pct  = abs(diff) / max(abs(claim_val), 0.01) * 100
                if abs(diff) <= TOLERANCE:
                    amap[key] = "ok"
                    ok_cnt += 1
                    ok_list.append((orig, db_val))
                else:
                    amap[key] = "fehlt"
                    diff_cnt += 1
                    # Ursache bestimmen
                    if pct < 1.0:
                        reason = "Rundungsdifferenz (< 1 %)"
                    elif pct < 10.0:
                        reason = "Teillieferung, Wechselkurs oder abweichender Abrechnungszeitraum möglich"
                    else:
                        reason = "Wesentliche Abweichung – Manuelle Prüfung erforderlich"
                    diff_list.append({
                        "orig": orig, "dach": dach,
                        "claim": claim_val, "db": db_val,
                        "diff": diff, "pct": pct,
                        "reason": reason,
                    })

        # Claim-Einträge ohne db_Tariff-Zeile
        for orig, cv in nicht_in_db_list:
            pass  # already mapped above

        # ── 4. Detail-Text aufbauen ───────────────────────────────────────
        def _dn(v):
            return _de_num(v) if v is not None else "–"

        ts = _dt.now().strftime("%d.%m.%Y  %H:%M:%S")
        lines = []
        lines.append(("h1",
            f"═══  ABGLEICH-BERICHT  db_Tariff  ↔  Claim-Report  ═══  {ts}\n"))
        lines.append(("grey",
            f"Claim-Report: {len(claim_map)} Entries  │  "
            f"db_Tariff: {len(db_entry_data)} unique Entry-Nos\n"))
        lines.append(("grey", "─" * 80 + "\n\n"))

        # ── Differenzen ──
        if diff_list:
            lines.append(("bold", f"⚠  BETRAGSABWEICHUNGEN  ({len(diff_list)})\n"))
            lines.append(("grey", "─" * 80 + "\n"))
            for it in diff_list:
                sign = "+" if it["diff"] > 0 else ""
                lines.append(("orange",
                    f"⚠  Entry  {it['orig']}  [{it['dach']}]\n"))
                lines.append(("",
                    f"   Tariff Claim (Claim-Report) : {_dn(it['claim']):>14} USD\n"
                    f"   Summe TAX  (db_Tariff)      : {_dn(it['db']):>14} USD\n"
                    f"   Differenz                   : {sign}{_dn(it['diff']):>14} USD"
                    f"  ({sign}{it['pct']:.1f} %)\n"))
                lines.append(("orange",
                    f"   → Ursache: {it['reason']}\n\n"))

        # ── Summe TAX fehlt ──
        if fehlt_list:
            lines.append(("bold", f"🔴  SUMME TAX FEHLT IN db_Tariff  ({len(fehlt_list)})\n"))
            lines.append(("grey", "─" * 80 + "\n"))
            for it in fehlt_list:
                lines.append(("red",
                    f"🔴  Entry  {it['orig']}  [{it['dach']}]\n"))
                lines.append(("",
                    f"   Tariff Claim (Claim-Report) : {_dn(it['claim']):>14} USD\n"
                    f"   Summe TAX  (db_Tariff)      : {'leer':>14}\n"))
                if it["hts_vals"]:
                    top_hts = sorted(it["hts_vals"].items(),
                                     key=lambda x: -x[1])[:5]
                    hts_str = "  ".join(f"{h}={_dn(v)}" for h, v in top_hts)
                    lines.append(("",
                        f"   HTS-Spalten befüllt        : {hts_str}\n"))
                    lines.append(("red",
                        "   → Ursache: HTS-Beträge vorhanden, aber 'Summe TAX' nicht berechnet/exportiert.\n\n"))
                elif it["tax_vol"]:
                    lines.append(("",
                        f"   Tax Vol. Total (col. 4)    : {_dn(it['tax_vol']):>14} USD\n"))
                    lines.append(("red",
                        "   → Ursache: Tax Volume vorhanden, aber 'Summe TAX' Spalte leer.\n\n"))
                else:
                    lines.append(("red",
                        "   → Ursache: Alle Steuerfelder leer – Entry evtl. unvollständig importiert.\n\n"))

        # ── Nicht in db_Tariff ──
        if nicht_in_db_list:
            lines.append(("bold", f"❌  CLAIM-EINTRÄGE NICHT IN db_Tariff  ({len(nicht_in_db_list)})\n"))
            lines.append(("grey", "─" * 80 + "\n"))
            for orig, cv in nicht_in_db_list:
                lines.append(("red",
                    f"   · {orig}  (Tariff Claim: {_dn(cv)} USD)\n"))
            lines.append(("red",
                "\n   → Ursache: Entry-No im Claim-Report vorhanden, aber Excel-Import fehlt.\n\n"))

        # ── Nur in db_Tariff ──
        if nur_db_list:
            lines.append(("bold", f"🟡  NUR IN db_Tariff, KEIN Claim-Report-Match  ({len(nur_db_list)})\n"))
            lines.append(("grey", "─" * 80 + "\n"))
            for orig, dach, sv in nur_db_list:
                lines.append(("grey",
                    f"   · {orig}  [{dach}]  Summe TAX: {_dn(sv)}\n"))
            lines.append(("grey",
                "\n   → Hinweis: Diese Entries sind in db_Tariff, haben aber keinen Claim-Report-Eintrag.\n\n"))

        # ── OK ──
        if ok_list:
            lines.append(("bold", f"✅  ÜBEREINSTIMMUNGEN  ({len(ok_list)})\n"))
            lines.append(("grey", "─" * 80 + "\n"))
            for orig, sv in ok_list:
                lines.append(("green",
                    f"   ✅ {orig}  –  {_dn(sv)} USD\n"))

        # Detail-Text einschreiben
        dt = self._dbt_detail_text
        dt.configure(state="normal")
        dt.delete("1.0", "end")
        for tag, txt in lines:
            dt.insert("end", txt, tag if tag else ())
        dt.configure(state="disabled")
        dt.see("1.0")

        # Detail-Panel anzeigen (treeview kurz neu verankern)
        if not self._dbt_detail_outer.winfo_ismapped():
            self._dbt_tv_frame.pack_forget()
            self._dbt_detail_outer.pack(side="bottom", fill="x",
                                        pady=(6, 0), in_=self._dbt_outer)
            self._dbt_tv_frame.pack(fill="both", expand=True, in_=self._dbt_outer)

        # ── 5. Abgleich-Map + Status ──────────────────────────────────────
        self._dbt_abgleich_map    = amap
        self._dbt_abgleich_active = True
        self._dbt_filter()

        nicht_in_db_cnt = len(nicht_in_db_list)
        self._dbt_status_var.set(
            f"Abgleich:  ✅ {ok_cnt} OK  "
            f"⚠ {diff_cnt} Differenz  "
            f"🔴 {fehlt_cnt} Summe TAX fehlt  "
            f"🟡 {nur_db} nur db_Tariff  "
            f"❌ {nicht_in_db_cnt} nicht in db_Tariff"
        )

    def _dbt_reset_abgleich(self):
        """Setzt Abgleich-Kennzeichnung und Detail-Panel zurück."""
        self._dbt_abgleich_active = False
        self._dbt_abgleich_map    = {}
        # Detail-Panel verstecken
        if self._dbt_detail_outer.winfo_ismapped():
            self._dbt_detail_outer.pack_forget()
        dt = self._dbt_detail_text
        dt.configure(state="normal")
        dt.delete("1.0", "end")
        dt.configure(state="disabled")
        self._dbt_filter()
        total = len(getattr(self, "_dbt_all_rows", []))
        self._dbt_status_var.set(f"{total} Einträge in der Datenbank.")
    def _dbt_export_excel(self):
        """Exportiert die aktuelle db_Tariff-Ansicht (gefiltert + Summenzeile)
        als formatierte Excel-Datei und öffnet sie danach automatisch."""
        if openpyxl is None:
            messagebox.showerror("openpyxl fehlt",
                                 "Bitte installieren:\n    pip install openpyxl")
            return

        rows = getattr(self, "_dbt_all_rows", [])
        if not rows:
            messagebox.showinfo("Keine Daten",
                                "Die db_Tariff-Tabelle ist leer. "
                                "Bitte zuerst eine Excel-Datei importieren.")
            return

        from tkinter import filedialog
        out_path = filedialog.asksaveasfilename(
            title="db_Tariff Report speichern",
            defaultextension=".xlsx",
            filetypes=[("Excel-Dateien", "*.xlsx")],
            initialfile="db_Tariff_Report.xlsx",
            initialdir=APP_DIR,
        )
        if not out_path:
            return

        self._dbt_status_var.set("📊 Erstelle Excel-Report …")
        self.root.update_idletasks()

        try:
            from datetime import datetime as _dtnow
            import openpyxl as _xl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter

            # ── Farben / Styles ──────────────────────────────────────────
            HDR_BG    = "1F4E79"   # Dunkelblau – Kopfzeile
            HDR_FG    = "FFFFFF"
            SUM_BG    = "1E3A5F"   # Noch dunkler – Summenzeile
            SUM_FG    = "FFFFFF"
            ALT1_BG   = "EFF6FF"   # Hellblau – gerade Zeilen
            ALT2_BG   = "FFFFFF"   # Weiß – ungerade Zeilen
            WARN_BG   = "FEE2E2"   # Rot – fehlt (Abgleich)
            OK_BG     = "D1FAE5"   # Grün – ok (Abgleich)
            NUR_DB_BG = "FEF9C3"   # Gelb – nur db_Tariff

            def _hdr_style():
                return {
                    "font": Font(bold=True, color=HDR_FG, name="Segoe UI", size=9),
                    "fill": PatternFill("solid", fgColor=HDR_BG),
                    "alignment": Alignment(horizontal="center",
                                           vertical="center", wrap_text=True),
                }

            def _sum_style():
                return {
                    "font": Font(bold=True, color=SUM_FG, name="Segoe UI", size=9),
                    "fill": PatternFill("solid", fgColor=SUM_BG),
                    "alignment": Alignment(horizontal="right", vertical="center"),
                }

            def _row_fill(idx, ab_status=""):
                if ab_status == "fehlt":
                    return PatternFill("solid", fgColor=WARN_BG)
                if ab_status == "ok":
                    return PatternFill("solid", fgColor=OK_BG)
                if ab_status == "nur_db":
                    return PatternFill("solid", fgColor=NUR_DB_BG)
                return PatternFill("solid", fgColor=ALT1_BG if idx % 2 == 0 else ALT2_BG)

            thin = Side(style="thin", color="B0C4DE")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # ── Workbook anlegen ─────────────────────────────────────────
            wb = _xl.Workbook()
            ws = wb.active
            ws.title = "db_Tariff"

            # Titel-Zeile
            ts = _dtnow.now().strftime("%d.%m.%Y  %H:%M")
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=len(DB_TARIFF_HEADERS))
            title_cell = ws.cell(row=1, column=1,
                                 value=f"db_Tariff  –  Report  |  {ts}")
            title_cell.font     = Font(bold=True, color="FFFFFF",
                                       name="Segoe UI", size=11)
            title_cell.fill     = PatternFill("solid", fgColor="0F2D54")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # Header-Zeile (Zeile 2)
            for ci, hdr in enumerate(DB_TARIFF_HEADERS, start=1):
                cell = ws.cell(row=2, column=ci, value=hdr)
                for k, v in _hdr_style().items():
                    setattr(cell, k, v)
                cell.border = border
            ws.row_dimensions[2].height = 28

            # Abgleich-Map (falls aktiv)
            import re as _re
            def _norm_key(row):
                _ed = str(row[DB_TARIFF_COLUMNS.index("entry_document")] or "").strip()
                _en = str(row[DB_TARIFF_COLUMNS.index("entry_no_new")]   or "").strip()
                k = _re.sub(r'\D', '', _ed) if _ed else _re.sub(r'\D', '', _en)
                return k

            amap   = getattr(self, "_dbt_abgleich_map", {})
            ab_act = getattr(self, "_dbt_abgleich_active", False)

            # ── Datenzeilen ──────────────────────────────────────────────
            for ri, row in enumerate(rows):
                excel_row = ri + 3     # Zeile 1=Titel, 2=Header
                ab_status = ""
                if ab_act:
                    ab_status = amap.get(_norm_key(row), "")
                fill = _row_fill(ri, ab_status)

                for ci, val in enumerate(row):
                    cell = ws.cell(row=excel_row, column=ci + 1, value=val)
                    cell.fill      = fill
                    cell.font      = Font(name="Segoe UI", size=9)
                    cell.border    = border
                    cell.alignment = Alignment(vertical="center",
                                               horizontal="right"
                                               if ci < len(DB_TARIFF_COLUMNS)
                                               and DB_TARIFF_COLUMNS[ci] in _DBT_SUM_COLS
                                               else "left")

            # ── Summenzeile ──────────────────────────────────────────────
            sum_row_idx = len(rows) + 3
            n_cols = len(DB_TARIFF_COLUMNS)
            sums   = [None] * n_cols
            for ci, col in enumerate(DB_TARIFF_COLUMNS):
                if col not in _DBT_SUM_COLS:
                    continue
                total = 0.0; has_val = False
                for row in rows:
                    v = _parse_de(row[ci] if ci < len(row) else None)
                    if v is not None:
                        total += v; has_val = True
                if has_val:
                    sums[ci] = total

            sum_style = _sum_style()
            for ci in range(n_cols):
                val = "∑ GESAMT" if ci == 0 else sums[ci]
                cell = ws.cell(row=sum_row_idx, column=ci + 1, value=val)
                for k, v in sum_style.items():
                    setattr(cell, k, v)
                cell.border = border
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
            ws.row_dimensions[sum_row_idx].height = 18

            # ── Spaltenbreiten ────────────────────────────────────────────
            for ci, hdr in enumerate(DB_TARIFF_HEADERS, start=1):
                col_letter = get_column_letter(ci)
                col_name   = DB_TARIFF_COLUMNS[ci - 1]
                # Breite nach Inhalt: Betragsspalten schmäler, Text breiter
                if col_name in _DBT_SUM_COLS:
                    width = max(11, min(18, len(hdr) + 3))
                elif col_name in ("vendor_name", "sap_customer_name",
                                  "invoice_no_delivery_note", "po_order_no"):
                    width = 24
                elif col_name in ("entry_document", "entry_no_new",
                                  "dachser_duty_doc_no", "dachser_dokument"):
                    width = 18
                else:
                    width = max(10, min(20, len(hdr) + 2))
                ws.column_dimensions[col_letter].width = width

            # Zeilen fixieren (Titel + Header scrollen nicht mit)
            ws.freeze_panes = "A3"

            # Auto-Filter auf Header-Zeile
            ws.auto_filter.ref = (
                f"A2:{get_column_letter(len(DB_TARIFF_HEADERS))}2"
            )

            # ── Abgleich-Legende (falls aktiv) ────────────────────────────
            if ab_act:
                ws_leg = wb.create_sheet("Legende Abgleich")
                legend = [
                    ("Farbe", "Bedeutung"),
                    ("🟢 Grün",  "Summe TAX stimmt mit Claim-Report überein (Toleranz ±0.05 USD)"),
                    ("🔴 Rot",   "Differenz oder Summe TAX fehlt in db_Tariff"),
                    ("🟡 Gelb",  "Eintrag nur in db_Tariff – kein Match im Claim-Report"),
                    ("Weiß",    "Kein Abgleich durchgeführt / außerhalb Abgleich-Scope"),
                ]
                for r, (a, b) in enumerate(legend, start=1):
                    ws_leg.cell(row=r, column=1, value=a).font = Font(bold=(r == 1))
                    ws_leg.cell(row=r, column=2, value=b)
                ws_leg.column_dimensions["A"].width = 12
                ws_leg.column_dimensions["B"].width = 65

            wb.save(out_path)

            n_rows = len(rows)
            self._dbt_status_var.set(
                f"✅ Report gespeichert: {os.path.basename(out_path)}  "
                f"({n_rows} Zeilen + Summenzeile)"
            )
            # Datei automatisch öffnen
            try:
                os.startfile(out_path)
            except Exception:
                import subprocess
                subprocess.Popen(["start", "", out_path], shell=True)

        except Exception as exc:
            self._dbt_status_var.set(f"❌ Export-Fehler: {exc}")
            messagebox.showerror("Export-Fehler", str(exc))

    # -----------------------------------------------------------------------
    # CBP Liquidation Tab
    # -----------------------------------------------------------------------
    def _build_liq_tab(self):
        """CBP-Report > Liquidation – Excel Drag & Drop Import + Tabelle."""
        parent = self.tab_liq
        self._liq_status_var = tk.StringVar(value="Bitte Excel-Datei hier ablegen …")

        # ── Top-Toolbar ──────────────────────────────────────────────────────
        toolbar = tk.Frame(parent, bg=CARD_BG, height=54)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        tk.Button(
            toolbar, text="↻  Aktualisieren", command=self._load_liq_data,
            bg=SIDEBAR_ACT, fg="white", relief="flat", padx=14, pady=6,
            font=("Segoe UI", 10, "bold"), cursor="hand2",
        ).pack(side="left", padx=16, pady=10)

        tk.Button(
            toolbar, text="🗑  Tabelle leeren", command=self._liq_clear_table,
            bg="#D32F2F", fg="white", relief="flat", padx=14, pady=6,
            font=("Segoe UI", 10), cursor="hand2",
        ).pack(side="left", padx=(0, 16), pady=10)

        tk.Label(toolbar, textvariable=self._liq_status_var,
                 bg=CARD_BG, fg="#5A6A80", font=("Segoe UI", 10)).pack(
            side="left", padx=8)

        tk.Frame(parent, bg="#E4E8EF", height=1).pack(fill="x")

        # ── Drag & Drop Zone ─────────────────────────────────────────────────
        drop_frame = tk.Frame(parent, bg="#EAF3FB", height=90)
        drop_frame.pack(fill="x", padx=20, pady=12)
        drop_frame.pack_propagate(False)

        drop_lbl = tk.Label(
            drop_frame,
            text="💰   Excel-Datei hier ablegen  (CBP Liquidation Refund Report – Report Sheet)",
            bg="#EAF3FB", fg="#1A5B8A",
            font=("Segoe UI", 12, "bold"),
        )
        drop_lbl.place(relx=0.5, rely=0.5, anchor="center")

        try:
            from tkinterdnd2 import DND_FILES
            for w in (drop_frame, drop_lbl):
                w.drop_target_register(DND_FILES)
                w.dnd_bind("<<Drop>>", self._liq_on_drop)
        except Exception:
            pass

        # ── Suchleiste ───────────────────────────────────────────────────────
        sf = tk.Frame(parent, bg=MAIN_BG)
        sf.pack(fill="x", padx=20, pady=(0, 6))
        tk.Label(sf, text="🔍", bg=MAIN_BG, fg="#8A9EB5",
                 font=("Segoe UI", 12)).pack(side="left", padx=(0, 6))
        self._liq_search_var = tk.StringVar()
        self._liq_search_var.trace_add("write", lambda *_: self._liq_filter())
        tk.Entry(sf, textvariable=self._liq_search_var, relief="flat",
                 font=("Segoe UI", 11), bg="white", fg="#1A2742",
                 insertbackground="#1A2742").pack(side="left", fill="x",
                                                   expand=True, ipady=5)

        # ── Tabelle ──────────────────────────────────────────────────────────
        tv_frame = tk.Frame(parent, bg=MAIN_BG)
        tv_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        vsb = ttk.Scrollbar(tv_frame, orient="vertical")
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal")
        self._liq_tv = ttk.Treeview(
            tv_frame,
            columns=LIQ_COLUMNS,
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._liq_tv.yview)
        hsb.config(command=self._liq_tv.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._liq_tv.pack(side="left", fill="both", expand=True)

        col_widths = {
            "refund_id": 90, "payee_id": 110, "company_name": 160,
            "co": 60, "address": 260, "refund_date": 90,
            "refund_status": 90, "refund_secondary_status": 120,
            "refund_type": 140, "document_number": 120,
            "total_refund_amount": 110, "check_ach_trace_number": 130,
            "check_or_ach": 80, "center_id": 75, "team_number": 75,
            "port_code": 75,
        }
        for col, hdr in zip(LIQ_COLUMNS, LIQ_HEADERS):
            self._liq_tv.heading(col, text=hdr,
                                 command=lambda c=col: self._liq_sort(c))
            self._liq_tv.column(col, width=col_widths.get(col, 100),
                                minwidth=60, stretch=False, anchor="w")

        self._liq_tv.tag_configure("odd",  background="#F7F9FC")
        self._liq_tv.tag_configure("even", background="white")

        self._liq_ensure_db()
        self._load_liq_data()
        self._liq_sort_col = None
        self._liq_sort_rev = False

    # -----------------------------------------------------------------------
    def _build_abgleich_manual_tab(self):
        """Plausibilität > Abgleich-Manual – Excel Drag & Drop Import + Tabelle."""
        parent = self.tab_abgleich_manual
        self._plausi_status_var = tk.StringVar(value="Bitte Excel-Datei hier ablegen …")

        # ── Top-Toolbar ──────────────────────────────────────────────────────
        toolbar = tk.Frame(parent, bg=CARD_BG, height=54)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        tk.Button(
            toolbar, text="↻  Aktualisieren", command=self._load_plausi_data,
            bg=SIDEBAR_ACT, fg="white", relief="flat", padx=14, pady=6,
            font=("Segoe UI", 10, "bold"), cursor="hand2",
        ).pack(side="left", padx=16, pady=10)

        tk.Button(
            toolbar, text="🗑  Tabelle leeren", command=self._plausi_clear_table,
            bg="#D32F2F", fg="white", relief="flat", padx=14, pady=6,
            font=("Segoe UI", 10), cursor="hand2",
        ).pack(side="left", padx=(0, 16), pady=10)

        tk.Label(toolbar, textvariable=self._plausi_status_var,
                 bg=CARD_BG, fg="#5A6A80", font=("Segoe UI", 10)).pack(
            side="left", padx=8)

        tk.Frame(parent, bg="#E4E8EF", height=1).pack(fill="x")

        # ── Drag & Drop Zone ─────────────────────────────────────────────────
        drop_frame = tk.Frame(parent, bg="#EAF3FB", height=90)
        drop_frame.pack(fill="x", padx=20, pady=12)
        drop_frame.pack_propagate(False)

        drop_lbl = tk.Label(
            drop_frame,
            text="🔍   Excel-Datei hier ablegen  (Plausi_ZollBetrag – Sheet: Plausi)",
            bg="#EAF3FB", fg="#1A5B8A",
            font=("Segoe UI", 12, "bold"),
        )
        drop_lbl.place(relx=0.5, rely=0.5, anchor="center")

        try:
            from tkinterdnd2 import DND_FILES
            for w in (drop_frame, drop_lbl):
                w.drop_target_register(DND_FILES)
                w.dnd_bind("<<Drop>>", self._plausi_on_drop)
        except Exception:
            pass

        # ── Suchleiste ───────────────────────────────────────────────────────
        sf = tk.Frame(parent, bg=MAIN_BG)
        sf.pack(fill="x", padx=20, pady=(0, 6))
        tk.Label(sf, text="🔍", bg=MAIN_BG, fg="#8A9EB5",
                 font=("Segoe UI", 12)).pack(side="left", padx=(0, 6))
        self._plausi_search_var = tk.StringVar()
        self._plausi_search_var.trace_add("write", lambda *_: self._plausi_filter())
        tk.Entry(sf, textvariable=self._plausi_search_var, relief="flat",
                 font=("Segoe UI", 11), bg="white", fg="#1A2742",
                 insertbackground="#1A2742").pack(side="left", fill="x",
                                                   expand=True, ipady=5)

        # ── Tabelle ──────────────────────────────────────────────────────────
        tv_frame = tk.Frame(parent, bg=MAIN_BG)
        tv_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        vsb = ttk.Scrollbar(tv_frame, orient="vertical")
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal")
        self._plausi_tv = ttk.Treeview(
            tv_frame,
            columns=PLAUSI_COLUMNS,
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self._plausi_tv.yview)
        hsb.config(command=self._plausi_tv.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._plausi_tv.pack(side="left", fill="both", expand=True)

        col_widths = {
            "year_month": 80, "dachser_duty_doc_no": 120, "entry_no": 120,
            "tax_volume_total_doc": 100, "tax_volume_total_doc_currency": 70,
            "invoices_base_for_tax": 100, "invoices_currency_base_for_tax": 70,
            "merchandise_processing_fee": 80, "harbour_fee": 75,
            "hts_3926_90_9989": 90, "hts_8483_90_1050": 90,
            "hts_8483_90_8080": 90, "hts_8708_99_6805": 90,
            "hts_8708_99_8180": 90, "hts_9903_81_90": 90,
            "hts_9903_01_10": 80, "hts_9903_01_24": 90, "hts_9903_88_01": 90,
            "summe_tax": 90, "vendor": 80, "vendor_name": 180,
            "vendor_origin_o": 70, "vendor_origin_e": 70,
            "invoice_no_from_vendor": 130, "po_order_no": 110,
            "catensys_mat_no_sap": 130, "catensys_mat_desc_sap": 200,
            "quantity": 70, "rate": 70, "invoice_doc_value": 100,
            "doc_currency": 70, "local_value_sap": 100, "local_currency": 70,
            "reference_doc": 110, "material_doc": 100, "batch": 90,
            "air_see_truck": 75, "customer_partno": 110,
            "sap_customer_no": 100, "sap_customer_descr": 160,
            "project_contingent": 110,
        }
        for col, hdr in zip(PLAUSI_COLUMNS, PLAUSI_HEADERS):
            self._plausi_tv.heading(col, text=hdr,
                                    command=lambda c=col: self._plausi_sort(c))
            self._plausi_tv.column(col, width=col_widths.get(col, 90),
                                   minwidth=50, stretch=False, anchor="w")

        self._plausi_tv.tag_configure("odd",  background="#F7F9FC")
        self._plausi_tv.tag_configure("even", background="white")

        self._plausi_ensure_db()
        self._load_plausi_data()
        self._plausi_sort_col = None
        self._plausi_sort_rev = False

    def _plausi_ensure_db(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cols_def = ", ".join(f"{c} TEXT" for c in PLAUSI_COLUMNS)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS plausi_abgleich_manual (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_date TEXT DEFAULT (date('now')),
                {cols_def}
            )
        """)
        conn.commit()
        conn.close()

    def _plausi_on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw
        if not os.path.isfile(path):
            self._plausi_status_var.set(f"Datei nicht gefunden: {path}")
            return
        if not path.lower().endswith((".xlsx", ".xls")):
            self._plausi_status_var.set("Nur Excel-Dateien (.xlsx / .xls) werden unterstützt.")
            return
        self._plausi_status_var.set(f"Importiere: {os.path.basename(path)} …")
        threading.Thread(target=self._plausi_import_excel, args=(path,),
                         daemon=True).start()

    def _plausi_import_excel(self, path):
        try:
            import openpyxl
            from datetime import datetime as _dt
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_name = None
            for n in wb.sheetnames:
                if n.strip().lower() == "plausi":
                    sheet_name = n
                    break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Header Zeile 5 (index 4), Daten ab Zeile 6 (index 5)
            data_rows = rows[5:]
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            inserted = 0
            source = os.path.basename(path)
            for row in data_rows:
                vals = list(row[:41])
                while len(vals) < 41:
                    vals.append(None)
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                norm = []
                for v in vals:
                    if v is None:
                        norm.append(None)
                    elif isinstance(v, _dt):
                        norm.append(v.strftime("%Y-%m-%d"))
                    elif isinstance(v, float):
                        norm.append(f"{v:.6g}")
                    else:
                        s = str(v).strip()
                        norm.append(s if s else None)
                cols_sql = ", ".join(PLAUSI_COLUMNS)
                ph = ", ".join(["?"] * len(PLAUSI_COLUMNS))
                cur.execute(
                    f"INSERT INTO plausi_abgleich_manual ({cols_sql}) VALUES ({ph})",
                    norm,
                )
                inserted += 1
            conn.commit()
            conn.close()
            self.root.after(0, lambda n=inserted, s=source: (
                self._plausi_status_var.set(f"✓ {n} Zeilen importiert aus '{s}'."),
                self._load_plausi_data(),
            ))
        except ImportError:
            self.root.after(0, lambda: self._plausi_status_var.set(
                "openpyxl nicht installiert – bitte: pip install openpyxl"))
        except Exception as exc:
            self.root.after(0, lambda e=exc: self._plausi_status_var.set(
                f"Import-Fehler: {e}"))

    def _load_plausi_data(self, *_):
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute(
                f"SELECT {', '.join(PLAUSI_COLUMNS)} FROM plausi_abgleich_manual ORDER BY id DESC"
            )
            rows = cur.fetchall()
            conn.close()
        except Exception:
            rows = []
        self._plausi_all_rows = rows
        self._plausi_filter()

    def _plausi_filter(self, *_):
        term = self._plausi_search_var.get().lower()
        tv = self._plausi_tv
        for item in tv.get_children():
            tv.delete(item)
        count = 0
        for row in self._plausi_all_rows:
            if term and not any(term in str(v).lower() for v in row if v):
                continue
            tag = "even" if count % 2 == 0 else "odd"
            tv.insert("", "end", values=row, tags=(tag,))
            count += 1
        total = len(self._plausi_all_rows)
        if term:
            self._plausi_status_var.set(f"{count} von {total} Einträgen angezeigt.")
        else:
            self._plausi_status_var.set(f"{total} Einträge in der Datenbank.")

    def _plausi_sort(self, col):
        if self._plausi_sort_col == col:
            self._plausi_sort_rev = not self._plausi_sort_rev
        else:
            self._plausi_sort_col = col
            self._plausi_sort_rev = False
        idx = PLAUSI_COLUMNS.index(col)
        self._plausi_all_rows.sort(
            key=lambda r: (r[idx] is None, str(r[idx] or "").lower()),
            reverse=self._plausi_sort_rev,
        )
        self._plausi_filter()

    def _plausi_clear_table(self):
        import tkinter.messagebox as mb
        if not mb.askyesno(
            "Tabelle leeren",
            "Alle Abgleich-Manual-Daten löschen?\nDieser Vorgang kann nicht rückgängig gemacht werden.",
        ):
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM plausi_abgleich_manual")
            conn.commit()
            conn.close()
        except Exception as exc:
            self._plausi_status_var.set(f"Fehler: {exc}")
            return
        self._load_plausi_data()

    # -----------------------------------------------------------------------
    def _build_abgleich_db_tab(self):
        parent = self.tab_abgleich_db
        self._adb_status_var = tk.StringVar(
            value="Abgleich noch nicht ausgeführt – Schaltfläche 'Abgleich ausführen' klicken.")

        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = tk.Frame(parent, bg=CARD_BG, height=54)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        tk.Button(
            toolbar, text="⚖  Abgleich ausführen",
            command=self._run_abgleich_db,
            bg=SIDEBAR_ACT, fg="white", relief="flat", padx=16, pady=6,
            font=("Segoe UI", 10, "bold"), cursor="hand2",
        ).pack(side="left", padx=16, pady=10)

        tk.Label(toolbar, textvariable=self._adb_status_var,
                 bg=CARD_BG, fg="#5A6A80",
                 font=("Segoe UI", 10)).pack(side="left", padx=12)

        tk.Frame(parent, bg="#E4E8EF", height=1).pack(fill="x")

        # ── Legende ──────────────────────────────────────────────────────────
        leg = tk.Frame(parent, bg=MAIN_BG)
        leg.pack(fill="x", padx=20, pady=(8, 4))
        for txt, bg, fg in [
            ("✅ Übereinstimmung",          "#DCFCE7", "#065F46"),
            ("⚠  Differenz (unvollständig)", "#FEF3C7", "#92400E"),
            ("❌ Fehlt in DB",               "#FEE2E2", "#991B1B"),
        ]:
            tk.Label(leg, text=txt, bg=bg, fg=fg,
                     font=("Segoe UI", 9, "bold"),
                     padx=10, pady=3, relief="flat").pack(side="left", padx=(0, 8))

        tk.Frame(parent, bg="#E4E8EF", height=1).pack(fill="x", padx=20)

        # ── Spalten-Header ───────────────────────────────────────────────────
        HDR_BG = "#1A2742"
        hdr = tk.Frame(parent, bg=HDR_BG)
        hdr.pack(fill="x", padx=20, pady=(4, 0))
        for txt, w in [
            ("#",              36),
            ("Entry-No",      145),
            ("Jahr/Monat",     90),
            ("Soll TAX (Manual)", 130),
            ("Ist TAX (DB)",  130),
            ("Differenz",     110),
            ("Pos. DB",        70),
            ("Status",        170),
            ("Aktion",        130),
        ]:
            tk.Label(hdr, text=txt, bg=HDR_BG, fg="white",
                     font=("Segoe UI", 9, "bold"),
                     width=0, anchor="w").pack(
                side="left", padx=6, pady=5, ipadx=2)

        # ── Scrollbarer Tabellen-Körper ───────────────────────────────────────
        outer = tk.Frame(parent, bg=MAIN_BG,
                         highlightthickness=1, highlightbackground="#D0D8E4")
        outer.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        canvas = tk.Canvas(outer, bg=MAIN_BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._adb_frame = tk.Frame(canvas, bg=MAIN_BG)
        _win = canvas.create_window((0, 0), window=self._adb_frame, anchor="nw")

        def _cfg(e):   canvas.configure(scrollregion=canvas.bbox("all"))
        def _ccfg(e):  canvas.itemconfig(_win, width=e.width)
        self._adb_frame.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _ccfg)
        canvas.bind("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self._adb_empty_lbl = tk.Label(
            self._adb_frame,
            text="Noch kein Abgleich ausgeführt.",
            bg=MAIN_BG, fg="#8A9EB5", font=("Segoe UI", 10))

    # ── Abgleich-Logik ───────────────────────────────────────────────────────
    def _run_abgleich_db(self):
        """Vergleicht plausi_abgleich_manual mit entry_lines und zeigt Differenzen."""
        self._adb_status_var.set("Abgleich wird berechnet …")
        self.root.after(50, self._do_abgleich_db)

    def _do_abgleich_db(self):
        try:
            conn = sqlite3.connect(DB_PATH)

            # ── SOLL: Erwarteter TAX-Betrag pro Entry No aus dem Manual-Excel ──
            # Wir nehmen die Summary-Zeilen (tax_volume_total_doc IS NOT NULL)
            # und addieren deren Wert pro Entry No.
            soll_rows = conn.execute("""
                SELECT entry_no,
                       MAX(year_month) as year_month,
                       SUM(CAST(COALESCE(tax_volume_total_doc, 0) AS REAL)) as soll
                FROM plausi_abgleich_manual
                WHERE entry_no IS NOT NULL
                  AND tax_volume_total_doc IS NOT NULL
                  AND CAST(tax_volume_total_doc AS REAL) > 0
                GROUP BY entry_no
                ORDER BY entry_no
            """).fetchall()

            if not soll_rows:
                # Fallback: summe_tax aus Detailzeilen
                soll_rows = conn.execute("""
                    SELECT entry_no,
                           MAX(year_month) as year_month,
                           SUM(CAST(COALESCE(summe_tax, 0) AS REAL)) as soll
                    FROM plausi_abgleich_manual
                    WHERE entry_no IS NOT NULL
                      AND summe_tax IS NOT NULL
                    GROUP BY entry_no
                    ORDER BY entry_no
                """).fetchall()

            # ── IST: Tatsächliche Zollbeträge in entry_lines (inkl. MPF) ──────
            ist_rows = conn.execute("""
                SELECT e.filer_code_entry_no,
                       COUNT(DISTINCT el.id)                                   AS line_count,
                       SUM(CAST(COALESCE(el.duty_amount, 0) AS REAL))         AS sum_duty,
                       SUM(CAST(COALESCE(el.mpf_amount,  0) AS REAL))         AS sum_mpf
                FROM entry_lines el
                JOIN entries e ON el.entry_id = e.id
                WHERE e.filer_code_entry_no IS NOT NULL
                GROUP BY e.filer_code_entry_no
            """).fetchall()
            ist_map = {r[0]: (r[1], r[2], r[3]) for r in ist_rows}

            # ── Archiv-Pfade aus processed_files ────────────────────────────
            arch_rows = conn.execute("""
                SELECT e.filer_code_entry_no,
                       pf.archived_path,
                       pf.file_name
                FROM entries e
                JOIN processed_files pf ON pf.file_name = e.source_file
                WHERE pf.archived_path IS NOT NULL
                GROUP BY e.filer_code_entry_no
                ORDER BY pf.id DESC
            """).fetchall()
            arch_map = {}
            for eno, apath, fname in arch_rows:
                if eno not in arch_map:
                    arch_map[eno] = (apath, fname)

            conn.close()

            # ── Alle Entry-Nos (Union aus Soll + Ist) ───────────────────────
            soll_map = {r[0]: (r[1], r[2]) for r in soll_rows}
            all_entries = sorted(set(list(soll_map.keys()) + list(ist_map.keys())))

        except Exception as exc:
            self._adb_status_var.set(f"Fehler beim Abgleich: {exc}")
            return

        # ── Tabelle aufbauen ─────────────────────────────────────────────────
        frame = self._adb_frame
        for w in frame.winfo_children():
            w.destroy()

        if not all_entries:
            tk.Label(frame,
                     text="Keine Daten – bitte zuerst Abgleich-Manual importieren und PDFs einlesen.",
                     bg=MAIN_BG, fg="#8A9EB5",
                     font=("Segoe UI", 10)).pack(pady=20)
            self._adb_status_var.set("Keine Daten vorhanden.")
            return

        ok_count  = 0
        diff_count = 0
        miss_count = 0
        TOLERANCE  = 0.05   # < 5 Cent gilt als Übereinstimmung

        for idx, entry_no in enumerate(all_entries):
            year_month, soll = soll_map.get(entry_no, ("–", None))
            ist_data         = ist_map.get(entry_no)
            arch_data        = arch_map.get(entry_no)

            if ist_data:
                line_cnt, sum_duty, sum_mpf = ist_data
                ist_total = sum_duty + sum_mpf
            else:
                line_cnt  = 0
                ist_total = None

            # Differenz & Status
            if soll is None and ist_total is None:
                status_txt = "⚠  nur in DB"
                row_bg, row_fg = "#FEF3C7", "#92400E"
                diff_str = "–"
                diff_count += 1
            elif ist_total is None or line_cnt == 0:
                status_txt = "❌ Fehlt in DB"
                row_bg, row_fg = "#FEE2E2", "#991B1B"
                diff_str = f"{soll:,.2f}" if soll else "–"
                miss_count += 1
            elif soll is None:
                status_txt = "⚠  kein Manual-Wert"
                row_bg, row_fg = "#FEF3C7", "#92400E"
                diff_str = "–"
                diff_count += 1
            else:
                diff = soll - ist_total
                diff_str = f"{diff:+,.2f}"
                if abs(diff) <= TOLERANCE:
                    status_txt = "✅ Übereinstimmung"
                    row_bg, row_fg = "#DCFCE7", "#065F46"
                    ok_count += 1
                else:
                    pct = abs(diff / soll * 100) if soll else 0
                    status_txt = f"⚠  Diff {pct:.1f}%"
                    row_bg, row_fg = "#FEF3C7", "#92400E"
                    diff_count += 1

            # Zeile rendern
            row_f = tk.Frame(frame, bg=row_bg,
                             highlightthickness=1,
                             highlightbackground="#D4D8E0")
            row_f.pack(fill="x", pady=(0, 1))

            tk.Label(row_f, text=str(idx+1), bg=row_bg, fg="#8A9EB5",
                     font=("Segoe UI", 9), width=3,
                     anchor="center").pack(side="left", padx=(6,2), pady=6)

            tk.Label(row_f, text=entry_no, bg=row_bg, fg="#1A2742",
                     font=("Segoe UI", 9, "bold"), width=18,
                     anchor="w").pack(side="left", padx=(2,4))

            tk.Label(row_f, text=year_month or "–", bg=row_bg, fg="#5A6A80",
                     font=("Segoe UI", 9), width=9,
                     anchor="w").pack(side="left", padx=(0,4))

            # Soll
            tk.Label(row_f,
                     text=f"{soll:,.2f}" if soll is not None else "–",
                     bg=row_bg, fg="#1A2742",
                     font=("Segoe UI", 9), width=14,
                     anchor="e").pack(side="left", padx=(0,6))

            # Ist
            tk.Label(row_f,
                     text=f"{ist_total:,.2f}" if ist_total is not None else "–",
                     bg=row_bg, fg="#1A2742",
                     font=("Segoe UI", 9), width=14,
                     anchor="e").pack(side="left", padx=(0,6))

            # Differenz (fett wenn Abweichung)
            diff_font = ("Segoe UI", 9, "bold") if "+" in diff_str or ("-" in diff_str and diff_str != "–") else ("Segoe UI", 9)
            diff_color = "#DC2626" if "+" in diff_str or ("-" in diff_str and diff_str != "–") else row_fg
            tk.Label(row_f, text=diff_str, bg=row_bg, fg=diff_color,
                     font=diff_font, width=12,
                     anchor="e").pack(side="left", padx=(0,6))

            # Positionen DB
            pos_fg = "#065F46" if line_cnt > 0 else "#DC2626"
            tk.Label(row_f, text=str(line_cnt), bg=row_bg, fg=pos_fg,
                     font=("Segoe UI", 9, "bold"), width=6,
                     anchor="center").pack(side="left", padx=(0,6))

            # Status-Chip
            tk.Label(row_f, text=status_txt, bg=row_bg, fg=row_fg,
                     font=("Segoe UI", 8, "bold"),
                     padx=6, pady=2).pack(side="left", padx=(0,8))

            # Aktion
            needs_reread = status_txt.startswith(("⚠", "❌"))
            if needs_reread and arch_data:
                apath, afname = arch_data
                if os.path.isfile(apath):
                    tk.Button(
                        row_f,
                        text="🔄 PDF erneut lesen",
                        command=lambda p=apath, f=afname, e=entry_no: \
                            self._abgleich_reread_pdf(p, f, e),
                        bg="#1D4ED8", fg="white", relief="flat",
                        padx=8, pady=2,
                        font=("Segoe UI", 8, "bold"),
                        cursor="hand2",
                    ).pack(side="left", padx=(0,8), pady=4)
                else:
                    tk.Label(row_f, text="⚠ Archiv fehlt",
                             bg=row_bg, fg="#DC2626",
                             font=("Segoe UI", 8)).pack(side="left", padx=(0,8))
            elif needs_reread:
                tk.Label(row_f, text="kein Archiv",
                         bg=row_bg, fg="#9CA3AF",
                         font=("Segoe UI", 8)).pack(side="left", padx=(0,8))

        total = len(all_entries)
        self._adb_status_var.set(
            f"{total} Entry-Nos  |  ✅ {ok_count} OK  "
            f"⚠ {diff_count} Differenz  ❌ {miss_count} fehlt in DB"
        )

    def _abgleich_reread_pdf(self, arch_path, fname, entry_no):
        """Löscht entry_lines für diese Entry-No und liest PDF erneut."""
        if not os.path.isfile(arch_path):
            messagebox.showerror("Datei fehlt",
                f"Archivierte PDF nicht gefunden:\n{arch_path}")
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            # Entry aus DB entfernen damit sauber neu importiert wird
            cur = conn.cursor()
            cur.execute(
                "SELECT id FROM entries WHERE filer_code_entry_no = ?",
                (entry_no,))
            eids = [r[0] for r in cur.fetchall()]
            if eids:
                ph = ",".join("?" * len(eids))
                conn.execute(f"DELETE FROM entry_lines WHERE entry_id IN ({ph})", eids)
                conn.execute(
                    "DELETE FROM entries WHERE filer_code_entry_no = ?",
                    (entry_no,))
            conn.commit()
            conn.close()
        except Exception as exc:
            messagebox.showerror("DB-Fehler", str(exc))
            return

        self._adb_status_var.set(f"Lese erneut: {fname} …")
        # Nach dem Batch-Import Abgleich automatisch neu ausführen
        self._abgleich_pending_refresh = True
        self._process_pdf_batch([arch_path])

    # -----------------------------------------------------------------------
    def _liq_ensure_db(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cols_def = ", ".join(f"{c} TEXT" for c in LIQ_COLUMNS)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS cbp_liquidation (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_date TEXT DEFAULT (date('now')),
                {cols_def}
            )
        """)
        conn.commit()
        conn.close()

    def _liq_on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw
        if not os.path.isfile(path):
            self._liq_status_var.set(f"Datei nicht gefunden: {path}")
            return
        if not path.lower().endswith((".xlsx", ".xls")):
            self._liq_status_var.set("Nur Excel-Dateien (.xlsx / .xls) werden unterstützt.")
            return
        self._liq_status_var.set(f"Importiere: {os.path.basename(path)} …")
        threading.Thread(target=self._liq_import_excel, args=(path,),
                         daemon=True).start()

    def _liq_import_excel(self, path):
        try:
            import openpyxl
            from datetime import datetime as _dt
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            # Sheet "Report" suchen (case-insensitive)
            sheet_name = None
            for n in wb.sheetnames:
                if n.strip().lower() == "report":
                    sheet_name = n
                    break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Header in Zeile 6 (index 5), Daten ab Zeile 7 (index 6)
            data_rows = rows[6:]

            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            inserted = 0
            source = os.path.basename(path)
            for row in data_rows:
                vals = list(row[:16])
                while len(vals) < 16:
                    vals.append(None)
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                norm = []
                for v in vals:
                    if v is None:
                        norm.append(None)
                    elif isinstance(v, _dt):
                        norm.append(v.strftime("%Y-%m-%d"))
                    else:
                        s = str(v).strip()
                        norm.append(s if s else None)
                cols_sql = ", ".join(LIQ_COLUMNS)
                ph = ", ".join(["?"] * len(LIQ_COLUMNS))
                cur.execute(
                    f"INSERT INTO cbp_liquidation ({cols_sql}) VALUES ({ph})",
                    norm,
                )
                inserted += 1
            conn.commit()
            conn.close()
            self.root.after(0, lambda n=inserted, s=source: (
                self._liq_status_var.set(f"✓ {n} Zeilen importiert aus '{s}'."),
                self._load_liq_data(),
            ))
        except ImportError:
            self.root.after(0, lambda: self._liq_status_var.set(
                "openpyxl nicht installiert – bitte: pip install openpyxl"))
        except Exception as exc:
            self.root.after(0, lambda e=exc: self._liq_status_var.set(
                f"Import-Fehler: {e}"))

    def _load_liq_data(self, *_):
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            cur.execute(
                f"SELECT {', '.join(LIQ_COLUMNS)} FROM cbp_liquidation ORDER BY id DESC"
            )
            rows = cur.fetchall()
            conn.close()
        except Exception:
            rows = []
        self._liq_all_rows = rows
        self._liq_filter()

    def _liq_filter(self, *_):
        term = self._liq_search_var.get().lower()
        tv = self._liq_tv
        for item in tv.get_children():
            tv.delete(item)
        count = 0
        for row in self._liq_all_rows:
            if term and not any(term in str(v).lower() for v in row if v):
                continue
            tag = "even" if count % 2 == 0 else "odd"
            tv.insert("", "end", values=row, tags=(tag,))
            count += 1
        total = len(self._liq_all_rows)
        if term:
            self._liq_status_var.set(f"{count} von {total} Einträgen angezeigt.")
        else:
            self._liq_status_var.set(f"{total} Einträge in der Datenbank.")

    def _liq_sort(self, col):
        if self._liq_sort_col == col:
            self._liq_sort_rev = not self._liq_sort_rev
        else:
            self._liq_sort_col = col
            self._liq_sort_rev = False
        idx = LIQ_COLUMNS.index(col)
        self._liq_all_rows.sort(
            key=lambda r: (r[idx] is None, str(r[idx] or "").lower()),
            reverse=self._liq_sort_rev,
        )
        self._liq_filter()

    def _liq_clear_table(self):
        import tkinter.messagebox as mb
        if not mb.askyesno(
            "Tabelle leeren",
            "Alle Liquidation-Daten aus der Datenbank löschen?\nDieser Vorgang kann nicht rückgängig gemacht werden.",
        ):
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM cbp_liquidation")
            conn.commit()
            conn.close()
        except Exception as exc:
            import tkinter.messagebox as _mb2
            _mb2.showerror("Fehler", str(exc), parent=self)
            return
        self._liq_load()
        self._liq_status_var.set("Tabelle geleert.")


if __name__ == "__main__":
    import traceback as _tb, os as _os
    _LOG = _os.path.join(_os.path.dirname(_os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__)), "startup_error.log")
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except Exception:
        import tkinter as tk
        root = tk.Tk()
    try:
        app = App(root)
        root.mainloop()
    except Exception as _msg:
        _err = _tb.format_exc()
        with open(_LOG, "w", encoding="utf-8") as _lf:
            _lf.write(_err)
        try:
            import tkinter.messagebox as _mb2
            _mb2.showerror("Startup-Fehler",
                           f"{str(_msg)[:1500]}\n\nDetails: {_LOG}")
        except Exception:
            pass
